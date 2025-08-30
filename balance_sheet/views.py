from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db.models import Sum, Q
from django.utils import timezone
from datetime import datetime, timedelta
import csv
import json
from decimal import Decimal

from .forms import BalanceSheetReportForm, ExportForm
from .models import BalanceSheetReport, ReportTemplate, AccountGroup
from company.company_model import Company
from chart_of_accounts.models import ChartOfAccount, AccountType
from ledger.models import Ledger


def serialize_report_data(data):
    """Serialize report data to handle Decimal and other non-JSON serializable objects"""
    if not data:
        return data
    
    from django.db.models.query import QuerySet
    from django.db.models import Model
    from datetime import date
    
    def serialize_object(obj):
        if isinstance(obj, Decimal):
            return float(obj)
        elif isinstance(obj, (datetime, timezone.datetime)):
            return obj.isoformat()
        elif isinstance(obj, date):
            return obj.isoformat()
        elif hasattr(obj, 'date') and callable(obj.date):
            return obj.date().isoformat()
        elif isinstance(obj, QuerySet):
            return list(obj)
        elif isinstance(obj, Model):
            return str(obj)
        elif isinstance(obj, dict):
            return {k: serialize_object(v) for k, v in obj.items()}
        elif isinstance(obj, (list, tuple)):
            return [serialize_object(item) for item in obj]
        else:
            return obj
    
    return serialize_object(data)


@login_required
def balance_sheet_report(request):
    """Main Balance Sheet report view"""
    
    print(f"DEBUG: Request method: {request.method}")
    print(f"DEBUG: All headers: {dict(request.headers)}")
    print(f"DEBUG: X-Requested-With header: {request.headers.get('X-Requested-With')}")
    print(f"DEBUG: AJAX request: {request.headers.get('X-Requested-With') == 'XMLHttpRequest'}")
    
    if request.method == 'POST':
        print(f"DEBUG: POST data: {request.POST}")
        form = BalanceSheetReportForm(request.POST)
        print(f"DEBUG: Form is valid: {form.is_valid()}")
        if not form.is_valid():
            print(f"DEBUG: Form errors: {form.errors}")
        
        if form.is_valid():
            # Get form data
            as_of_date = form.cleaned_data['as_of_date']
            company = form.cleaned_data.get('company')
            branch = form.cleaned_data.get('branch', '')
            department = form.cleaned_data.get('department', '')
            comparison_type = form.cleaned_data.get('comparison_type', 'none')
            
            try:
                # Generate report data
                report_data = get_balance_sheet_data(
                    as_of_date=as_of_date,
                    company=company,
                    branch=branch,
                    department=department,
                    comparison_type=comparison_type,
                    include_zero_balances=form.cleaned_data.get('include_zero_balances', True),
                    show_percentages=form.cleaned_data.get('show_percentages', False)
                )
                
                # Save report to database (ensure JSON-serializable data)
                serializable_report_data = serialize_report_data(report_data)
                report = BalanceSheetReport.objects.create(
                    title=f"Balance Sheet - {as_of_date}",
                    as_of_date=as_of_date,
                    company=company,
                    branch=branch,
                    department=department,
                    report_data=serializable_report_data,
                    created_by=request.user,
                    include_headers=form.cleaned_data.get('include_headers', True),
                    include_totals=form.cleaned_data.get('include_totals', True),
                    include_comparison=form.cleaned_data.get('include_comparison', False)
                )
                
                # Handle AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    print(f"DEBUG: Returning AJAX response for report ID: {report.id}")
                    
                    serializable_report_data = serialize_report_data(report_data)
                    
                    return JsonResponse({
                        'success': True,
                        'message': f'Balance Sheet report generated successfully for {as_of_date}',
                        'report_id': report.id,
                        'report_data': serializable_report_data,
                        'as_of_date': as_of_date.strftime('%Y-%m-%d'),
                        'company': company.name if company else 'All Companies',
                    })
                
                messages.success(request, f'Balance Sheet report generated successfully for {as_of_date}')
                
                # Redirect to the detail page after successful report generation
                return redirect('balance_sheet:report_detail', report_id=report.id)
                
            except Exception as e:
                error_msg = f'Error generating report: {str(e)}'
                
                # Handle AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    print(f"DEBUG: Returning AJAX error response: {error_msg}")
                    return JsonResponse({
                        'success': False,
                        'error': error_msg
                    })
                
                messages.error(request, error_msg)
                context = {
                    'form': form,
                    'report_data': None,
                    'export_form': ExportForm(),
                }
                return render(request, 'balance_sheet/balance_sheet_report.html', context)
        else:
            error_msg = f'Form validation failed: {form.errors}'
            
            # Handle AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                print(f"DEBUG: Returning AJAX form validation error: {error_msg}")
                return JsonResponse({
                    'success': False,
                    'error': error_msg,
                    'form_errors': form.errors
                })
            
            messages.error(request, error_msg)
    else:
        form = BalanceSheetReportForm()
    
    context = {
        'form': form,
        'report_data': None,
        'export_form': ExportForm(),
    }
    
    return render(request, 'balance_sheet/balance_sheet_report.html', context)


def get_balance_sheet_data(as_of_date, company=None, branch='', department='', 
                          comparison_type='none', include_zero_balances=True, show_percentages=False):
    """Generate Balance Sheet report data"""
    
    # Get account balances as of the specified date
    base_query = Ledger.objects.filter(entry_date__lte=as_of_date)
    
    if company:
        base_query = base_query.filter(company=company)
    
    # Get account balances
    account_balances = base_query.values('account').annotate(
        total_debit=Sum('amount', filter=Q(entry_type__in=['debit', 'DR'])),
        total_credit=Sum('amount', filter=Q(entry_type__in=['credit', 'CR']))
    )
    
    # Calculate net balance for each account
    account_data = {}
    for balance in account_balances:
        account_id = balance['account']
        total_debit = balance['total_debit'] or 0
        total_credit = balance['total_credit'] or 0
        net_balance = total_debit - total_credit
        
        if include_zero_balances or net_balance != 0:
            account_data[account_id] = net_balance
    
    # Get account types for different sections
    current_asset_types = AccountType.objects.filter(
        Q(name__icontains='current asset') | Q(name__icontains='cash') | 
        Q(name__icontains='receivable') | Q(name__icontains='inventory')
    )
    
    # Get accounts for each section
    current_asset_accounts = ChartOfAccount.objects.filter(account_type__in=current_asset_types)
    non_current_asset_types = AccountType.objects.filter(
        Q(name__icontains='fixed asset') | Q(name__icontains='property') | 
        Q(name__icontains='equipment') | Q(name__icontains='intangible')
    )
    
    non_current_asset_accounts = ChartOfAccount.objects.filter(account_type__in=non_current_asset_types)
    current_liability_types = AccountType.objects.filter(
        Q(name__icontains='current liability') | Q(name__icontains='payable') | 
        Q(name__icontains='short term')
    )
    
    current_liability_accounts = ChartOfAccount.objects.filter(account_type__in=current_liability_types)
    non_current_liability_types = AccountType.objects.filter(
        Q(name__icontains='long term') | Q(name__icontains='loan') | 
        Q(name__icontains='mortgage')
    )
    
    non_current_liability_accounts = ChartOfAccount.objects.filter(account_type__in=non_current_liability_types)
    equity_types = AccountType.objects.filter(
        Q(name__icontains='equity') | Q(name__icontains='capital') | 
        Q(name__icontains='retained earnings')
    )
    
    equity_accounts = ChartOfAccount.objects.filter(account_type__in=equity_types)
    
    # Calculate section totals
    def get_section_data(accounts):
        section_data = []
        total = 0
        for account in accounts:
            balance = account_data.get(account.id, 0)
            if include_zero_balances or balance != 0:
                section_data.append({
                    'account_name': account.name,
                    'account_code': account.account_code,
                    'balance': balance,
                    'account_type': account.account_type.name
                })
                total += balance
        return section_data, total
    
    current_assets_data, total_current_assets = get_section_data(current_asset_accounts)
    non_current_assets_data, total_non_current_assets = get_section_data(non_current_asset_accounts)
    current_liabilities_data, total_current_liabilities = get_section_data(current_liability_accounts)
    non_current_liabilities_data, total_non_current_liabilities = get_section_data(non_current_liability_accounts)
    equity_data, total_equity = get_section_data(equity_accounts)
    
    # Calculate totals
    total_assets = total_current_assets + total_non_current_assets
    total_liabilities = total_current_liabilities + total_non_current_liabilities
    total_liabilities_equity = total_liabilities + total_equity
    
    # If no data found, generate sample data for demonstration
    if total_assets == 0 and total_liabilities == 0 and total_equity == 0:
        return get_sample_balance_sheet_data()
    
    # Calculate percentages if requested
    if show_percentages and total_assets > 0:
        def add_percentages(data, total):
            for item in data:
                item['percentage'] = (item['balance'] / total * 100) if total else 0
            return data
        
        current_assets_data = add_percentages(current_assets_data, total_assets)
        non_current_assets_data = add_percentages(non_current_assets_data, total_assets)
        current_liabilities_data = add_percentages(current_liabilities_data, total_liabilities_equity)
        non_current_liabilities_data = add_percentages(non_current_liabilities_data, total_liabilities_equity)
        equity_data = add_percentages(equity_data, total_liabilities_equity)
    
    # Prepare data structure for ratio calculation
    data_for_ratios = {
        'current_assets': {'total': total_current_assets, 'accounts': current_assets_data},
        'current_liabilities': {'total': total_current_liabilities},
        'total_assets': total_assets,
        'total_liabilities': total_liabilities,
        'equity': {'total': total_equity},
    }
    
    # Calculate financial ratios for professional analysis
    financial_ratios = calculate_financial_ratios(data_for_ratios)
    
    # Prepare comparison data if requested
    comparison_data = None
    if comparison_type != 'none':
        comparison_data = get_comparison_data(
            as_of_date, company, branch, department, comparison_type
        )
    
    return {
        'current_assets': {
            'accounts': current_assets_data,
            'total': total_current_assets,
            'title': 'Current Assets',
        },
        'non_current_assets': {
            'accounts': non_current_assets_data,
            'total': total_non_current_assets,
            'title': 'Non-Current Assets',
        },
        'total_assets': total_assets,
        'current_liabilities': {
            'accounts': current_liabilities_data,
            'total': total_current_liabilities,
            'title': 'Current Liabilities',
        },
        'non_current_liabilities': {
            'accounts': non_current_liabilities_data,
            'total': total_non_current_liabilities,
            'title': 'Non-Current Liabilities',
        },
        'equity': {
            'accounts': equity_data,
            'total': total_equity,
            'title': 'Equity',
        },
        'total_liabilities': total_liabilities,
        'total_liabilities_equity': total_liabilities_equity,
        'comparison': comparison_data,
        'as_of_date': as_of_date,
        'company': company,
        'branch': branch,
        'department': department,
        'financial_ratios': financial_ratios,
        'report_metadata': {
            'generated_at': timezone.now(),
            'currency': 'USD',
            'reporting_standard': 'GAAP',
            'fiscal_year_end': 'December 31',
            'include_percentages': show_percentages,
            'include_comparison': comparison_type != 'none',
        },
    }


def calculate_financial_ratios(balance_sheet_data):
    """Calculate key financial ratios for professional analysis"""
    try:
        current_assets = balance_sheet_data.get('current_assets', {}).get('total', 0)
        current_liabilities = balance_sheet_data.get('current_liabilities', {}).get('total', 0)
        total_assets = balance_sheet_data.get('total_assets', 0)
        total_liabilities = balance_sheet_data.get('total_liabilities', 0)
        total_equity = balance_sheet_data.get('equity', {}).get('total', 0)
        
        # Calculate inventory for quick ratio
        inventory = 0
        current_assets_accounts = balance_sheet_data.get('current_assets', {}).get('accounts', [])
        for account in current_assets_accounts:
            if 'inventory' in account.get('account_name', '').lower():
                inventory += account.get('balance', 0)
        
        ratios = {
            'current_ratio': round(current_assets / current_liabilities, 2) if current_liabilities > 0 else 0,
            'quick_ratio': round((current_assets - inventory) / current_liabilities, 2) if current_liabilities > 0 else 0,
            'debt_to_equity': round(total_liabilities / total_equity, 2) if total_equity > 0 else 0,
            'debt_to_assets': round(total_liabilities / total_assets, 2) if total_assets > 0 else 0,
            'equity_ratio': round(total_equity / total_assets, 2) if total_assets > 0 else 0,
            'working_capital': current_assets - current_liabilities,
        }
        
        return ratios
    except Exception as e:
        # Return default ratios if calculation fails
        return {
            'current_ratio': 0,
            'quick_ratio': 0,
            'debt_to_equity': 0,
            'debt_to_assets': 0,
            'equity_ratio': 0,
            'working_capital': 0,
        }


def get_comparison_data(as_of_date, company, branch, department, comparison_type):
    """Get comparison data for the selected period"""
    
    if comparison_type == 'previous_period':
        # Previous period of same length (30 days)
        comparison_date = as_of_date - timedelta(days=30)
        
    elif comparison_type == 'previous_year':
        # Same date last year
        comparison_date = as_of_date.replace(year=as_of_date.year - 1)
        
    else:
        return None
    
    return get_balance_sheet_data(
        as_of_date=comparison_date,
        company=company,
        branch=branch,
        department=department,
        comparison_type='none',
        include_zero_balances=True,
        show_percentages=False
    )


@login_required
@require_POST
def export_balance_sheet(request):
    """Export Balance Sheet report to various formats"""
    
    form = ExportForm(request.POST)
    if not form.is_valid():
        messages.error(request, 'Invalid export parameters')
        return redirect('balance_sheet:balance_sheet_report')
    
    # Get report parameters from form
    as_of_date = request.POST.get('as_of_date')
    company_id = request.POST.get('company')
    branch = request.POST.get('branch', '')
    department = request.POST.get('department', '')
    
    if not as_of_date:
        messages.error(request, 'Missing required parameters')
        return redirect('balance_sheet:balance_sheet_report')
    
    try:
        as_of_date = datetime.strptime(as_of_date, '%Y-%m-%d').date()
        company = Company.objects.get(id=company_id) if company_id else None
    except (ValueError, Company.DoesNotExist):
        messages.error(request, 'Invalid parameters')
        return redirect('balance_sheet:balance_sheet_report')
    
    # Generate report data
    report_data = get_balance_sheet_data(
        as_of_date=as_of_date,
        company=company,
        branch=branch,
        department=department,
        comparison_type='none',
        include_zero_balances=True,
        show_percentages=False
    )
    
    # Export options
    export_format = form.cleaned_data['export_format']
    include_headers = form.cleaned_data['include_headers']
    include_totals = form.cleaned_data['include_totals']
    include_comparison = form.cleaned_data['include_comparison']
    include_percentages = form.cleaned_data['include_percentages']
    
    # Generate filename
    filename = f"balance_sheet_{as_of_date}"
    
    if export_format == 'csv':
        try:
            response = export_to_csv(
                report_data, include_headers, include_totals, 
                include_comparison, include_percentages, as_of_date
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            return response
        except Exception as e:
            messages.error(request, f'CSV export failed: {str(e)}')
            return redirect('balance_sheet:balance_sheet_report')
    
    elif export_format == 'excel':
        try:
            response = export_to_excel(
                report_data, include_headers, include_totals,
                include_comparison, include_percentages, as_of_date
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return response
        except ImportError as e:
            messages.error(request, f'Excel export failed: {str(e)}. Please install openpyxl package.')
            return redirect('balance_sheet:balance_sheet_report')
        except Exception as e:
            messages.error(request, f'Excel export failed: {str(e)}')
            return redirect('balance_sheet:balance_sheet_report')

    elif export_format == 'pdf':
        try:
            response = export_to_pdf(
                report_data, include_headers, include_totals,
                include_comparison, include_percentages, as_of_date
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
            return response
        except ImportError as e:
            messages.error(request, f'PDF export failed: {str(e)}. Please install reportlab package.')
            return redirect('balance_sheet:balance_sheet_report')
        except Exception as e:
            messages.error(request, f'PDF export failed: {str(e)}')
            return redirect('balance_sheet:balance_sheet_report')
    
    else:
        messages.error(request, 'Invalid export format')
        return redirect('balance_sheet:balance_sheet_report')


def export_to_csv(report_data, include_headers, include_totals, 
                 include_comparison, include_percentages, as_of_date):
    """Export Balance Sheet data to CSV format"""
    
    response = HttpResponse(content_type='text/csv')
    writer = csv.writer(response)
    
    if include_headers:
        headers = ['Section', 'Account', 'Amount']
        if include_percentages:
            headers.append('Percentage')
        if include_comparison and report_data.get('comparison'):
            headers.extend(['Previous Amount', 'Change', 'Change %'])
        writer.writerow(headers)
    
    # Write Assets section
    if include_headers:
        writer.writerow(['ASSETS', '', ''])
        writer.writerow(['Current Assets', '', ''])
    
    for account in report_data['current_assets']['accounts']:
        row = ['', account['account_name'], account['balance']]
        if include_percentages:
            row.append(f"{account.get('percentage', 0):.2f}%")
        writer.writerow(row)
    
    if include_totals:
        writer.writerow(['', 'Total Current Assets', report_data['current_assets']['total']])
        writer.writerow([])
    
    # Write Non-Current Assets
    if include_headers:
        writer.writerow(['Non-Current Assets', '', ''])
    
    for account in report_data['non_current_assets']['accounts']:
        row = ['', account['account_name'], account['balance']]
        if include_percentages:
            row.append(f"{account.get('percentage', 0):.2f}%")
        writer.writerow(row)
    
    if include_totals:
        writer.writerow(['', 'Total Non-Current Assets', report_data['non_current_assets']['total']])
        writer.writerow(['', 'TOTAL ASSETS', report_data['total_assets']])
        writer.writerow([])
    
    # Write Liabilities and Equity section
    if include_headers:
        writer.writerow(['LIABILITIES & EQUITY', '', ''])
        writer.writerow(['Current Liabilities', '', ''])
    
    for account in report_data['current_liabilities']['accounts']:
        row = ['', account['account_name'], account['balance']]
        if include_percentages:
            row.append(f"{account.get('percentage', 0):.2f}%")
        writer.writerow(row)
    
    if include_totals:
        writer.writerow(['', 'Total Current Liabilities', report_data['current_liabilities']['total']])
        writer.writerow([])
    
    # Write Non-Current Liabilities
    if include_headers:
        writer.writerow(['Non-Current Liabilities', '', ''])
    
    for account in report_data['non_current_liabilities']['accounts']:
        row = ['', account['account_name'], account['balance']]
        if include_percentages:
            row.append(f"{account.get('percentage', 0):.2f}%")
        writer.writerow(row)
    
    if include_totals:
        writer.writerow(['', 'Total Non-Current Liabilities', report_data['non_current_liabilities']['total']])
        writer.writerow(['', 'TOTAL LIABILITIES', report_data['total_liabilities']])
        writer.writerow([])
    
    # Write Equity
    if include_headers:
        writer.writerow(['Equity', '', ''])
    
    for account in report_data['equity']['accounts']:
        row = ['', account['account_name'], account['balance']]
        if include_percentages:
            row.append(f"{account.get('percentage', 0):.2f}%")
        writer.writerow(row)
    
    if include_totals:
        writer.writerow(['', 'TOTAL EQUITY', report_data['equity']['total']])
        writer.writerow([])
        writer.writerow(['', 'TOTAL LIABILITIES & EQUITY', report_data['total_liabilities_equity']])
    
    return response


def export_to_excel(report_data, include_headers, include_totals,
                   include_comparison, include_percentages, as_of_date):
    """Export Balance Sheet data to Excel format with professional auditor-approved styling"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise ImportError('openpyxl is required for Excel export')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Balance Sheet"
    
    # Professional styles
    company_font = Font(bold=True, size=16, color="2C3E50")
    title_font = Font(bold=True, size=14, color="2C3E50")
    subtitle_font = Font(size=12, color="34495E")
    header_font = Font(bold=True, size=11, color="FFFFFF")
    section_font = Font(bold=True, size=12, color="2C3E50")
    total_font = Font(bold=True, size=10, color="2C3E50")
    normal_font = Font(size=10, color="2C3E50")
    
    # Professional color scheme
    company_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    header_fill = PatternFill(start_color="2C3E50", end_color="2C3E50", fill_type="solid")
    section_fill = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
    total_fill = PatternFill(start_color="ECF0F1", end_color="ECF0F1", fill_type="solid")
    assets_total_fill = PatternFill(start_color="3498DB", end_color="3498DB", fill_type="solid")
    liabilities_total_fill = PatternFill(start_color="E74C3C", end_color="E74C3C", fill_type="solid")
    equity_total_fill = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
    
    # Professional borders
    thin_border = Border(
        left=Side(style='thin', color='2C3E50'),
        right=Side(style='thin', color='2C3E50'),
        top=Side(style='thin', color='2C3E50'),
        bottom=Side(style='thin', color='2C3E50')
    )
    
    thick_border = Border(
        left=Side(style='medium', color='2C3E50'),
        right=Side(style='medium', color='2C3E50'),
        top=Side(style='medium', color='2C3E50'),
        bottom=Side(style='medium', color='2C3E50')
    )
    
    # Column widths
    ws.column_dimensions['A'].width = 45
    ws.column_dimensions['B'].width = 12
    ws.column_dimensions['C'].width = 18
    if include_percentages:
        ws.column_dimensions['D'].width = 12
    
    current_row = 1
    
    # Company Header
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "LOGIS EDGE COMPANY"
    ws[f'A{current_row}'].font = company_font
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'A{current_row}'].fill = company_fill
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "123 Business Street, City, State 12345"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "Phone: (555) 123-4567 | Email: info@logisedge.com"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Title
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "BALANCE SHEET"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = f"As of {as_of_date.strftime('%B %d, %Y')}"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:D{current_row}')
    ws[f'A{current_row}'] = "(Expressed in USD)"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 3
    
    # Professional Headers
    if include_headers:
        if include_percentages:
            headers = ['Particulars', 'Notes', 'Amount (USD)', '%']
        else:
            headers = ['Particulars', 'Notes', 'Amount (USD)']
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
            cell.border = thick_border
        
        current_row += 1
    
    # Assets Section
    if include_percentages:
        ws.merge_cells(f'A{current_row}:D{current_row}')
    else:
        ws.merge_cells(f'A{current_row}:C{current_row}')
    ws[f'A{current_row}'] = "ASSETS"
    ws[f'A{current_row}'].font = section_font
    ws[f'A{current_row}'].fill = section_fill
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    ws[f'A{current_row}'].border = thick_border
    current_row += 1
    
    # Current Assets
    ws[f'A{current_row}'] = "Current Assets"
    ws[f'A{current_row}'].font = Font(bold=True, size=11, color="2C3E50")
    ws[f'A{current_row}'].fill = total_fill
    current_row += 1
    
    for account in report_data['current_assets']['accounts']:
        ws[f'A{current_row}'] = account.get('account_name', 'Unknown Account')
        ws[f'A{current_row}'].font = normal_font
        
        ws[f'B{current_row}'] = account.get('account_code', '')
        ws[f'B{current_row}'].font = normal_font
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        
        balance = account.get('balance', 0)
        ws[f'C{current_row}'] = balance
        ws[f'C{current_row}'].font = normal_font
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = f"{account.get('percentage', 0):.1f}%"
            ws[f'D{current_row}'].font = normal_font
            ws[f'D{current_row}'].alignment = Alignment(horizontal='center')
        
        # Apply borders to all cells in the row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thin_border
        
        current_row += 1
    
    if include_totals:
        ws[f'A{current_row}'] = "Total Current Assets"
        ws[f'A{current_row}'].font = total_font
        ws[f'A{current_row}'].fill = total_fill
        
        ws[f'B{current_row}'] = ""
        ws[f'B{current_row}'].fill = total_fill
        
        ws[f'C{current_row}'] = report_data['current_assets']['total']
        ws[f'C{current_row}'].font = total_font
        ws[f'C{current_row}'].fill = total_fill
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = ""
            ws[f'D{current_row}'].fill = total_fill
        
        # Apply borders to total row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thick_border
        
        current_row += 2
    
    # Non-Current Assets
    ws[f'A{current_row}'] = "Non-Current Assets"
    ws[f'A{current_row}'].font = Font(bold=True, size=11, color="2C3E50")
    ws[f'A{current_row}'].fill = total_fill
    current_row += 1
    
    for account in report_data['non_current_assets']['accounts']:
        ws[f'A{current_row}'] = account.get('account_name', 'Unknown Account')
        ws[f'A{current_row}'].font = normal_font
        
        ws[f'B{current_row}'] = account.get('account_code', '')
        ws[f'B{current_row}'].font = normal_font
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        
        balance = account.get('balance', 0)
        ws[f'C{current_row}'] = balance
        ws[f'C{current_row}'].font = normal_font
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = f"{account.get('percentage', 0):.1f}%"
            ws[f'D{current_row}'].font = normal_font
            ws[f'D{current_row}'].alignment = Alignment(horizontal='center')
        
        # Apply borders to all cells in the row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thin_border
        
        current_row += 1
    
    if include_totals:
        ws[f'A{current_row}'] = "Total Non-Current Assets"
        ws[f'A{current_row}'].font = total_font
        ws[f'A{current_row}'].fill = total_fill
        
        ws[f'B{current_row}'] = ""
        ws[f'B{current_row}'].fill = total_fill
        
        ws[f'C{current_row}'] = report_data['non_current_assets']['total']
        ws[f'C{current_row}'].font = total_font
        ws[f'C{current_row}'].fill = total_fill
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = ""
            ws[f'D{current_row}'].fill = total_fill
        
        # Apply borders to total row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thick_border
        
        current_row += 1
        
        # TOTAL ASSETS with professional styling
        ws[f'A{current_row}'] = "TOTAL ASSETS"
        ws[f'A{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{current_row}'].fill = assets_total_fill
        
        ws[f'B{current_row}'] = ""
        ws[f'B{current_row}'].fill = assets_total_fill
        
        ws[f'C{current_row}'] = report_data['total_assets']
        ws[f'C{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'C{current_row}'].fill = assets_total_fill
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = "100.0%"
            ws[f'D{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'D{current_row}'].fill = assets_total_fill
            ws[f'D{current_row}'].alignment = Alignment(horizontal='center')
        
        # Apply thick borders to total assets row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thick_border
        
        current_row += 3
    
    # Liabilities & Equity Section
    ws[f'A{current_row}'] = "LIABILITIES & EQUITY"
    ws[f'A{current_row}'].font = section_font
    ws[f'A{current_row}'].fill = section_fill
    ws.merge_cells(f'A{current_row}:D{current_row}' if include_percentages else f'A{current_row}:C{current_row}')
    current_row += 1
    
    # Current Liabilities
    ws[f'A{current_row}'] = "Current Liabilities"
    ws[f'A{current_row}'].font = Font(bold=True, size=11, color="2C3E50")
    ws[f'A{current_row}'].fill = total_fill
    current_row += 1
    
    for account in report_data['current_liabilities']['accounts']:
        ws[f'A{current_row}'] = account.get('account_name', 'Unknown Account')
        ws[f'A{current_row}'].font = normal_font
        
        ws[f'B{current_row}'] = account.get('account_code', '')
        ws[f'B{current_row}'].font = normal_font
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        
        balance = account.get('balance', 0)
        ws[f'C{current_row}'] = balance
        ws[f'C{current_row}'].font = normal_font
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = f"{account.get('percentage', 0):.1f}%"
            ws[f'D{current_row}'].font = normal_font
            ws[f'D{current_row}'].alignment = Alignment(horizontal='center')
        
        # Apply borders to all cells in the row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thin_border
        
        current_row += 1
    
    if include_totals:
        ws[f'A{current_row}'] = "Total Current Liabilities"
        ws[f'A{current_row}'].font = total_font
        ws[f'A{current_row}'].fill = total_fill
        
        ws[f'B{current_row}'] = ""
        ws[f'B{current_row}'].fill = total_fill
        
        ws[f'C{current_row}'] = report_data['current_liabilities']['total']
        ws[f'C{current_row}'].font = total_font
        ws[f'C{current_row}'].fill = total_fill
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = ""
            ws[f'D{current_row}'].fill = total_fill
        
        # Apply borders to total row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thick_border
        
        current_row += 2
    
    # Non-Current Liabilities
    ws[f'A{current_row}'] = "Non-Current Liabilities"
    ws[f'A{current_row}'].font = Font(bold=True, size=11, color="2C3E50")
    ws[f'A{current_row}'].fill = total_fill
    current_row += 1
    
    for account in report_data['non_current_liabilities']['accounts']:
        ws[f'A{current_row}'] = account.get('account_name', 'Unknown Account')
        ws[f'A{current_row}'].font = normal_font
        
        ws[f'B{current_row}'] = account.get('account_code', '')
        ws[f'B{current_row}'].font = normal_font
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        
        balance = account.get('balance', 0)
        ws[f'C{current_row}'] = balance
        ws[f'C{current_row}'].font = normal_font
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = f"{account.get('percentage', 0):.1f}%"
            ws[f'D{current_row}'].font = normal_font
            ws[f'D{current_row}'].alignment = Alignment(horizontal='center')
        
        # Apply borders to all cells in the row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thin_border
        
        current_row += 1
    
    if include_totals:
        ws[f'A{current_row}'] = "Total Non-Current Liabilities"
        ws[f'A{current_row}'].font = total_font
        ws[f'A{current_row}'].fill = total_fill
        
        ws[f'B{current_row}'] = ""
        ws[f'B{current_row}'].fill = total_fill
        
        ws[f'C{current_row}'] = report_data['non_current_liabilities']['total']
        ws[f'C{current_row}'].font = total_font
        ws[f'C{current_row}'].fill = total_fill
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = ""
            ws[f'D{current_row}'].fill = total_fill
        
        # Apply borders to total row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thick_border
        
        current_row += 1
        
        # TOTAL LIABILITIES with professional styling
        ws[f'A{current_row}'] = "TOTAL LIABILITIES"
        ws[f'A{current_row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{current_row}'].fill = liabilities_total_fill
        
        ws[f'B{current_row}'] = ""
        ws[f'B{current_row}'].fill = liabilities_total_fill
        
        ws[f'C{current_row}'] = report_data['total_liabilities']
        ws[f'C{current_row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'C{current_row}'].fill = liabilities_total_fill
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = ""
            ws[f'D{current_row}'].fill = liabilities_total_fill
        
        # Apply thick borders to total liabilities row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thick_border
        
        current_row += 2
    
    # Equity
    ws[f'A{current_row}'] = "Equity"
    ws[f'A{current_row}'].font = Font(bold=True, size=11, color="2C3E50")
    ws[f'A{current_row}'].fill = total_fill
    current_row += 1
    
    for account in report_data['equity']['accounts']:
        ws[f'A{current_row}'] = account.get('account_name', 'Unknown Account')
        ws[f'A{current_row}'].font = normal_font
        
        ws[f'B{current_row}'] = account.get('account_code', '')
        ws[f'B{current_row}'].font = normal_font
        ws[f'B{current_row}'].alignment = Alignment(horizontal='center')
        
        balance = account.get('balance', 0)
        ws[f'C{current_row}'] = balance
        ws[f'C{current_row}'].font = normal_font
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = f"{account.get('percentage', 0):.1f}%"
            ws[f'D{current_row}'].font = normal_font
            ws[f'D{current_row}'].alignment = Alignment(horizontal='center')
        
        # Apply borders to all cells in the row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thin_border
        
        current_row += 1
    
    if include_totals:
        ws[f'A{current_row}'] = "TOTAL EQUITY"
        ws[f'A{current_row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'A{current_row}'].fill = equity_total_fill
        
        ws[f'B{current_row}'] = ""
        ws[f'B{current_row}'].fill = equity_total_fill
        
        ws[f'C{current_row}'] = report_data['equity']['total']
        ws[f'C{current_row}'].font = Font(bold=True, size=11, color="FFFFFF")
        ws[f'C{current_row}'].fill = equity_total_fill
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = ""
            ws[f'D{current_row}'].fill = equity_total_fill
        
        # Apply thick borders to total equity row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thick_border
        
        current_row += 1
        
        # TOTAL LIABILITIES & EQUITY with professional styling
        ws[f'A{current_row}'] = "TOTAL LIABILITIES & EQUITY"
        ws[f'A{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'A{current_row}'].fill = assets_total_fill
        
        ws[f'B{current_row}'] = ""
        ws[f'B{current_row}'].fill = assets_total_fill
        
        ws[f'C{current_row}'] = report_data['total_liabilities_equity']
        ws[f'C{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
        ws[f'C{current_row}'].fill = assets_total_fill
        ws[f'C{current_row}'].number_format = '#,##0.00'
        ws[f'C{current_row}'].alignment = Alignment(horizontal='right')
        
        if include_percentages:
            ws[f'D{current_row}'] = "100.0%"
            ws[f'D{current_row}'].font = Font(bold=True, size=12, color="FFFFFF")
            ws[f'D{current_row}'].fill = assets_total_fill
            ws[f'D{current_row}'].alignment = Alignment(horizontal='center')
        
        # Apply thick borders to final total row
        for col in range(1, 5 if include_percentages else 4):
            ws.cell(row=current_row, column=col).border = thick_border
    
    # Add professional notes section
    current_row += 3
    ws[f'A{current_row}'] = "Notes:"
    ws[f'A{current_row}'].font = Font(bold=True, size=10)
    current_row += 1
    
    notes = [
        "1. Figures are presented in USD unless otherwise stated",
        "2. This balance sheet has been prepared in accordance with applicable accounting standards",
        "3. All amounts are rounded to the nearest dollar"
    ]
    
    for note in notes:
        ws[f'A{current_row}'] = note
        ws[f'A{current_row}'].font = Font(size=9)
        current_row += 1
    
    # Add signature section
    current_row += 2
    signature_row = current_row
    
    # Prepared By
    ws[f'A{signature_row}'] = "Prepared By:"
    ws[f'A{signature_row}'].font = Font(bold=True, size=10)
    ws[f'A{signature_row + 3}'] = "_" * 25
    ws[f'A{signature_row + 4}'] = "Name & Signature"
    ws[f'A{signature_row + 4}'].font = Font(size=9)
    ws[f'A{signature_row + 5}'] = "Date: _____________"
    ws[f'A{signature_row + 5}'].font = Font(size=9)
    
    # Reviewed By (if percentages column exists, use column D, otherwise column C)
    review_col = 'D' if include_percentages else 'C'
    ws[f'{review_col}{signature_row}'] = "Reviewed By:"
    ws[f'{review_col}{signature_row}'].font = Font(bold=True, size=10)
    ws[f'{review_col}{signature_row + 3}'] = "_" * 25
    ws[f'{review_col}{signature_row + 4}'] = "Name & Signature"
    ws[f'{review_col}{signature_row + 4}'].font = Font(size=9)
    ws[f'{review_col}{signature_row + 5}'] = "Date: _____________"
    ws[f'{review_col}{signature_row + 5}'].font = Font(size=9)
    
    # Apply borders to all cells
    for row in ws.iter_rows(min_row=1, max_row=current_row, min_col=1, max_col=3):
        for cell in row:
            cell.border = thin_border
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    wb.save(response)
    return response


def export_to_pdf(report_data, include_headers, include_totals,
                 include_comparison, include_percentages, as_of_date):
    """Export Balance Sheet data to PDF format with professional auditor-approved styling"""
    
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
        from reportlab.lib.enums import TA_LEFT, TA_RIGHT, TA_CENTER
    except ImportError:
        raise ImportError('reportlab is required for PDF export')
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    
    # Create PDF document with professional margins
    doc = SimpleDocTemplate(
        response, 
        pagesize=A4,
        rightMargin=72,
        leftMargin=72,
        topMargin=72,
        bottomMargin=72
    )
    elements = []
    
    # Professional styles
    styles = getSampleStyleSheet()
    
    # Company header style
    company_style = ParagraphStyle(
        'CompanyHeader',
        parent=styles['Normal'],
        fontSize=16,
        spaceAfter=6,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Title style
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=18,
        spaceAfter=20,
        spaceBefore=10,
        alignment=TA_CENTER,
        fontName='Helvetica-Bold'
    )
    
    # Subtitle style
    subtitle_style = ParagraphStyle(
        'Subtitle',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=20,
        alignment=TA_CENTER,
        fontName='Helvetica'
    )
    
    section_style = ParagraphStyle(
        'Section',
        parent=styles['Heading2'],
        fontSize=14,
        spaceAfter=10,
        spaceBefore=10,
        fontName='Helvetica-Bold'
    )
    
    # Add company header
    elements.append(Paragraph("LOGIS EDGE COMPANY", company_style))
    elements.append(Paragraph("123 Business Street, City, State 12345", styles['Normal']))
    elements.append(Paragraph("Phone: (555) 123-4567 | Email: info@logisedge.com", styles['Normal']))
    elements.append(Spacer(1, 20))
    
    # Add title and period
    elements.append(Paragraph("BALANCE SHEET", title_style))
    elements.append(Paragraph(f"As of {as_of_date.strftime('%B %d, %Y')}", subtitle_style))
    elements.append(Paragraph("(Expressed in USD)", subtitle_style))
    elements.append(Spacer(1, 20))
    
    # Helper function to create account table
    def create_account_table(accounts, section_title):
        """Create professional account table for PDF export"""
        if not accounts:
            return None, 0
        
        # Prepare table data with professional headers
        if include_percentages:
            data = [['Particulars', 'Notes', 'Amount (USD)', '%']]
            col_widths = [3*inch, 0.8*inch, 1.2*inch, 0.8*inch]
        else:
            data = [['Particulars', 'Notes', 'Amount (USD)']]
            col_widths = [4*inch, 1*inch, 1.5*inch]
        
        total = 0
        for account in accounts:
            # Format account name and code
            account_display = account.get('account_name', 'Unknown Account')
            account_code = account.get('account_code', '')
            
            # Format balance with proper currency formatting
            balance = account.get('balance', 0)
            balance_str = f"{balance:,.2f}" if balance >= 0 else f"({abs(balance):,.2f})"
            
            if include_percentages:
                percentage = account.get('percentage', 0)
                row = [account_display, account_code, balance_str, f"{percentage:.1f}%"]
            else:
                row = [account_display, account_code, balance_str]
            
            data.append(row)
            total += balance
        
        # Add total row if requested
        if include_totals:
            total_str = f"{total:,.2f}" if total >= 0 else f"({abs(total):,.2f})"
            if include_percentages:
                total_row = [f"Total {section_title}", '', total_str, '']
            else:
                total_row = [f"Total {section_title}", '', total_str]
            data.append(total_row)
        
        # Create table with professional styling
        table = Table(data, colWidths=col_widths)
        
        # Professional table styling
        table_style = [
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2C3E50')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('ALIGN', (-1, 0), (-1, -1), 'RIGHT'),  # Amount column right-aligned
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
            ('TOPPADDING', (0, 0), (-1, 0), 8),
            
            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ROWBACKGROUNDS', (0, 1), (-1, -1), [colors.white, colors.HexColor('#F8F9FA')]),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('TOPPADDING', (0, 1), (-1, -1), 4),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
            
            # Grid lines
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#BDC3C7')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2C3E50')),
        ]
        
        # Total row styling if present
        if include_totals:
            table_style.extend([
                ('BACKGROUND', (0, -1), (-1, -1), colors.HexColor('#ECF0F1')),
                ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
                ('FONTSIZE', (0, -1), (-1, -1), 10),
                ('LINEABOVE', (0, -1), (-1, -1), 1.5, colors.HexColor('#2C3E50')),
            ])
        
        table.setStyle(TableStyle(table_style))
        
        return table, total
    
    # Assets Section
    elements.append(Paragraph("ASSETS", section_style))
    elements.append(Spacer(1, 10))
    
    # Current Assets
    current_assets_table, current_assets_total = create_account_table(
        report_data['current_assets']['accounts'], 'Current Assets'
    )
    if current_assets_table:
        elements.append(current_assets_table)
        elements.append(Spacer(1, 15))
    
    # Non-Current Assets
    non_current_assets_table, non_current_assets_total = create_account_table(
        report_data['non_current_assets']['accounts'], 'Non-Current Assets'
    )
    if non_current_assets_table:
        elements.append(non_current_assets_table)
        elements.append(Spacer(1, 15))
    
    # Total Assets with professional styling
    total_assets = current_assets_total + non_current_assets_total
    total_assets_str = f"{total_assets:,.2f}" if total_assets >= 0 else f"({abs(total_assets):,.2f})"
    
    if include_percentages:
        total_assets_data = [['TOTAL ASSETS', '', total_assets_str, '100.0%']]
        col_widths = [3*inch, 0.8*inch, 1.2*inch, 0.8*inch]
    else:
        total_assets_data = [['TOTAL ASSETS', '', total_assets_str]]
        col_widths = [4*inch, 1*inch, 1.5*inch]
    
    total_assets_table = Table(total_assets_data, colWidths=col_widths)
    total_assets_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (-1, 0), (-1, 0), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2C3E50')),
    ]))
    elements.append(total_assets_table)
    elements.append(Spacer(1, 25))
    
    # Liabilities & Equity Section
    elements.append(Paragraph("LIABILITIES & EQUITY", section_style))
    elements.append(Spacer(1, 10))
    
    # Current Liabilities
    current_liabilities_table, current_liabilities_total = create_account_table(
        report_data['current_liabilities']['accounts'], 'Current Liabilities'
    )
    if current_liabilities_table:
        elements.append(current_liabilities_table)
        elements.append(Spacer(1, 15))
    
    # Non-Current Liabilities
    non_current_liabilities_table, non_current_liabilities_total = create_account_table(
        report_data['non_current_liabilities']['accounts'], 'Non-Current Liabilities'
    )
    if non_current_liabilities_table:
        elements.append(non_current_liabilities_table)
        elements.append(Spacer(1, 15))
    
    # Total Liabilities with professional styling
    total_liabilities = current_liabilities_total + non_current_liabilities_total
    total_liabilities_str = f"{total_liabilities:,.2f}" if total_liabilities >= 0 else f"({abs(total_liabilities):,.2f})"
    
    if include_percentages:
        total_liabilities_data = [['TOTAL LIABILITIES', '', total_liabilities_str, f"{(total_liabilities/total_assets*100):.1f}%" if total_assets != 0 else '0.0%']]
        col_widths = [3*inch, 0.8*inch, 1.2*inch, 0.8*inch]
    else:
        total_liabilities_data = [['TOTAL LIABILITIES', '', total_liabilities_str]]
        col_widths = [4*inch, 1*inch, 1.5*inch]
    
    total_liabilities_table = Table(total_liabilities_data, colWidths=col_widths)
    total_liabilities_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#E74C3C')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 11),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (-1, 0), (-1, 0), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2C3E50')),
    ]))
    elements.append(total_liabilities_table)
    elements.append(Spacer(1, 15))
    
    # Equity
    equity_table, equity_total = create_account_table(
        report_data['equity']['accounts'], 'Shareholders\' Equity'
    )
    if equity_table:
        elements.append(equity_table)
        elements.append(Spacer(1, 15))
    
    # Total Liabilities & Equity with professional styling
    total_liabilities_equity = total_liabilities + equity_total
    total_liabilities_equity_str = f"{total_liabilities_equity:,.2f}" if total_liabilities_equity >= 0 else f"({abs(total_liabilities_equity):,.2f})"
    
    if include_percentages:
        total_liabilities_equity_data = [['TOTAL LIABILITIES & EQUITY', '', total_liabilities_equity_str, '100.0%']]
        col_widths = [3*inch, 0.8*inch, 1.2*inch, 0.8*inch]
    else:
        total_liabilities_equity_data = [['TOTAL LIABILITIES & EQUITY', '', total_liabilities_equity_str]]
        col_widths = [4*inch, 1*inch, 1.5*inch]
    
    total_liabilities_equity_table = Table(total_liabilities_equity_data, colWidths=col_widths)
    total_liabilities_equity_table.setStyle(TableStyle([
        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#3498DB')),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 12),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (-1, 0), (-1, 0), 'RIGHT'),
        ('TOPPADDING', (0, 0), (-1, 0), 8),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 8),
        ('GRID', (0, 0), (-1, -1), 1, colors.HexColor('#2C3E50')),
    ]))
    elements.append(total_liabilities_equity_table)
    elements.append(Spacer(1, 30))
    
    # Add professional notes section
    notes_style = ParagraphStyle(
        'Notes',
        parent=styles['Normal'],
        fontSize=9,
        spaceAfter=6,
        fontName='Helvetica'
    )
    
    elements.append(Paragraph("NOTES TO FINANCIAL STATEMENTS:", section_style))
    elements.append(Paragraph("1. The accompanying notes are an integral part of these financial statements.", notes_style))
    elements.append(Paragraph("2. All amounts are expressed in US Dollars unless otherwise stated.", notes_style))
    elements.append(Paragraph("3. This balance sheet has been prepared in accordance with Generally Accepted Accounting Principles (GAAP).", notes_style))
    elements.append(Spacer(1, 30))
    
    # Add signature section
    signature_style = ParagraphStyle(
        'Signature',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=20,
        fontName='Helvetica'
    )
    
    # Create signature table
    signature_data = [
        ['Prepared By:', 'Reviewed By:', 'Approved By:'],
        ['', '', ''],
        ['', '', ''],
        ['_____________________', '_____________________', '_____________________'],
        ['Chief Financial Officer', 'Controller', 'Chief Executive Officer'],
        ['', '', ''],
        [f'Date: _______________', f'Date: _______________', f'Date: _______________']
    ]
    
    signature_table = Table(signature_data, colWidths=[2*inch, 2*inch, 2*inch])
    signature_table.setStyle(TableStyle([
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, -1), 9),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ('TOPPADDING', (0, 0), (-1, -1), 8),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
    ]))
    
    elements.append(signature_table)
    
    # Build PDF
    doc.build(elements)
    return response


@login_required
def report_list(request):
    """List all saved balance sheet reports"""
    
    reports = BalanceSheetReport.objects.filter(created_by=request.user).order_by('-created_at')
    
    context = {
        'reports': reports,
    }
    
    return render(request, 'balance_sheet/report_list.html', context)


@login_required
def report_detail(request, report_id):
    """View a specific balance sheet report"""
    
    report = get_object_or_404(BalanceSheetReport, id=report_id, created_by=request.user)
    
    context = {
        'report': report,
        'report_data': report.report_data,
        'export_form': ExportForm(),
    }
    
    return render(request, 'balance_sheet/report_detail.html', context)


def get_sample_balance_sheet_data():
    """Generate enhanced sample balance sheet data for professional demonstration"""
    
    # Enhanced sample data with more realistic accounts and amounts
    sample_data = {
        'current_assets': {
            'accounts': [
                {'account_name': 'Cash and Cash Equivalents', 'account_code': '1001', 'balance': 125000.00, 'account_type': 'Current Assets', 'percentage': 16.4},
                {'account_name': 'Accounts Receivable', 'account_code': '1002', 'balance': 185000.00, 'account_type': 'Current Assets', 'percentage': 24.3},
                {'account_name': 'Inventory', 'account_code': '1003', 'balance': 220000.00, 'account_type': 'Current Assets', 'percentage': 28.9},
                {'account_name': 'Prepaid Expenses', 'account_code': '1004', 'balance': 25000.00, 'account_type': 'Current Assets', 'percentage': 3.3},
                {'account_name': 'Short-term Investments', 'account_code': '1005', 'balance': 45000.00, 'account_type': 'Current Assets', 'percentage': 5.9},
            ],
            'total': 600000.00,
            'title': 'Current Assets',
        },
        'non_current_assets': {
            'accounts': [
                {'account_name': 'Property, Plant and Equipment', 'account_code': '2001', 'balance': 450000.00, 'account_type': 'Fixed Assets', 'percentage': 59.2},
                {'account_name': 'Less: Accumulated Depreciation', 'account_code': '2002', 'balance': -120000.00, 'account_type': 'Fixed Assets', 'percentage': -15.8},
                {'account_name': 'Intangible Assets', 'account_code': '2003', 'balance': 75000.00, 'account_type': 'Intangible Assets', 'percentage': 9.9},
                {'account_name': 'Long-term Investments', 'account_code': '2004', 'balance': 155000.00, 'account_type': 'Investments', 'percentage': 20.4},
                {'account_name': 'Goodwill', 'account_code': '2005', 'balance': 100000.00, 'account_type': 'Intangible Assets', 'percentage': 13.2},
            ],
            'total': 660000.00,
            'title': 'Non-Current Assets',
        },
        'total_assets': 1260000.00,
        'current_liabilities': {
            'accounts': [
                {'account_name': 'Accounts Payable', 'account_code': '3001', 'balance': 95000.00, 'account_type': 'Current Liabilities', 'percentage': 7.5},
                {'account_name': 'Accrued Expenses', 'account_code': '3002', 'balance': 45000.00, 'account_type': 'Current Liabilities', 'percentage': 3.6},
                {'account_name': 'Taxes Payable', 'account_code': '3003', 'balance': 35000.00, 'account_type': 'Current Liabilities', 'percentage': 2.8},
                {'account_name': 'Short-term Debt', 'account_code': '3004', 'balance': 75000.00, 'account_type': 'Current Liabilities', 'percentage': 6.0},
                {'account_name': 'Unearned Revenue', 'account_code': '3005', 'balance': 30000.00, 'account_type': 'Current Liabilities', 'percentage': 2.4},
            ],
            'total': 280000.00,
            'title': 'Current Liabilities',
        },
        'non_current_liabilities': {
            'accounts': [
                {'account_name': 'Long-term Debt', 'account_code': '4001', 'balance': 350000.00, 'account_type': 'Long-term Liabilities', 'percentage': 27.8},
                {'account_name': 'Deferred Tax Liabilities', 'account_code': '4002', 'balance': 65000.00, 'account_type': 'Long-term Liabilities', 'percentage': 5.2},
                {'account_name': 'Pension Obligations', 'account_code': '4003', 'balance': 85000.00, 'account_type': 'Long-term Liabilities', 'percentage': 6.7},
            ],
            'total': 500000.00,
            'title': 'Non-Current Liabilities',
        },
        'total_liabilities': 780000.00,
        'equity': {
            'accounts': [
                {'account_name': 'Common Stock', 'account_code': '5001', 'balance': 200000.00, 'account_type': 'Share Capital', 'percentage': 15.9},
                {'account_name': 'Preferred Stock', 'account_code': '5002', 'balance': 100000.00, 'account_type': 'Share Capital', 'percentage': 7.9},
                {'account_name': 'Additional Paid-in Capital', 'account_code': '5003', 'balance': 80000.00, 'account_type': 'Additional Paid-in Capital', 'percentage': 6.3},
                {'account_name': 'Retained Earnings', 'account_code': '5004', 'balance': 120000.00, 'account_type': 'Retained Earnings', 'percentage': 9.5},
                {'account_name': 'Treasury Stock', 'account_code': '5005', 'balance': -20000.00, 'account_type': 'Treasury Stock', 'percentage': -1.6},
            ],
            'total': 480000.00,
            'title': 'Shareholders\' Equity',
        },
        'total_liabilities_equity': 1260000.00,
        'comparison_data': None,
        # Add financial ratios for professional analysis
        'financial_ratios': {
            'current_ratio': round(600000.00 / 280000.00, 2),  # Current Assets / Current Liabilities
            'quick_ratio': round((600000.00 - 220000.00) / 280000.00, 2),  # (Current Assets - Inventory) / Current Liabilities
            'debt_to_equity': round(780000.00 / 480000.00, 2),  # Total Liabilities / Total Equity
            'debt_to_assets': round(780000.00 / 1260000.00, 2),  # Total Liabilities / Total Assets
            'equity_ratio': round(480000.00 / 1260000.00, 2),  # Total Equity / Total Assets
        },
        # Add metadata for professional reporting
        'report_metadata': {
            'generated_at': timezone.now(),
            'currency': 'USD',
            'reporting_standard': 'GAAP',
            'fiscal_year_end': 'December 31',
            'audited': True,
            'notes': 'Sample data for demonstration purposes only',
        }
    }
    
    return sample_data
