from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import HttpResponse, JsonResponse
from django.views.decorators.http import require_POST
from django.db.models import Sum, Q, Case, When, F, DecimalField
from django.utils import timezone
from datetime import datetime, timedelta
import csv
import json

from .forms import ProfitLossReportForm, ExportForm
from .models import ProfitLossReport, ReportTemplate

from company.company_model import Company
from chart_of_accounts.models import ChartOfAccount, AccountType
from ledger.models import Ledger


@login_required
def profit_loss_report(request):
    """Main Profit & Loss report view"""
    
    print(f"DEBUG: Request method: {request.method}")
    print(f"DEBUG: POST data: {request.POST}")
    print(f"DEBUG: User: {request.user}")
    
    if request.method == 'POST':
        print("DEBUG: Processing POST request")
        form = ProfitLossReportForm(request.POST)
        print(f"DEBUG: Form is valid: {form.is_valid()}")
        if not form.is_valid():
            print(f"DEBUG: Form errors: {form.errors}")
        
        if form.is_valid():
            print("DEBUG: Form validation passed")
            # Get form data
            from_date = form.cleaned_data['from_date']
            to_date = form.cleaned_data['to_date']
            company = form.cleaned_data.get('company')
            comparison_type = form.cleaned_data.get('comparison_type', 'none')
            
            try:
                print(f"DEBUG: Generating report data for dates {from_date} to {to_date}, company: {company}")
                # Generate report data
                report_data = get_profit_loss_data(
                    from_date=from_date,
                    to_date=to_date,
                    company=company,
                    comparison_type=comparison_type,
                    include_zero_balances=form.cleaned_data.get('include_zero_balances', True),
                    show_percentages=form.cleaned_data.get('show_percentages', True)
                )
                print(f"DEBUG: Report data generated: {report_data}")
                
                # Check if report has meaningful data
                has_data = (
                    report_data and (
                        abs(report_data.get('total_revenue', 0)) > 0 or
                        abs(report_data.get('total_cogs', 0)) > 0 or
                        abs(report_data.get('total_expenses', 0)) > 0 or
                        abs(report_data.get('total_other_income', 0)) > 0 or
                        abs(report_data.get('total_other_expenses', 0)) > 0
                    )
                )
                print(f"DEBUG: Has data: {has_data}")
                print(f"DEBUG: Report data keys: {list(report_data.keys()) if report_data else 'None'}")
                
                if not has_data:
                    warning_msg = f'No financial data found for the selected period ({from_date} to {to_date}). Please check your date range or ensure ledger entries exist for this period.'
                    
                    # Handle AJAX requests
                    if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                        return JsonResponse({
                            'success': False,
                            'error': warning_msg,
                            'no_data': True
                        })
                    
                    messages.warning(request, warning_msg)
                    context = {
                        'form': form,
                        'report_data': None,
                        'no_data_message': True,
                        'from_date': from_date,
                        'to_date': to_date,
                    }
                    return render(request, 'profit_loss_statement/profit_loss_report.html', context)
                
                # Save report to database
                print("DEBUG: Saving report to database")
                report = ProfitLossReport.objects.create(
                    title=f"Profit & Loss Statement - {from_date} to {to_date}",
                    from_date=from_date,
                    to_date=to_date,
                    company=company,
                    comparison_type=comparison_type,
                    report_data=report_data,
                    created_by=request.user,
                    include_headers=form.cleaned_data.get('include_headers', True),
                    include_totals=form.cleaned_data.get('include_totals', True),
                    include_comparison=form.cleaned_data.get('include_comparison', False)
                )
                print(f"DEBUG: Report saved with ID: {report.id}")
                
                context = {
                    'form': form,
                    'report_data': report_data,
                    'report': report,
                    'from_date': from_date,
                    'to_date': to_date,
                    'company': company,
                    'show_percentages': form.cleaned_data.get('show_percentages', True),
                }
                
                # Handle AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    # Convert QuerySets and Decimals to JSON serializable types
                    def serialize_report_data(data):
                        if not data:
                            return data
                        
                        from decimal import Decimal
                        from django.db.models.query import QuerySet
                        from django.db.models.fields import DecimalField
                        
                        def serialize_object(obj):
                            if isinstance(obj, Decimal):
                                return float(obj)
                            elif isinstance(obj, QuerySet):
                                return list(obj)
                            elif isinstance(obj, dict):
                                return {k: serialize_object(v) for k, v in obj.items()}
                            elif isinstance(obj, (list, tuple)):
                                return [serialize_object(item) for item in obj]
                            else:
                                return obj
                        
                        return serialize_object(data)
                    
                    serializable_report_data = serialize_report_data(report_data)
                    
                    return JsonResponse({
                        'success': True,
                        'report_data': serializable_report_data,
                        'from_date': from_date.strftime('%Y-%m-%d'),
                        'to_date': to_date.strftime('%Y-%m-%d'),
                        'company': company.name if company else 'All Companies',
                        'show_percentages': form.cleaned_data.get('show_percentages', True),
                        'report_id': report.id
                    })
                
                print("DEBUG: Rendering template with report data")
                print(f"DEBUG: Context keys: {list(context.keys())}")
                return render(request, 'profit_loss_statement/profit_loss_report.html', context)
                
            except Exception as e:
                error_msg = f'Error generating report: {str(e)}'
                
                # Handle AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': False,
                        'error': error_msg
                    })
                
                messages.error(request, error_msg)
                context = {
                    'form': form,
                    'report_data': None,
                }
                return render(request, 'profit_loss_statement/profit_loss_report.html', context)
        else:
            error_msg = f'Form validation failed: {form.errors}'
            
            # Handle AJAX requests
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                return JsonResponse({
                    'success': False,
                    'error': error_msg,
                    'form_errors': form.errors
                })
            
            messages.error(request, error_msg)
    else:
        form = ProfitLossReportForm()
    
    context = {
        'form': form,
        'report_data': None,
    }
    
    return render(request, 'profit_loss_statement/profit_loss_report.html', context)


def get_profit_loss_data(from_date, to_date, company=None, comparison_type='none', 
                        include_zero_balances=True, show_percentages=True):
    """Generate Profit & Loss report data"""
    
    # Base query for ledger entries
    base_query = Ledger.objects.filter(
        entry_date__gte=from_date,
        entry_date__lte=to_date
    )
    
    if company:
        base_query = base_query.filter(company=company)
    
    # Get account types for different sections
    revenue_account_types = AccountType.objects.filter(
        category='REVENUE'
    ).exclude(name__icontains='other')
    
    # For COGS, we'll use specific expense types that relate to cost of goods
    cogs_account_types = AccountType.objects.filter(
        Q(name__icontains='cost') | Q(name__icontains='cogs')
    )
    
    # Operating expenses (exclude COGS-related expenses)
    expense_account_types = AccountType.objects.filter(
        category='EXPENSE'
    ).exclude(
        Q(name__icontains='cost') | Q(name__icontains='cogs')
    )
    
    # Other income and expenses
    other_income_types = AccountType.objects.filter(
        category='REVENUE',
        name__icontains='other'
    )
    
    other_expense_types = AccountType.objects.filter(
        category='EXPENSE',
        name__icontains='other'
    )
    
    # Calculate revenue (Credit entries increase revenue, Debit entries decrease it)
    revenue_accounts = ChartOfAccount.objects.filter(account_type__in=revenue_account_types)
    revenue_data = list(base_query.filter(account__in=revenue_accounts).values(
        'account__name', 'account__account_type__name'
    ).annotate(
        credit_total=Sum(Case(When(entry_type='CR', then='amount'), default=0, output_field=DecimalField())),
        debit_total=Sum(Case(When(entry_type='DR', then='amount'), default=0, output_field=DecimalField())),
        total=F('credit_total') - F('debit_total')
    ).order_by('account__account_type__name', 'account__name'))

    # Convert Decimal values to floats for JSON serialization
    for item in revenue_data:
        item['total'] = float(item['total'] or 0)
        item['credit_total'] = float(item['credit_total'] or 0)
        item['debit_total'] = float(item['debit_total'] or 0)

    total_revenue = sum(item['total'] for item in revenue_data)
    
    # Calculate COGS (Debit entries increase COGS, Credit entries decrease it)
    cogs_accounts = ChartOfAccount.objects.filter(account_type__in=cogs_account_types)
    cogs_data = list(base_query.filter(account__in=cogs_accounts).values(
        'account__name', 'account__account_type__name'
    ).annotate(
        credit_total=Sum(Case(When(entry_type='CR', then='amount'), default=0, output_field=DecimalField())),
        debit_total=Sum(Case(When(entry_type='DR', then='amount'), default=0, output_field=DecimalField())),
        total=F('debit_total') - F('credit_total')
    ).order_by('account__account_type__name', 'account__name'))

    # Convert Decimal values to floats for JSON serialization
    for item in cogs_data:
        item['total'] = float(item['total'] or 0)
        item['credit_total'] = float(item['credit_total'] or 0)
        item['debit_total'] = float(item['debit_total'] or 0)

    total_cogs = sum(item['total'] for item in cogs_data)
    
    # Calculate operating expenses (Debit entries increase expenses, Credit entries decrease them)
    expense_accounts = ChartOfAccount.objects.filter(account_type__in=expense_account_types)
    expense_data = list(base_query.filter(account__in=expense_accounts).values(
        'account__name', 'account__account_type__name'
    ).annotate(
        credit_total=Sum(Case(When(entry_type='CR', then='amount'), default=0, output_field=DecimalField())),
        debit_total=Sum(Case(When(entry_type='DR', then='amount'), default=0, output_field=DecimalField())),
        total=F('debit_total') - F('credit_total')
    ).order_by('account__account_type__name', 'account__name'))

    # Convert Decimal values to floats for JSON serialization
    for item in expense_data:
        item['total'] = float(item['total'] or 0)
        item['credit_total'] = float(item['credit_total'] or 0)
        item['debit_total'] = float(item['debit_total'] or 0)

    total_expenses = sum(item['total'] for item in expense_data)
    
    # Calculate other income/expenses
    other_income_accounts = ChartOfAccount.objects.filter(account_type__in=other_income_types)
    other_income_data = list(base_query.filter(account__in=other_income_accounts).values(
        'account__name', 'account__account_type__name'
    ).annotate(
        credit_total=Sum(Case(When(entry_type='CR', then='amount'), default=0, output_field=DecimalField())),
        debit_total=Sum(Case(When(entry_type='DR', then='amount'), default=0, output_field=DecimalField())),
        total=F('credit_total') - F('debit_total')
    ).order_by('account__account_type__name', 'account__name'))

    # Convert Decimal values to floats for JSON serialization
    for item in other_income_data:
        item['total'] = float(item['total'] or 0)
        item['credit_total'] = float(item['credit_total'] or 0)
        item['debit_total'] = float(item['debit_total'] or 0)

    total_other_income = sum(item['total'] for item in other_income_data)

    other_expense_accounts = ChartOfAccount.objects.filter(account_type__in=other_expense_types)
    other_expense_data = list(base_query.filter(account__in=other_expense_accounts).values(
        'account__name', 'account__account_type__name'
    ).annotate(
        credit_total=Sum(Case(When(entry_type='CR', then='amount'), default=0, output_field=DecimalField())),
        debit_total=Sum(Case(When(entry_type='DR', then='amount'), default=0, output_field=DecimalField())),
        total=F('debit_total') - F('credit_total')
    ).order_by('account__account_type__name', 'account__name'))

    # Convert Decimal values to floats for JSON serialization
    for item in other_expense_data:
        item['total'] = float(item['total'] or 0)
        item['credit_total'] = float(item['credit_total'] or 0)
        item['debit_total'] = float(item['debit_total'] or 0)

    total_other_expenses = sum(item['total'] for item in other_expense_data)
    
    # Calculate key metrics
    gross_profit = total_revenue - total_cogs
    operating_profit = gross_profit - total_expenses
    net_profit = operating_profit + total_other_income - total_other_expenses

    # Precompute percentages (avoid division by zero)
    def percent(val, base):
        try:
            return (val / base * 100) if base else 0
        except Exception:
            return 0

    cogs_pct = float(percent(total_cogs, total_revenue))
    gross_profit_pct = float(percent(gross_profit, total_revenue))
    expenses_pct = float(percent(total_expenses, total_revenue))
    operating_profit_pct = float(percent(operating_profit, total_revenue))
    other_income_pct = float(percent(total_other_income, total_revenue))
    other_expenses_pct = float(percent(total_other_expenses, total_revenue))
    net_profit_pct = float(percent(net_profit, total_revenue))

    # Add percentages if requested
    if show_percentages and total_revenue > 0:
        for item in revenue_data:
            item['percentage'] = float((item['total'] / total_revenue) * 100)
        for item in cogs_data:
            item['percentage'] = float((item['total'] / total_revenue) * 100)
        for item in expense_data:
            item['percentage'] = float((item['total'] / total_revenue) * 100)
        for item in other_income_data:
            item['percentage'] = float((item['total'] / total_revenue) * 100)
        for item in other_expense_data:
            item['percentage'] = float((item['total'] / total_revenue) * 100)

    # Prepare comparison data if requested
    comparison_data = None
    if comparison_type != 'none':
        comparison_data = get_comparison_data(
            from_date, to_date, company, comparison_type
        )
    
    return {
        'revenue': {
            'accounts': revenue_data,
            'total': float(total_revenue),
            'title': 'Revenue',
        },
        'cogs': {
            'accounts': cogs_data,
            'total': float(total_cogs),
            'title': 'Cost of Goods Sold',
            'pct_of_revenue': cogs_pct,
        },
        'gross_profit': float(gross_profit),
        'gross_profit_pct': gross_profit_pct,
        'expenses': {
            'accounts': expense_data,
            'total': float(total_expenses),
            'title': 'Operating Expenses',
            'pct_of_revenue': expenses_pct,
        },
        'operating_profit': float(operating_profit),
        'operating_profit_pct': operating_profit_pct,
        'other_income': {
            'accounts': other_income_data,
            'total': float(total_other_income),
            'title': 'Other Income',
            'pct_of_revenue': other_income_pct,
        },
        'other_expenses': {
            'accounts': other_expense_data,
            'total': float(total_other_expenses),
            'title': 'Other Expenses',
            'pct_of_revenue': other_expenses_pct,
        },
        'net_profit': float(net_profit),
        'net_profit_pct': net_profit_pct,
        'comparison': comparison_data,
        'total_revenue': float(total_revenue),
        'total_cogs': float(total_cogs),
        'total_expenses': float(total_expenses),
        'total_other_income': float(total_other_income),
        'total_other_expenses': float(total_other_expenses),
    }


def get_comparison_data(from_date, to_date, company, comparison_type):
    """Get comparison data for the selected period"""
    
    if comparison_type == 'previous_period':
        # Previous period of same length
        period_length = (to_date - from_date).days
        prev_to_date = from_date - timedelta(days=1)
        prev_from_date = prev_to_date - timedelta(days=period_length)
        
    elif comparison_type == 'previous_year':
        # Same period last year
        prev_from_date = from_date.replace(year=from_date.year - 1)
        prev_to_date = to_date.replace(year=to_date.year - 1)
        
    else:
        return None
    
    return get_profit_loss_data(
        from_date=prev_from_date,
        to_date=prev_to_date,
        company=company,
        comparison_type='none',
        include_zero_balances=True,
        show_percentages=True
    )


@login_required
def export_profit_loss(request):
    """Export Profit & Loss report to various formats"""
    
    # Check if this is a GET request for an existing report
    if request.method == 'GET':
        report_id = request.GET.get('report_id')
        if report_id:
            try:
                # Get the existing report
                report = ProfitLossReport.objects.get(id=report_id, created_by=request.user)
                report_data = report.report_data
                from_date = report.from_date
                to_date = report.to_date
                company = report.company
                
                # Set default export options for existing reports
                export_format = request.GET.get('format', 'excel')
                include_headers = True
                include_totals = True
                include_comparison = False
                include_percentages = True
                
                # Generate filename
                filename = f"profit_loss_{from_date}_to_{to_date}"
                
                if export_format == 'csv':
                    response = export_to_csv(
                        report_data, include_headers, include_totals, 
                        include_comparison, include_percentages, from_date, to_date
                    )
                    response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
                    return response
                
                elif export_format == 'excel':
                    try:
                        response = export_to_excel(
                            report_data, include_headers, include_totals,
                            include_comparison, include_percentages, from_date, to_date, company
                        )
                        response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
                        return response
                    except ImportError as e:
                        messages.error(request, str(e))
                        return redirect('profit_loss_statement:report_detail', report_id=report_id)

                elif export_format == 'pdf':
                    try:
                        response = export_to_pdf(
                            report_data, include_headers, include_totals,
                            include_comparison, include_percentages, from_date, to_date, company
                        )
                        response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
                        return response
                    except ImportError as e:
                        messages.error(request, str(e))
                        return redirect('profit_loss_statement:report_detail', report_id=report_id)
                
                else:
                    messages.error(request, 'Invalid export format')
                    return redirect('profit_loss_statement:report_detail', report_id=report_id)
                    
            except ProfitLossReport.DoesNotExist:
                messages.error(request, 'Report not found')
                return redirect('profit_loss_statement:report_list')
    
    # Handle POST request for new report generation
    elif request.method == 'POST':
        form = ExportForm(request.POST)
        if not form.is_valid():
            messages.error(request, 'Invalid export parameters')
            return redirect('profit_loss_statement:profit_loss_report')
        
        # Get report parameters from form
        from_date = request.POST.get('from_date')
        to_date = request.POST.get('to_date')
        company_id = request.POST.get('company')
        
        if not all([from_date, to_date]):
            messages.error(request, 'Missing required parameters')
            return redirect('profit_loss_statement:profit_loss_report')
        
        try:
            from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
            to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
            company = Company.objects.get(id=company_id) if company_id else None
        except (ValueError, Company.DoesNotExist):
            messages.error(request, 'Invalid parameters')
            return redirect('profit_loss_statement:profit_loss_report')
        
        # Generate report data
        report_data = get_profit_loss_data(
            from_date=from_date,
            to_date=to_date,
            company=company,
            comparison_type='none',
            include_zero_balances=True,
            show_percentages=True
        )
        
        # Export options
        export_format = form.cleaned_data['export_format']
        include_headers = form.cleaned_data['include_headers']
        include_totals = form.cleaned_data['include_totals']
        include_comparison = form.cleaned_data['include_comparison']
        include_percentages = form.cleaned_data['include_percentages']
        
        # Generate filename
        filename = f"profit_loss_{from_date}_to_{to_date}"
        
        if export_format == 'csv':
            response = export_to_csv(
                report_data, include_headers, include_totals, 
                include_comparison, include_percentages, from_date, to_date
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
            return response
        
        elif export_format == 'excel':
            try:
                response = export_to_excel(
                    report_data, include_headers, include_totals,
                    include_comparison, include_percentages, from_date, to_date, company
            )
                response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
                return response
            except ImportError as e:
                messages.error(request, str(e))
                return redirect('profit_loss_statement:profit_loss_report')

        elif export_format == 'pdf':
            try:
                response = export_to_pdf(
                    report_data, include_headers, include_totals,
                    include_comparison, include_percentages, from_date, to_date, company
                )
                response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
                return response
            except ImportError as e:
                messages.error(request, str(e))
                return redirect('profit_loss_statement:profit_loss_report')
        
        else:
            messages.error(request, 'Invalid export format')
            return redirect('profit_loss_statement:profit_loss_report')
    
    # If neither GET nor POST, redirect to report list
    return redirect('profit_loss_statement:profit_loss_report')


def export_to_csv(report_data, include_headers, include_totals, 
                 include_comparison, include_percentages, from_date, to_date):
    """Export Profit & Loss data to CSV format"""
    
    response = HttpResponse(content_type='text/csv')
    writer = csv.writer(response)
    
    if include_headers:
        headers = ['Section', 'Account', 'Amount']
        if include_percentages:
            headers.append('Percentage')
        if include_comparison and report_data.get('comparison'):
            headers.extend(['Previous Amount', 'Change', 'Change %'])
        writer.writerow(headers)
    
    # Write revenue section
    if include_headers:
        writer.writerow([report_data['revenue']['title'], '', ''])
    
    for account in report_data['revenue']['accounts']:
        row = ['', account['account__name'], account['total']]
        if include_percentages:
            row.append(f"{account.get('percentage', 0):.2f}%")
        writer.writerow(row)
    
    if include_totals:
        writer.writerow(['', 'Total Revenue', report_data['revenue']['total']])
        writer.writerow([])
    
    # Write COGS section
    if include_headers:
        writer.writerow([report_data['cogs']['title'], '', ''])
    
    for account in report_data['cogs']['accounts']:
        row = ['', account['account__name'], account['total']]
        if include_percentages:
            row.append(f"{account.get('percentage', 0):.2f}%")
        writer.writerow(row)
    
    if include_totals:
        writer.writerow(['', 'Total COGS', report_data['cogs']['total']])
        writer.writerow(['', 'Gross Profit', report_data['gross_profit']])
        writer.writerow([])
    
    # Write expenses section
    if include_headers:
        writer.writerow([report_data['expenses']['title'], '', ''])
    
    for account in report_data['expenses']['accounts']:
        row = ['', account['account__name'], account['total']]
        if include_percentages:
            row.append(f"{account.get('percentage', 0):.2f}%")
        writer.writerow(row)
    
    if include_totals:
        writer.writerow(['', 'Total Expenses', report_data['expenses']['total']])
        writer.writerow(['', 'Operating Profit', report_data['operating_profit']])
        writer.writerow([])
    
    # Write other income/expenses
    if report_data['other_income']['accounts']:
        if include_headers:
            writer.writerow([report_data['other_income']['title'], '', ''])
        
        for account in report_data['other_income']['accounts']:
            row = ['', account['account__name'], account['total']]
            if include_percentages:
                row.append(f"{account.get('percentage', 0):.2f}%")
            writer.writerow(row)
        
        if include_totals:
            writer.writerow(['', 'Total Other Income', report_data['other_income']['total']])
    
    if report_data['other_expenses']['accounts']:
        if include_headers:
            writer.writerow([report_data['other_expenses']['title'], '', ''])
        
        for account in report_data['other_expenses']['accounts']:
            row = ['', account['account__name'], account['total']]
            if include_percentages:
                row.append(f"{account.get('percentage', 0):.2f}%")
            writer.writerow(row)
        
        if include_totals:
            writer.writerow(['', 'Total Other Expenses', report_data['other_expenses']['total']])
    
    # Write final net profit
    if include_totals:
        writer.writerow([])
        writer.writerow(['', 'NET PROFIT', report_data['net_profit']])
    
    return response


def export_to_excel(report_data, include_headers, include_totals,
                   include_comparison, include_percentages, from_date, to_date, company=None):
    """Export Profit & Loss data to Excel format with professional formatting"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side, NamedStyle
        from openpyxl.utils import get_column_letter
        from openpyxl.worksheet.page import PageMargins
        from datetime import datetime
    except ImportError:
        raise ImportError('openpyxl is required for Excel export')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Profit & Loss Statement"
    
    # Set page margins and orientation
    ws.page_margins = PageMargins(left=0.7, right=0.7, top=1, bottom=1, header=0.3, footer=0.3)
    ws.page_setup.orientation = ws.ORIENTATION_PORTRAIT
    ws.page_setup.paperSize = ws.PAPERSIZE_A4
    
    # Professional styles
    company_font = Font(bold=True, size=16, color="2d3748")
    title_font = Font(bold=True, size=14, color="2d3748")
    subtitle_font = Font(bold=True, size=12, color="4a5568")
    header_font = Font(bold=True, size=11, color="ffffff")
    header_fill = PatternFill(start_color="2d3748", end_color="2d3748", fill_type="solid")
    section_font = Font(bold=True, size=11, color="2d3748")
    section_fill = PatternFill(start_color="f7fafc", end_color="f7fafc", fill_type="solid")
    total_font = Font(bold=True, size=10, color="2d3748")
    total_fill = PatternFill(start_color="edf2f7", end_color="edf2f7", fill_type="solid")
    net_profit_font = Font(bold=True, size=12, color="2d3748")
    net_profit_fill = PatternFill(start_color="edf2f7", end_color="edf2f7", fill_type="solid")
    
    # Border styles
    thin_border = Border(
        left=Side(style='thin', color='e2e8f0'),
        right=Side(style='thin', color='e2e8f0'),
        top=Side(style='thin', color='e2e8f0'),
        bottom=Side(style='thin', color='e2e8f0')
    )
    
    thick_border = Border(
        left=Side(style='medium', color='2d3748'),
        right=Side(style='medium', color='2d3748'),
        top=Side(style='medium', color='2d3748'),
        bottom=Side(style='medium', color='2d3748')
    )
    
    current_row = 1
    
    # Company header (if company provided)
    if company:
        # Company name
        ws.merge_cells(f'A{current_row}:D{current_row}')
        company_cell = ws[f'A{current_row}']
        company_cell.value = company.name
        company_cell.font = company_font
        company_cell.alignment = Alignment(horizontal='center')
        current_row += 1
        
        # Company address
        if company.address:
            ws.merge_cells(f'A{current_row}:D{current_row}')
            address_cell = ws[f'A{current_row}']
            address_cell.value = company.address
            address_cell.font = Font(size=10, color="4a5568")
            address_cell.alignment = Alignment(horizontal='center')
            current_row += 1
        
        # Contact info
        contact_info = []
        if hasattr(company, 'phone') and company.phone:
            contact_info.append(f"Tel: {company.phone}")
        if hasattr(company, 'email') and company.email:
            contact_info.append(f"Email: {company.email}")
        
        if contact_info:
            ws.merge_cells(f'A{current_row}:D{current_row}')
            contact_cell = ws[f'A{current_row}']
            contact_cell.value = " | ".join(contact_info)
            contact_cell.font = Font(size=10, color="4a5568")
            contact_cell.alignment = Alignment(horizontal='center')
            current_row += 1
        
        current_row += 1  # Extra space
    
    # Statement title
    ws.merge_cells(f'A{current_row}:D{current_row}')
    title_cell = ws[f'A{current_row}']
    title_cell.value = "PROFIT & LOSS STATEMENT"
    title_cell.font = title_font
    title_cell.alignment = Alignment(horizontal='center')
    current_row += 1
    
    # Period
    ws.merge_cells(f'A{current_row}:D{current_row}')
    period_cell = ws[f'A{current_row}']
    period_cell.value = f"For the period from {from_date} to {to_date}"
    period_cell.font = subtitle_font
    period_cell.alignment = Alignment(horizontal='center')
    current_row += 1
    
    # Generation timestamp
    ws.merge_cells(f'A{current_row}:D{current_row}')
    timestamp_cell = ws[f'A{current_row}']
    timestamp_cell.value = f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}"
    timestamp_cell.font = Font(size=9, color="718096")
    timestamp_cell.alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Define columns and helper functions
    col_account = 1
    col_amount = 2
    col_percentage = 3 if include_percentages else None
    
    def format_amount(amount):
        """Format amount with proper currency formatting"""
        if amount is None:
            return "0.00"
        try:
            amount = float(amount)
            if amount < 0:
                return f"({abs(amount):,.2f})"
            return f"{amount:,.2f}"
        except (ValueError, TypeError):
            return "0.00"
    
    def add_section_header(row, title):
        """Add a section header with proper styling"""
        ws.merge_cells(f'A{row}:B{row}')
        section_cell = ws[f'A{row}']
        section_cell.value = title
        section_cell.font = section_font
        section_cell.fill = section_fill
        section_cell.alignment = Alignment(horizontal='left')
        section_cell.border = thin_border
        
        # Add border to amount column
        amount_cell = ws[f'B{row}']
        amount_cell.border = thin_border
        
        if include_percentages:
            perc_cell = ws[f'C{row}']
            perc_cell.border = thin_border
        
        return row + 1
    
    def add_account_row(row, account_name, amount, percentage=None, indent=True):
        """Add an account row with proper formatting"""
        # Account name with indentation
        account_cell = ws[f'A{row}']
        account_cell.value = f"    {account_name}" if indent else account_name
        account_cell.font = Font(size=10)
        account_cell.alignment = Alignment(horizontal='left')
        account_cell.border = thin_border
        
        # Amount
        amount_cell = ws[f'B{row}']
        amount_cell.value = format_amount(amount)
        amount_cell.font = Font(size=10)
        amount_cell.alignment = Alignment(horizontal='right')
        amount_cell.border = thin_border
        
        # Percentage if included
        if include_percentages:
            perc_cell = ws[f'C{row}']
            if percentage is not None:
                perc_cell.value = f"{percentage:.1f}%"
            perc_cell.font = Font(size=10)
            perc_cell.alignment = Alignment(horizontal='right')
            perc_cell.border = thin_border
        
        return row + 1
    
    def add_total_row(row, title, amount, is_net_profit=False):
        """Add a total row with proper styling"""
        # Title
        title_cell = ws[f'A{row}']
        title_cell.value = title
        title_cell.font = net_profit_font if is_net_profit else total_font
        title_cell.fill = net_profit_fill if is_net_profit else total_fill
        title_cell.alignment = Alignment(horizontal='left')
        title_cell.border = thick_border if is_net_profit else thin_border
        
        # Amount
        amount_cell = ws[f'B{row}']
        amount_cell.value = format_amount(amount)
        amount_cell.font = net_profit_font if is_net_profit else total_font
        amount_cell.fill = net_profit_fill if is_net_profit else total_fill
        amount_cell.alignment = Alignment(horizontal='right')
        amount_cell.border = thick_border if is_net_profit else thin_border
        
        # Percentage column if included
        if include_percentages:
            perc_cell = ws[f'C{row}']
            perc_cell.fill = net_profit_fill if is_net_profit else total_fill
            perc_cell.border = thick_border if is_net_profit else thin_border
        
        return row + 1
    
    def add_spacing_row(row):
        """Add an empty spacing row"""
        return row + 1
    
    # Add table headers
    headers = ['PARTICULARS', 'AMOUNT (AED)']
    if include_percentages:
        headers.append('PERCENTAGE')
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center')
        cell.border = thick_border
    
    current_row += 1
    
    # Revenue section
    current_row = add_section_header(current_row, "REVENUE")
    
    for account in report_data['revenue']['accounts']:
        current_row = add_account_row(
            current_row, 
            account['account__name'], 
            account['total'],
            account.get('percentage') if include_percentages else None
        )
    
    if include_totals and report_data['revenue']['accounts']:
        current_row = add_total_row(current_row, "Total Revenue", report_data['revenue']['total'])
        current_row = add_spacing_row(current_row)
    
    # COGS section
    if report_data['cogs']['accounts']:
        current_row = add_section_header(current_row, "COST OF GOODS SOLD")
        
        for account in report_data['cogs']['accounts']:
            current_row = add_account_row(
                current_row, 
                account['account__name'], 
                account['total'],
                account.get('percentage') if include_percentages else None
            )
        
        if include_totals:
            current_row = add_total_row(current_row, "Total Cost of Goods Sold", report_data['cogs']['total'])
            current_row = add_spacing_row(current_row)
    
    # Gross Profit
    if include_totals:
        current_row = add_total_row(current_row, "GROSS PROFIT", report_data['gross_profit'])
        current_row = add_spacing_row(current_row)
    
    # Operating Expenses section
    if report_data['expenses']['accounts']:
        current_row = add_section_header(current_row, "OPERATING EXPENSES")
        
        for account in report_data['expenses']['accounts']:
            current_row = add_account_row(
                current_row, 
                account['account__name'], 
                account['total'],
                account.get('percentage') if include_percentages else None
            )
        
        if include_totals:
            current_row = add_total_row(current_row, "Total Operating Expenses", report_data['expenses']['total'])
            current_row = add_spacing_row(current_row)
    
    # Operating Profit
    if include_totals:
        current_row = add_total_row(current_row, "OPERATING PROFIT", report_data['operating_profit'])
        current_row = add_spacing_row(current_row)
    
    # Other Income section
    if report_data['other_income']['accounts']:
        current_row = add_section_header(current_row, "OTHER INCOME")
        
        for account in report_data['other_income']['accounts']:
            current_row = add_account_row(
                current_row, 
                account['account__name'], 
                account['total'],
                account.get('percentage') if include_percentages else None
            )
        
        if include_totals:
            current_row = add_total_row(current_row, "Total Other Income", report_data['other_income']['total'])
            current_row = add_spacing_row(current_row)
    
    # Other Expenses section
    if report_data['other_expenses']['accounts']:
        current_row = add_section_header(current_row, "OTHER EXPENSES")
        
        for account in report_data['other_expenses']['accounts']:
            current_row = add_account_row(
                current_row, 
                account['account__name'], 
                account['total'],
                account.get('percentage') if include_percentages else None
            )
        
        if include_totals:
            current_row = add_total_row(current_row, "Total Other Expenses", report_data['other_expenses']['total'])
            current_row = add_spacing_row(current_row)
    
    # Net Profit - Final result
    if include_totals:
        current_row = add_total_row(current_row, "NET PROFIT", report_data['net_profit'], is_net_profit=True)
    
    # Set optimal column widths
    ws.column_dimensions['A'].width = 35  # Account names
    ws.column_dimensions['B'].width = 15  # Amounts
    if include_percentages:
        ws.column_dimensions['C'].width = 12  # Percentages
    
    # Add signature section
    current_row += 3
    
    # Signature headers
    signature_headers = ['Prepared by:', 'Reviewed by:', 'Approved by:']
    for i, header in enumerate(signature_headers):
        col = 1 + (i * 2)  # Columns A, C, E
        if col <= 5:  # Ensure we don't exceed reasonable column limits
            header_cell = ws.cell(row=current_row, column=col, value=header)
            header_cell.font = Font(bold=True, size=10)
            header_cell.alignment = Alignment(horizontal='center')
    
    current_row += 3
    
    # Signature lines
    for i in range(3):
        col = 1 + (i * 2)  # Columns A, C, E
        if col <= 5:
            line_cell = ws.cell(row=current_row, column=col, value="_________________")
            line_cell.font = Font(size=10)
            line_cell.alignment = Alignment(horizontal='center')
    
    current_row += 1
    
    # Name & Signature labels
    for i in range(3):
        col = 1 + (i * 2)  # Columns A, C, E
        if col <= 5:
            label_cell = ws.cell(row=current_row, column=col, value="Name & Signature")
            label_cell.font = Font(size=9, color="718096")
            label_cell.alignment = Alignment(horizontal='center')
    
    current_row += 2
    
    # Date lines
    for i in range(3):
        col = 1 + (i * 2)  # Columns A, C, E
        if col <= 5:
            date_cell = ws.cell(row=current_row, column=col, value="Date: ___________")
            date_cell.font = Font(size=9, color="718096")
            date_cell.alignment = Alignment(horizontal='center')
    
    # Set print area and page setup
    ws.print_area = f'A1:C{current_row}'
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    wb.save(response)
    
    return response


def export_to_pdf(report_data, include_headers, include_totals,
                 include_comparison, include_percentages, from_date, to_date, company=None):
    """Export Profit & Loss data to PDF format with professional formatting"""
    
    try:
        from reportlab.lib.pagesizes import letter, A4
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.lib import colors
        from reportlab.platypus.tableofcontents import TableOfContents
        from datetime import datetime
    except ImportError:
        raise ImportError('reportlab is required for PDF export')
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f"Profit_Loss_Statement_{from_date}_to_{to_date}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Create PDF document with margins
    doc = SimpleDocTemplate(
        response, 
        pagesize=A4,
        rightMargin=2*cm,
        leftMargin=2*cm,
        topMargin=2*cm,
        bottomMargin=2*cm
    )
    elements = []
    
    # Professional Styles
    styles = getSampleStyleSheet()
    
    # Company header style
    company_style = ParagraphStyle(
        'CompanyHeader',
        parent=styles['Normal'],
        fontSize=18,
        spaceAfter=5,
        alignment=0,  # Left alignment
        textColor=colors.HexColor('#1a365d'),
        fontName='Helvetica-Bold'
    )
    
    # Company details style
    company_details_style = ParagraphStyle(
        'CompanyDetails',
        parent=styles['Normal'],
        fontSize=10,
        spaceAfter=3,
        alignment=0,  # Left alignment
        textColor=colors.HexColor('#4a5568')
    )
    
    # Statement title style
    title_style = ParagraphStyle(
        'StatementTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=10,
        spaceBefore=20,
        alignment=2,  # Right alignment
        textColor=colors.HexColor('#2d3748'),
        fontName='Helvetica-Bold'
    )
    
    # Period style
    period_style = ParagraphStyle(
        'Period',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=30,
        alignment=2,  # Right alignment
        textColor=colors.HexColor('#4a5568'),
        fontName='Helvetica-Bold'
    )
    
    # Build header sections (Left: Company Details, Right: Title & Period)
    header_left = []
    
    # If no company is provided, try to get the first company from database
    if not company:
        try:
            company = Company.objects.first()
        except:
            company = None
    
    if company:
        header_left.append(Paragraph(f"<b>{company.name.upper()}</b>", company_style))
        if hasattr(company, 'address') and company.address:
            header_left.append(Paragraph(company.address, company_details_style))
        contact_info = []
        if hasattr(company, 'phone') and company.phone:
            contact_info.append(f"Tel: {company.phone}")
        if hasattr(company, 'email') and company.email:
            contact_info.append(f"Email: {company.email}")
        if contact_info:
            header_left.append(Paragraph(" | ".join(contact_info), company_details_style))
    
    header_right = [
        Paragraph("PROFIT & LOSS STATEMENT", title_style),
        Paragraph(f"For the period from {from_date} to {to_date}", period_style),
    ]
    
    header_table = Table(
        [[header_left, header_right]],
        colWidths=[doc.width * 0.6, doc.width * 0.4]
    )
    header_table.setStyle(TableStyle([
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('ALIGN', (0, 0), (0, 0), 'LEFT'),
        ('ALIGN', (1, 0), (1, 0), 'RIGHT'),
        ('LEFTPADDING', (0, 0), (-1, -1), 0),
        ('RIGHTPADDING', (0, 0), (-1, -1), 0),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
    ]))
    elements.append(header_table)
    elements.append(Spacer(1, 12))
    
    # Add generation timestamp
    timestamp_style = ParagraphStyle(
        'Timestamp',
        parent=styles['Normal'],
        fontSize=8,
        spaceAfter=20,
        alignment=1,
        textColor=colors.HexColor('#718096')
    )
    timestamp = Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", timestamp_style)
    elements.append(timestamp)
    
    # Prepare professional table data
    table_data = []
    
    # Add professional headers
    if include_headers:
        headers = ['PARTICULARS', 'AMOUNT (AED)']
        if include_percentages:
            headers.append('%')
        table_data.append(headers)
    
    # Helper function to format currency
    def format_amount(amount):
        return f"{amount:,.2f}" if amount != 0 else "0.00"
    
    # Add revenue section
    table_data.append(['REVENUE', '', ''])
    
    for account in report_data['revenue']['accounts']:
        row = [f"    {account['account__name']}", format_amount(account['total'])]
        if include_percentages:
            row.append(f"{account.get('percentage', 0):.2f}")
        table_data.append(row)
    
    if include_totals:
        table_data.append(['Total Revenue', format_amount(report_data['revenue']['total']), ''])
        table_data.append(['', '', ''])  # Spacing row
    
    # Add COGS section
    if report_data['cogs']['accounts']:
        table_data.append(['COST OF GOODS SOLD', '', ''])
        
        for account in report_data['cogs']['accounts']:
            row = [f"    {account['account__name']}", f"({format_amount(account['total'])})"]
            if include_percentages:
                row.append(f"{account.get('percentage', 0):.2f}")
            table_data.append(row)
        
        if include_totals:
            table_data.append(['Total Cost of Goods Sold', f"({format_amount(report_data['cogs']['total'])})", ''])
            table_data.append(['', '', ''])  # Spacing row
    
    # Add Gross Profit
    if include_totals:
        table_data.append(['GROSS PROFIT', format_amount(report_data['gross_profit']), ''])
        table_data.append(['', '', ''])  # Spacing row
    
    # Add operating expenses section
    if report_data['expenses']['accounts']:
        table_data.append(['OPERATING EXPENSES', '', ''])
        
        for account in report_data['expenses']['accounts']:
            row = [f"    {account['account__name']}", f"({format_amount(account['total'])})"]
            if include_percentages:
                row.append(f"{account.get('percentage', 0):.2f}")
            table_data.append(row)
        
        if include_totals:
            table_data.append(['Total Operating Expenses', f"({format_amount(report_data['expenses']['total'])})", ''])
            table_data.append(['', '', ''])  # Spacing row
    
    # Add Operating Profit
    if include_totals:
        table_data.append(['OPERATING PROFIT', format_amount(report_data['operating_profit']), ''])
        table_data.append(['', '', ''])  # Spacing row
    
    # Add other income section
    if report_data['other_income']['accounts']:
        table_data.append(['OTHER INCOME', '', ''])
        
        for account in report_data['other_income']['accounts']:
            row = [f"    {account['account__name']}", format_amount(account['total'])]
            if include_percentages:
                row.append(f"{account.get('percentage', 0):.2f}")
            table_data.append(row)
        
        if include_totals:
            table_data.append(['Total Other Income', format_amount(report_data['other_income']['total']), ''])
            table_data.append(['', '', ''])  # Spacing row
    
    # Add other expenses section
    if report_data['other_expenses']['accounts']:
        table_data.append(['OTHER EXPENSES', '', ''])
        
        for account in report_data['other_expenses']['accounts']:
            row = [f"    {account['account__name']}", f"({format_amount(account['total'])})"]
            if include_percentages:
                row.append(f"{account.get('percentage', 0):.2f}")
            table_data.append(row)
        
        if include_totals:
            table_data.append(['Total Other Expenses', f"({format_amount(report_data['other_expenses']['total'])})", ''])
            table_data.append(['', '', ''])  # Spacing row
    
    # Add final net profit with emphasis
    if include_totals:
        table_data.append(['NET PROFIT', format_amount(report_data['net_profit']), ''])
    
    # Create professional table
    if table_data:
        # Set column widths for better formatting
        col_widths = [4*inch, 1.5*inch]
        if include_percentages:
            col_widths.append(0.8*inch)
        
        table = Table(table_data, colWidths=col_widths, repeatRows=1)
        
        # Professional table style
        style_commands = [
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('TOPPADDING', (0, 0), (-1, 0), 12),
            
            # General table styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
            ('ALIGN', (1, 1), (-1, -1), 'RIGHT'),  # Right align amounts
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('LEFTPADDING', (0, 0), (-1, -1), 8),
            ('RIGHTPADDING', (0, 0), (-1, -1), 8),
            ('TOPPADDING', (0, 1), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 6),
            
            # Border styling
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2d3748')),
            ('LINEBELOW', (0, -1), (-1, -1), 2, colors.HexColor('#2d3748')),
        ]
        
        # Add section header styling
        section_rows = []
        total_rows = []
        net_profit_row = None
        
        for i, row in enumerate(table_data):
            if i == 0:  # Skip header row
                continue
            
            # Identify section headers (all caps, no indentation)
            if row[0] and not row[0].startswith('    ') and row[0].isupper() and row[0] not in ['GROSS PROFIT', 'OPERATING PROFIT', 'NET PROFIT']:
                section_rows.append(i)
            
            # Identify total rows
            if row[0] and ('Total' in row[0] or row[0] in ['GROSS PROFIT', 'OPERATING PROFIT']):
                total_rows.append(i)
            
            # Identify net profit row
            if row[0] == 'NET PROFIT':
                net_profit_row = i
        
        # Style section headers
        for row_idx in section_rows:
            style_commands.extend([
                ('BACKGROUND', (0, row_idx), (-1, row_idx), colors.HexColor('#f7fafc')),
                ('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'),
                ('FONTSIZE', (0, row_idx), (-1, row_idx), 11),
                ('LINEABOVE', (0, row_idx), (-1, row_idx), 1, colors.HexColor('#e2e8f0')),
                ('LINEBELOW', (0, row_idx), (-1, row_idx), 1, colors.HexColor('#e2e8f0')),
            ])
        
        # Style total rows
        for row_idx in total_rows:
            style_commands.extend([
                ('FONTNAME', (0, row_idx), (-1, row_idx), 'Helvetica-Bold'),
                ('LINEABOVE', (0, row_idx), (-1, row_idx), 1, colors.HexColor('#4a5568')),
                ('LINEBELOW', (0, row_idx), (-1, row_idx), 1, colors.HexColor('#4a5568')),
            ])
        
        # Style net profit row with emphasis
        if net_profit_row:
            style_commands.extend([
                ('BACKGROUND', (0, net_profit_row), (-1, net_profit_row), colors.HexColor('#edf2f7')),
                ('FONTNAME', (0, net_profit_row), (-1, net_profit_row), 'Helvetica-Bold'),
                ('FONTSIZE', (0, net_profit_row), (-1, net_profit_row), 12),
                ('LINEABOVE', (0, net_profit_row), (-1, net_profit_row), 2, colors.HexColor('#2d3748')),
                ('LINEBELOW', (0, net_profit_row), (-1, net_profit_row), 2, colors.HexColor('#2d3748')),
            ])
        
        # Apply all styles
        table.setStyle(TableStyle(style_commands))
        elements.append(table)
        
        # Add footer with signature lines
        elements.append(Spacer(1, 40))
        
        # Signature section
        signature_style = ParagraphStyle(
            'Signature',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=30,
            textColor=colors.HexColor('#4a5568')
        )
        
        signature_table_data = [
            ['Prepared by:', '', 'Reviewed by:', '', 'Approved by:'],
            ['', '', '', '', ''],
            ['_________________', '', '_________________', '', '_________________'],
            ['Name & Signature', '', 'Name & Signature', '', 'Name & Signature'],
            ['', '', '', '', ''],
            ['Date: ___________', '', 'Date: ___________', '', 'Date: ___________']
        ]
        
        signature_table = Table(signature_table_data, colWidths=[1.5*inch, 0.5*inch, 1.5*inch, 0.5*inch, 1.5*inch])
        signature_table.setStyle(TableStyle([
            ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 0), (-1, -1), 9),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
        ]))
        
        elements.append(signature_table)
    
    # Build PDF
    doc.build(elements)
    
    return response


@login_required
def report_list(request):
    """List of saved Profit & Loss reports"""
    
    reports = ProfitLossReport.objects.filter(created_by=request.user).order_by('-created_at')
    
    context = {
        'reports': reports,
    }
    
    return render(request, 'profit_loss_statement/report_list.html', context)


@login_required
def report_detail(request, report_id):
    """Detail view of a saved Profit & Loss report"""
    
    try:
        report = ProfitLossReport.objects.get(id=report_id, created_by=request.user)
    except ProfitLossReport.DoesNotExist:
        messages.error(request, 'Report not found')
        return redirect('profit_loss_statement:report_list')
    
    context = {
        'report': report,
        'report_data': report.report_data,
        'from_date': report.from_date,
        'to_date': report.to_date,
        'company': report.company,
    }
    
    return render(request, 'profit_loss_statement/report_detail.html', context)


@login_required
def delete_report(request, report_id):
    """Delete a saved Profit & Loss report"""
    
    try:
        report = ProfitLossReport.objects.get(id=report_id, created_by=request.user)
        report_title = report.title
        report.delete()
        messages.success(request, f'Report "{report_title}" has been deleted successfully.')
    except ProfitLossReport.DoesNotExist:
        messages.error(request, 'Report not found or you do not have permission to delete it.')
    
    return redirect('profit_loss_statement:report_list')