from django.shortcuts import render, get_object_or_404, redirect
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.views.decorators.http import require_GET, require_POST
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.core.paginator import Paginator
from django.forms import formset_factory
from django.db import transaction
from django.template.loader import get_template
from .models import PettyCashDay, PettyCashEntry, PettyCashBalance, PettyCashAudit
from .forms import PettyCashDayForm, PettyCashEntryForm, PettyCashEntryFormSet, PettyCashFilterForm, QuickEntryForm
from multi_currency.models import Currency
from general_journal.models import JournalEntry, JournalEntryLine
from chart_of_accounts.models import ChartOfAccount
from ledger.models import Ledger
from decimal import Decimal
from django.db.models import Q, Sum
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from io import BytesIO
import datetime

try:
    from weasyprint import HTML, CSS
    WEASYPRINT_AVAILABLE = True
except ImportError:
    WEASYPRINT_AVAILABLE = False


def get_petty_cash_ledger_balance():
    """Calculate actual petty cash balance from ledger entries"""
    # Get the main petty cash account (1000)
    try:
        petty_cash_account = ChartOfAccount.objects.get(account_code='1000')
        
        # Calculate balance for the main petty cash account (Debits - Credits for asset accounts)
        debits = Ledger.objects.filter(
            account=petty_cash_account,
            status='POSTED',
            entry_type='DR'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        credits = Ledger.objects.filter(
            account=petty_cash_account,
            status='POSTED',
            entry_type='CR'
        ).aggregate(total=Sum('amount'))['total'] or Decimal('0.00')
        
        balance = debits - credits
        return balance
        
    except ChartOfAccount.DoesNotExist:
        return Decimal('0.00')

@login_required
def petty_cash_register(request):
    """Main petty cash register view with day selection and entries"""
    
    # Get selected date (default to today)
    selected_date = request.GET.get('date')
    if selected_date:
        try:
            selected_date = timezone.datetime.strptime(selected_date, '%Y-%m-%d').date()
        except:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Calculate actual petty cash balance from ledger for reference
    actual_ledger_balance = get_petty_cash_ledger_balance()
    
    # Get previous day's closing balance to use as opening balance
    previous_day_balance = Decimal('0.00')
    try:
        previous_day = PettyCashDay.objects.filter(
            entry_date__lt=selected_date
        ).order_by('-entry_date').first()
        if previous_day:
            previous_day_balance = Decimal(str(previous_day.closing_balance))
    except:
        previous_day_balance = Decimal('0.00')
    
    # Use previous day's closing balance as opening balance for current day
    previous_balance = previous_day_balance
    
    # Get or create petty cash day
    petty_cash_day, created = PettyCashDay.objects.get_or_create(
        entry_date=selected_date,
        defaults={
            'created_by': request.user,
            'opening_balance': previous_balance
        }
    )
    
    if request.method == 'POST':
        # Handle form submission
        day_form = PettyCashDayForm(request.POST, instance=petty_cash_day)
        entry_formset = PettyCashEntryFormSet(request.POST, request.FILES, instance=petty_cash_day)
        
        if day_form.is_valid() and entry_formset.is_valid():
            with transaction.atomic():
                # Save day form
                day = day_form.save(commit=False)
                day.updated_by = request.user
                day.save()
                
                # Save entries
                entries = entry_formset.save(commit=False)
                for entry in entries:
                    entry.created_by = request.user
                    entry.save()
                
                # Delete marked entries
                for obj in entry_formset.deleted_objects:
                    obj.delete()
                
                # Create audit trail
                PettyCashAudit.objects.create(
                    petty_cash_day=day,
                    action='updated',
                    description=f'Petty cash entries updated for {day.entry_date}',
                    user=request.user
                )
                
                # Handle AJAX requests
                if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                    return JsonResponse({
                        'success': True,
                        'message': f'Petty cash entries for {day.entry_date} saved successfully.',
                        'redirect_url': f'/accounting/petty-cash/?date={day.entry_date.strftime("%Y-%m-%d")}'
                    })
                
                messages.success(request, f'Petty cash entries for {day.entry_date} saved successfully.')
                return redirect(f'/accounting/petty-cash/?date={day.entry_date.strftime("%Y-%m-%d")}')
        else:
            # Handle AJAX validation errors
            if request.headers.get('X-Requested-With') == 'XMLHttpRequest':
                # Debug logging
                print("=== PETTY CASH 400 ERROR DEBUG ===")
                print(f"POST data: {dict(request.POST)}")
                print(f"Day form errors: {day_form.errors}")
                print(f"Entry formset errors: {entry_formset.errors}")
                print(f"Entry formset non-form errors: {entry_formset.non_form_errors()}")
                
                errors = {}
                if day_form.errors:
                    errors.update(day_form.errors)
                if entry_formset.errors:
                    for i, form_errors in enumerate(entry_formset.errors):
                        if form_errors:
                            errors[f'entry_{i}'] = form_errors
                if entry_formset.non_form_errors():
                    errors['non_form_errors'] = entry_formset.non_form_errors()
                
                print(f"Final errors dict: {errors}")
                print("=== END DEBUG ===")
                
                return JsonResponse({
                    'success': False,
                    'message': 'Please correct the errors below.',
                    'errors': errors
                }, status=400)
            
            messages.error(request, 'Please correct the errors below.')
    else:
        day_form = PettyCashDayForm(instance=petty_cash_day)
        entry_formset = PettyCashEntryFormSet(instance=petty_cash_day)
    
    # Get recent days for navigation
    recent_days = PettyCashDay.objects.filter(
        entry_date__lt=selected_date
    ).order_by('-entry_date')[:10]
    
    # Get next day if exists
    next_day = PettyCashDay.objects.filter(
        entry_date__gt=selected_date
    ).order_by('entry_date').first()
    
    # Get journal entries that affect petty cash account for the selected date
    petty_cash_account_codes = ['1000']  # Only main Petty Cash account
    petty_cash_accounts = ChartOfAccount.objects.filter(account_code__in=petty_cash_account_codes)
    
    journal_entries = JournalEntry.objects.filter(
        date=selected_date,
        status='posted',
        lines__account__in=petty_cash_accounts
    ).distinct().prefetch_related('lines__account')
    
    # Process journal entries to get petty cash related lines
    journal_entries_data = []
    for journal_entry in journal_entries:
        petty_cash_lines = journal_entry.lines.filter(account__in=petty_cash_accounts)
        for line in petty_cash_lines:
            journal_entries_data.append({
                'journal_number': journal_entry.journal_number,
                'description': journal_entry.description,
                'account': line.account,
                'debit_amount': line.debit_amount,
                'credit_amount': line.credit_amount,
                'date': journal_entry.date,
                'created_by': journal_entry.created_by,
            })
    
    context = {
        'petty_cash_day': petty_cash_day,
        'day_form': day_form,
        'entry_formset': entry_formset,
        'selected_date': selected_date,
        'previous_balance': previous_balance,
        'previous_day_balance': previous_day_balance,
        'recent_days': recent_days,
        'next_day': next_day,
        'journal_entries': journal_entries_data,
        'actual_ledger_balance': actual_ledger_balance,
    }
    
    return render(request, 'petty_cash/register.html', context)


@login_required
def petty_cash_list(request):
    """List all petty cash days with filtering and pagination"""
    
    # Get filter parameters
    form = PettyCashFilterForm(request.GET)
    days = PettyCashDay.objects.all()
    
    # Apply filters
    if form.is_valid():
        date_from = form.cleaned_data.get('date_from')
        date_to = form.cleaned_data.get('date_to')
        status = form.cleaned_data.get('status')
        paid_by = form.cleaned_data.get('paid_by')
        search = form.cleaned_data.get('search')
        
        if date_from:
            days = days.filter(entry_date__gte=date_from)
        if date_to:
            days = days.filter(entry_date__lte=date_to)
        if status:
            days = days.filter(status=status)
        if paid_by:
            days = days.filter(entries__paid_by__icontains=paid_by).distinct()
        if search:
            days = days.filter(
                Q(entries__description__icontains=search) |
                Q(notes__icontains=search)
            ).distinct()
    
    # Pagination
    paginator = Paginator(days, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Summary statistics
    total_days = days.count()
    total_expenses = days.aggregate(total=Sum('total_expenses'))['total'] or 0
    avg_daily_expense = total_expenses / total_days if total_days > 0 else 0
    
    context = {
        'page_obj': page_obj,
        'form': form,
        'total_days': total_days,
        'total_expenses': total_expenses,
        'avg_daily_expense': avg_daily_expense,
    }
    
    return render(request, 'petty_cash/list.html', context)


@login_required
def petty_cash_detail(request, pk):
    """View petty cash day details"""
    
    petty_cash_day = get_object_or_404(PettyCashDay, pk=pk)
    
    context = {
        'petty_cash_day': petty_cash_day,
        'audit_trail': petty_cash_day.audit_trail.select_related('user').all()[:10],
    }
    
    return render(request, 'petty_cash/detail.html', context)


@login_required
def petty_cash_create(request):
    """Create a new petty cash day"""
    
    if request.method == 'POST':
        form = PettyCashDayForm(request.POST)
        
        if form.is_valid():
            petty_cash_day = form.save(commit=False)
            petty_cash_day.created_by = request.user
            petty_cash_day.save()
            
            # Create audit trail
            PettyCashAudit.objects.create(
                petty_cash_day=petty_cash_day,
                action='created',
                description=f'Petty cash day created for {petty_cash_day.entry_date}',
                user=request.user
            )
            
            messages.success(request, f'Petty cash day for {petty_cash_day.entry_date} created successfully.')
            return redirect('petty_cash:register')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PettyCashDayForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'petty_cash/create.html', context)


@login_required
def petty_cash_edit(request, pk):
    """Edit petty cash day"""
    
    petty_cash_day = get_object_or_404(PettyCashDay, pk=pk)
    
    if not petty_cash_day.can_edit:
        messages.error(request, 'This petty cash day cannot be edited.')
        return redirect('petty_cash:detail', pk=pk)
    
    if request.method == 'POST':
        form = PettyCashDayForm(request.POST, instance=petty_cash_day)
        
        if form.is_valid():
            petty_cash_day = form.save(commit=False)
            petty_cash_day.updated_by = request.user
            petty_cash_day.save()
            
            # Create audit trail
            PettyCashAudit.objects.create(
                petty_cash_day=petty_cash_day,
                action='updated',
                description=f'Petty cash day updated for {petty_cash_day.entry_date}',
                user=request.user
            )
            
            messages.success(request, f'Petty cash day for {petty_cash_day.entry_date} updated successfully.')
            return redirect('petty_cash:detail', pk=pk)
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = PettyCashDayForm(instance=petty_cash_day)
    
    context = {
        'petty_cash_day': petty_cash_day,
        'form': form,
    }
    
    return render(request, 'petty_cash/edit.html', context)


@login_required
@require_POST
def petty_cash_submit(request, pk):
    """Submit petty cash day for approval"""
    
    petty_cash_day = get_object_or_404(PettyCashDay, pk=pk)
    
    if petty_cash_day.status != 'draft':
        return JsonResponse({'success': False, 'message': 'Only draft entries can be submitted.'})
    
    petty_cash_day.status = 'submitted'
    petty_cash_day.updated_by = request.user
    petty_cash_day.save()
    
    # Create audit trail
    PettyCashAudit.objects.create(
        petty_cash_day=petty_cash_day,
        action='submitted',
        description='Petty cash day submitted for approval',
        user=request.user
    )
    
    messages.success(request, f'Petty cash day for {petty_cash_day.entry_date} submitted for approval.')
    return JsonResponse({'success': True, 'message': 'Petty cash day submitted successfully.'})


@login_required
@require_POST
def petty_cash_approve(request, pk):
    """Approve petty cash day"""
    
    petty_cash_day = get_object_or_404(PettyCashDay, pk=pk)
    
    if petty_cash_day.status != 'submitted':
        return JsonResponse({'success': False, 'message': 'Only submitted entries can be approved.'})
    
    petty_cash_day.status = 'approved'
    petty_cash_day.approved_by = request.user
    petty_cash_day.approved_at = timezone.now()
    petty_cash_day.save()
    
    # Create audit trail
    PettyCashAudit.objects.create(
        petty_cash_day=petty_cash_day,
        action='approved',
        description='Petty cash day approved',
        user=request.user
    )
    
    messages.success(request, f'Petty cash day for {petty_cash_day.entry_date} approved successfully.')
    return JsonResponse({'success': True, 'message': 'Petty cash day approved successfully.'})


@login_required
@require_POST
def petty_cash_lock(request, pk):
    """Lock petty cash day"""
    
    petty_cash_day = get_object_or_404(PettyCashDay, pk=pk)
    
    if petty_cash_day.status != 'approved':
        return JsonResponse({'success': False, 'message': 'Only approved entries can be locked.'})
    
    petty_cash_day.is_locked = True
    petty_cash_day.save()
    
    # Create audit trail
    PettyCashAudit.objects.create(
        petty_cash_day=petty_cash_day,
        action='locked',
        description='Petty cash day locked',
        user=request.user
    )
    
    messages.success(request, f'Petty cash day for {petty_cash_day.entry_date} locked successfully.')
    return JsonResponse({'success': True, 'message': 'Petty cash day locked successfully.'})


@login_required
def quick_entry(request):
    """Quick entry form for adding single entries"""
    
    if request.method == 'POST':
        form = QuickEntryForm(request.POST)
        
        if form.is_valid():
            # Get or create petty cash day for today
            today = timezone.now().date()
            petty_cash_day, created = PettyCashDay.objects.get_or_create(
                entry_date=today,
                defaults={
                    'created_by': request.user,
                    'opening_balance': Decimal('0.00')
                }
            )
            
            # Create entry
            entry = PettyCashEntry(
                petty_cash_day=petty_cash_day,
                description=form.cleaned_data['description'],
                amount=form.cleaned_data['amount'],
                paid_by=form.cleaned_data.get('paid_by', ''),
                notes=form.cleaned_data.get('notes', ''),
                created_by=request.user
            )
            entry.save()
            
            # Create audit trail
            PettyCashAudit.objects.create(
                petty_cash_day=petty_cash_day,
                action='entry_added',
                description=f'Quick entry added: {entry.description}',
                user=request.user,
                entry=entry
            )
            
            messages.success(request, f'Quick entry "{entry.description}" added successfully.')
            return redirect('petty_cash:register')
        else:
            messages.error(request, 'Please correct the errors below.')
    else:
        form = QuickEntryForm()
    
    context = {
        'form': form,
    }
    
    return render(request, 'petty_cash/quick_entry.html', context)


@login_required
@require_GET
def get_previous_balance(request):
    """AJAX endpoint to get previous day's closing balance"""
    
    date_str = request.GET.get('date')
    
    if not date_str:
        return JsonResponse({'error': 'Date is required'})
    
    try:
        date = timezone.datetime.strptime(date_str, '%Y-%m-%d').date()
        previous_day = PettyCashDay.objects.filter(
            entry_date__lt=date
        ).order_by('-entry_date').first()
        
        balance = previous_day.closing_balance if previous_day else Decimal('0.00')
        
        return JsonResponse({
            'balance': str(balance),
            'date': previous_day.entry_date.strftime('%Y-%m-%d') if previous_day else None
        })
    except Exception as e:
        return JsonResponse({'error': str(e)})


@login_required
@require_GET
def petty_cash_summary(request):
    """Get petty cash summary statistics"""
    
    # Get date range (default to last 30 days)
    days_back = int(request.GET.get('days', 30))
    end_date = timezone.now().date()
    start_date = end_date - timezone.timedelta(days=days_back)
    
    days = PettyCashDay.objects.filter(entry_date__range=[start_date, end_date])
    
    total_days = days.count()
    total_expenses = days.aggregate(total=Sum('total_expenses'))['total'] or 0
    avg_daily_expense = total_expenses / total_days if total_days > 0 else 0
    
    # Status breakdown
    status_stats = {}
    for status_code, status_name in PettyCashDay.STATUS_CHOICES:
        count = days.filter(status=status_code).count()
        amount = days.filter(status=status_code).aggregate(total=Sum('total_expenses'))['total'] or 0
        status_stats[status_name] = {
            'count': count,
            'amount': float(amount)
        }
    
    # Top expenses by description
    top_expenses = PettyCashEntry.objects.filter(
        petty_cash_day__entry_date__range=[start_date, end_date]
    ).values('description').annotate(
        total=Sum('amount'),
        count=Count('id')
    ).order_by('-total')[:10]
    
    data = {
        'period': {
            'start_date': start_date.strftime('%Y-%m-%d'),
            'end_date': end_date.strftime('%Y-%m-%d'),
            'days': days_back
        },
        'summary': {
            'total_days': total_days,
            'total_expenses': float(total_expenses),
            'avg_daily_expense': float(avg_daily_expense)
        },
        'status_stats': status_stats,
        'top_expenses': list(top_expenses)
    }
    
    return JsonResponse(data)


@login_required
def export_petty_cash_excel(request):
    """Export petty cash entries to Excel with professional formatting"""
    
    # Get selected date (default to today)
    selected_date = request.GET.get('date')
    if selected_date:
        try:
            selected_date = timezone.datetime.strptime(selected_date, '%Y-%m-%d').date()
        except:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get petty cash day
    try:
        petty_cash_day = PettyCashDay.objects.get(entry_date=selected_date)
        entries = petty_cash_day.entries.all().order_by('id')
    except PettyCashDay.DoesNotExist:
        entries = []
        petty_cash_day = None
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f"Petty Cash Register - {selected_date.strftime('%Y-%m-%d')}"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Title section
    ws.merge_cells('A1:G1')
    ws['A1'] = "PETTY CASH REGISTER"
    ws['A1'].font = Font(bold=True, size=16)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"Date: {selected_date.strftime('%B %d, %Y')}"
    ws['A2'].font = Font(bold=True, size=12)
    ws['A2'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A3:G3')
    ws['A3'] = "Petty Cash Ledger"
    ws['A3'].font = Font(bold=True, size=12)
    ws['A3'].alignment = Alignment(horizontal="center")
    
    # Opening balance
    opening_balance = petty_cash_day.opening_balance if petty_cash_day else Decimal('0.00')
    ws.merge_cells('A5:G5')
    ws['A5'] = f"Opening Balance: {opening_balance:,.2f} AED"
    ws['A5'].font = Font(bold=True, size=11)
    
    # Headers
    headers = ['Sr. No', 'Job No', 'Description', 'Debit', 'Credit', 'Balance', 'Notes']
    header_row = 7
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=header_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    running_balance = opening_balance
    row = header_row + 1
    
    for idx, entry in enumerate(entries, 1):
        # Calculate running balance
        running_balance -= entry.amount
        
        # Sr. No
        cell = ws.cell(row=row, column=1, value=idx)
        cell.alignment = Alignment(horizontal="center")
        cell.border = border
        
        # Job No
        cell = ws.cell(row=row, column=2, value=entry.job_no or '')
        cell.border = border
        
        # Description
        cell = ws.cell(row=row, column=3, value=entry.description)
        cell.border = border
        
        # Debit (Amount)
        cell = ws.cell(row=row, column=4, value=entry.amount)
        cell.number_format = '#,##0.00" AED"'
        cell.alignment = Alignment(horizontal="right")
        cell.border = border
        
        # Credit (empty for expenses)
        cell = ws.cell(row=row, column=5, value='')
        cell.border = border
        
        # Balance
        cell = ws.cell(row=row, column=6, value=running_balance)
        cell.number_format = '#,##0.00" AED"'
        cell.alignment = Alignment(horizontal="right")
        cell.border = border
        if running_balance < 0:
            cell.font = Font(color="FF0000")  # Red for negative
        else:
            cell.font = Font(color="008000")  # Green for positive
        
        # Notes
        cell = ws.cell(row=row, column=7, value=entry.notes or '')
        cell.border = border
        
        row += 1
    
    # Summary section
    if petty_cash_day:
        row += 1
        ws.merge_cells(f'A{row}:G{row}')
        ws[f'A{row}'] = "SUMMARY"
        ws[f'A{row}'].font = Font(bold=True, size=12)
        ws[f'A{row}'].alignment = Alignment(horizontal="center")
        
        row += 1
        ws[f'A{row}'] = "Total Expenses:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'D{row}'] = petty_cash_day.total_expenses
        ws[f'D{row}'].number_format = '#,##0.00" AED"'
        ws[f'D{row}'].font = Font(bold=True)
        ws[f'D{row}'].alignment = Alignment(horizontal="right")
        
        row += 1
        ws[f'A{row}'] = "Closing Balance:"
        ws[f'A{row}'].font = Font(bold=True)
        ws[f'F{row}'] = petty_cash_day.closing_balance
        ws[f'F{row}'].number_format = '#,##0.00" AED"'
        ws[f'F{row}'].font = Font(bold=True)
        ws[f'F{row}'].alignment = Alignment(horizontal="right")
        if petty_cash_day.closing_balance < 0:
            ws[f'F{row}'].font = Font(bold=True, color="FF0000")
        else:
            ws[f'F{row}'].font = Font(bold=True, color="008000")
    
    # Adjust column widths
    column_widths = [10, 15, 30, 15, 15, 15, 20]
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Save to BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Create response
    response = HttpResponse(
        output.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="petty_cash_register_{selected_date.strftime("%Y%m%d")}.xlsx"'
    
    return response


@login_required
def export_petty_cash_pdf(request):
    """Export petty cash entries to PDF with professional formatting"""
    
    if not WEASYPRINT_AVAILABLE:
        messages.error(request, 'PDF export is not available. Please install WeasyPrint.')
        return redirect('petty_cash:register')
    
    # Get selected date (default to today)
    selected_date = request.GET.get('date')
    if selected_date:
        try:
            selected_date = timezone.datetime.strptime(selected_date, '%Y-%m-%d').date()
        except:
            selected_date = timezone.now().date()
    else:
        selected_date = timezone.now().date()
    
    # Get petty cash day
    try:
        petty_cash_day = PettyCashDay.objects.get(entry_date=selected_date)
        entries = petty_cash_day.entries.all().order_by('id')
    except PettyCashDay.DoesNotExist:
        entries = []
        petty_cash_day = None
    
    # Calculate running balances
    opening_balance = petty_cash_day.opening_balance if petty_cash_day else Decimal('0.00')
    running_balance = opening_balance
    entries_with_balance = []
    
    for entry in entries:
        running_balance -= entry.amount
        entries_with_balance.append({
            'entry': entry,
            'balance': running_balance
        })
    
    # Render HTML template
    template = get_template('petty_cash/export_pdf.html')
    context = {
        'petty_cash_day': petty_cash_day,
        'entries_with_balance': entries_with_balance,
        'selected_date': selected_date,
        'opening_balance': opening_balance,
        'current_date': timezone.now().date(),
    }
    
    html_string = template.render(context)
    
    # CSS for PDF styling
    css_string = '''
    @page {
        size: A4;
        margin: 1cm;
    }
    
    body {
        font-family: Arial, sans-serif;
        font-size: 12px;
        line-height: 1.4;
        color: #333;
    }
    
    .header {
        text-align: center;
        margin-bottom: 30px;
        border-bottom: 2px solid #4472C4;
        padding-bottom: 15px;
    }
    
    .header h1 {
        margin: 0;
        color: #4472C4;
        font-size: 24px;
        font-weight: bold;
    }
    
    .header h2 {
        margin: 5px 0;
        color: #666;
        font-size: 16px;
    }
    
    .ledger-info {
        margin-bottom: 20px;
        padding: 15px;
        background-color: #f8f9fa;
        border-radius: 5px;
    }
    
    .opening-balance {
        font-size: 14px;
        font-weight: bold;
        margin-bottom: 20px;
        text-align: right;
    }
    
    table {
        width: 100%;
        border-collapse: collapse;
        margin-bottom: 20px;
    }
    
    th, td {
        border: 1px solid #ddd;
        padding: 8px;
        text-align: left;
    }
    
    th {
        background-color: #4472C4;
        color: white;
        font-weight: bold;
        text-align: center;
    }
    
    .text-center {
        text-align: center;
    }
    
    .text-right {
        text-align: right;
    }
    
    .amount {
        font-family: 'Courier New', monospace;
        font-weight: bold;
    }
    
    .positive {
        color: #28a745;
    }
    
    .negative {
        color: #dc3545;
    }
    
    .summary {
        margin-top: 30px;
        padding: 15px;
        background-color: #f8f9fa;
        border-radius: 5px;
    }
    
    .summary h3 {
        margin-top: 0;
        color: #4472C4;
    }
    
    .summary-row {
        display: flex;
        justify-content: space-between;
        margin-bottom: 10px;
        font-weight: bold;
    }
    
    .footer {
        margin-top: 40px;
        text-align: center;
        font-size: 10px;
        color: #666;
        border-top: 1px solid #ddd;
        padding-top: 10px;
    }
    '''
    
    # Generate PDF
    html = HTML(string=html_string)
    css = CSS(string=css_string)
    pdf = html.write_pdf(stylesheets=[css])
    
    # Create response
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="petty_cash_register_{selected_date.strftime("%Y%m%d")}.pdf"'
    
    return response
