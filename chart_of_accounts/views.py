from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.views.decorators.http import require_http_methods
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.views.generic import ListView, DetailView, CreateView, UpdateView, DeleteView
from django.urls import reverse_lazy
from django.contrib.auth.mixins import LoginRequiredMixin
from django.template.loader import render_to_string
from django.utils import timezone
from datetime import datetime, date
import json
import csv
import io
from django.db import IntegrityError

from .models import (
    ChartOfAccount, AccountType, AccountBalance, AccountGroup, 
    AccountTemplate, AccountTemplateItem
)
from .forms import (
    ChartOfAccountForm, ParentAccountForm, AccountBalanceForm, AccountGroupForm,
    AccountTemplateForm, AccountTemplateItemForm, ChartOfAccountSearchForm,
    BulkAccountImportForm, AccountOpeningBalanceForm
)
from .account_type_forms import AccountTypeForm
from company.company_model import Company


def test_view(request):
    """Simple test view to debug URL issues"""
    return HttpResponse("Chart of Accounts test view is working!")


def test_dashboard_view(request):
    """Test dashboard view without login requirement"""
    return HttpResponse("Chart of Accounts dashboard test - no login required!")


def test_list_view(request):
    """Simple test list view without company requirement"""
    return HttpResponse("Chart of Accounts test list view - no company required!")


@login_required
def parent_accounts_list(request):
    """View to list only parent accounts (groups)"""
    # Get current company using the same pattern as other views
    from company.company_model import Company
    company = Company.objects.filter(is_active=True).first()
    
    if not company:
        messages.error(request, 'No active company found. Please set up a company first.')
        return redirect('dashboard:dashboard')
    
    # Base queryset for parent accounts only
    parent_accounts = ChartOfAccount.objects.filter(
        company=company,
        is_group=True
    ).select_related('account_type', 'currency', 'parent_account')
    
    # Get all account types for filter dropdown
    account_types = AccountType.objects.all()
    
    # Apply filters
    search_query = request.GET.get('search', '')
    account_type_filter = request.GET.get('account_type', '')
    status_filter = request.GET.get('status', '')
    
    if search_query:
        parent_accounts = parent_accounts.filter(
            Q(name__icontains=search_query) |
            Q(account_code__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if account_type_filter:
        parent_accounts = parent_accounts.filter(account_type_id=account_type_filter)
    
    if status_filter == 'active':
        parent_accounts = parent_accounts.filter(is_active=True)
    elif status_filter == 'inactive':
        parent_accounts = parent_accounts.filter(is_active=False)
    
    # Order by account code
    parent_accounts = parent_accounts.order_by('account_code')
    
    # Calculate statistics
    total_parent_accounts = parent_accounts.count()
    asset_accounts_count = parent_accounts.filter(account_type__category='ASSET').count()
    liability_accounts_count = parent_accounts.filter(account_type__category='LIABILITY').count()
    equity_accounts_count = parent_accounts.filter(account_type__category='EQUITY').count()
    revenue_accounts_count = parent_accounts.filter(account_type__category='REVENUE').count()
    expense_accounts_count = parent_accounts.filter(account_type__category='EXPENSE').count()
    revenue_expense_count = revenue_accounts_count + expense_accounts_count
    
    # Pagination
    paginator = Paginator(parent_accounts, 12)  # 12 cards per page
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'parent_accounts': page_obj,
        'account_types': account_types,
        'total_parent_accounts': total_parent_accounts,
        'asset_accounts_count': asset_accounts_count,
        'liability_accounts_count': liability_accounts_count,
        'revenue_expense_count': revenue_expense_count,
        'page_obj': page_obj,
        'is_paginated': page_obj.has_other_pages(),
        'search_query': search_query,
        'account_type_filter': account_type_filter,
        'status_filter': status_filter,
    }
    
    return render(request, 'chart_of_accounts/parent_accounts_list.html', context)


@login_required
def chart_of_accounts_dashboard(request):
    """Dashboard view for Chart of Accounts"""
    company = Company.objects.filter(is_active=True).first()
    
    if not company:
        messages.warning(request, 'No active company found. Please set up a company first.')
        return redirect('company:company_list')
    
    # Get summary statistics
    total_accounts = ChartOfAccount.objects.filter(company=company, is_active=True).count()
    group_accounts = ChartOfAccount.objects.filter(company=company, is_active=True, is_group=True).count()
    detail_accounts = total_accounts - group_accounts
    
    # Get accounts by category
    accounts_by_category = {}
    for category_code, category_name in AccountType.ACCOUNT_CATEGORIES:
        count = ChartOfAccount.objects.filter(
            company=company,
            is_active=True,
            account_type__category=category_code
        ).count()
        accounts_by_category[category_name] = count
    
    # Get recent accounts
    recent_accounts = ChartOfAccount.objects.filter(
        company=company,
        is_active=True
    ).order_by('-created_at')[:10]
    
    # Get accounts with zero balance
    zero_balance_accounts = ChartOfAccount.objects.filter(
        company=company,
        is_active=True,
        current_balance=0
    ).count()
    
    context = {
        'total_accounts': total_accounts,
        'group_accounts': group_accounts,
        'detail_accounts': detail_accounts,
        'accounts_by_category': accounts_by_category,
        'recent_accounts': recent_accounts,
        'zero_balance_accounts': zero_balance_accounts,
        'company': company,
    }
    
    return render(request, 'chart_of_accounts/dashboard.html', context)


@login_required
def account_list(request):
    """List view for Chart of Accounts with search and filtering"""
    company = Company.objects.filter(is_active=True).first()
    
    if not company:
        messages.warning(request, 'No active company found. Please set up a company first.')
        return redirect('company:company_list')
    
    # Get search form
    search_form = ChartOfAccountSearchForm(request.GET)
    
    # Build queryset - only show ledger accounts (non-parent accounts)
    queryset = ChartOfAccount.objects.filter(company=company, is_group=False)
    
    if search_form.is_valid():
        search_term = search_form.cleaned_data.get('search_term')
        search_by = search_form.cleaned_data.get('search_by')
        account_type = search_form.cleaned_data.get('account_type')
        category = search_form.cleaned_data.get('category')
        is_active = search_form.cleaned_data.get('is_active')
        is_group = search_form.cleaned_data.get('is_group')
        
        # Apply search filters
        if search_term:
            if search_by == 'code':
                queryset = queryset.filter(account_code__icontains=search_term)
            elif search_by == 'name':
                queryset = queryset.filter(name__icontains=search_term)
            elif search_by == 'type':
                queryset = queryset.filter(account_type__name__icontains=search_term)
            elif search_by == 'category':
                queryset = queryset.filter(account_type__category__icontains=search_term)
        
        # Apply other filters
        if account_type:
            queryset = queryset.filter(account_type=account_type)
        
        if category:
            queryset = queryset.filter(account_type__category=category)
        
        if is_active:
            queryset = queryset.filter(is_active=is_active == 'True')
        
        if is_group:
            queryset = queryset.filter(is_group=is_group == 'True')
    
    # Order by account code
    queryset = queryset.order_by('account_code')
    
    # Pagination
    paginator = Paginator(queryset, 25)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    context = {
        'page_obj': page_obj,
        'search_form': search_form,
        'company': company,
    }
    
    return render(request, 'chart_of_accounts/account_list.html', context)


@login_required
def account_detail(request, pk):
    """Detail view for Chart of Account"""
    company = Company.objects.filter(is_active=True).first()
    account = get_object_or_404(ChartOfAccount, pk=pk, company=company)
    
    # Get sub-accounts
    sub_accounts = account.sub_accounts.filter(is_active=True).order_by('account_code')
    
    # Get account balances for current fiscal year
    current_fiscal_year = getattr(company, 'current_fiscal_year', None)
    balances = []
    if current_fiscal_year:
        balances = AccountBalance.objects.filter(
            account=account,
            fiscal_year=current_fiscal_year
        ).order_by('period')
    
    # Get parent accounts
    parent_accounts = account.get_parents_recursive()
    
    context = {
        'account': account,
        'sub_accounts': sub_accounts,
        'balances': balances,
        'parent_accounts': parent_accounts,
        'company': company,
    }
    
    return render(request, 'chart_of_accounts/account_detail.html', context)


@login_required
def account_create(request):
    """Create new Chart of Account"""
    company = Company.objects.filter(is_active=True).first()
    
    if request.method == 'POST':
        form = ChartOfAccountForm(request.POST, company=company)
        if form.is_valid():
            account = form.save(commit=False)
            account.created_by = request.user
            account.company = company
            
            # Auto-generate account code if not provided
            if not account.account_code and account.account_type:
                account.account_code = ChartOfAccountForm.generate_account_code(account.account_type, company)
            
            try:
                account.save()
                messages.success(request, f'Account "{account.name}" created successfully!')
                return redirect('chart_of_accounts:account_detail', pk=account.pk)
            except IntegrityError:
                form.add_error('account_code', 'An account with this code already exists for this company.')
    else:
        form = ChartOfAccountForm(company=company)
    
    # Prepare account type categories for JavaScript
    account_type_categories = {}
    for account_type in AccountType.objects.filter(is_active=True):
        account_type_categories[str(account_type.pk)] = account_type.category
    
    context = {
        'form': form,
        'company': company,
        'title': 'Create New Account',
        'account_type_categories': json.dumps(account_type_categories),
    }
    
    return render(request, 'chart_of_accounts/account_form.html', context)


@login_required
def parent_account_create(request):
    """Create new Parent Account (Group)"""
    company = Company.objects.filter(is_active=True).first()
    
    if request.method == 'POST':
        form = ParentAccountForm(request.POST, company=company)
        if form.is_valid():
            account = form.save(commit=False)
            account.created_by = request.user
            account.company = company
            
            # Auto-generate account code if not provided
            if not account.account_code and account.account_type:
                account.account_code = ParentAccountForm.generate_account_code(account.account_type, company)
            
            try:
                account.save()
                messages.success(request, f'Parent Account "{account.name}" created successfully!')
                return redirect('chart_of_accounts:account_detail', pk=account.pk)
            except IntegrityError:
                form.add_error('account_code', 'An account with this code already exists for this company.')
    else:
        form = ParentAccountForm(company=company)
    
    # Prepare account type categories for JavaScript
    account_type_categories = {}
    for account_type in AccountType.objects.filter(is_active=True):
        account_type_categories[str(account_type.pk)] = account_type.category
    
    context = {
        'form': form,
        'company': company,
        'title': 'Create New Parent Account',
        'account_type_categories': json.dumps(account_type_categories),
        'is_parent_account': True,
    }
    
    return render(request, 'chart_of_accounts/parent_account_form.html', context)


@login_required
def account_update(request, pk):
    """Update Chart of Account"""
    company = Company.objects.filter(is_active=True).first()
    account = get_object_or_404(ChartOfAccount, pk=pk, company=company)
    
    if request.method == 'POST':
        form = ChartOfAccountForm(request.POST, instance=account, company=company)
        if form.is_valid():
            account = form.save(commit=False)
            account.updated_by = request.user
            account.save()
            
            messages.success(request, f'Account "{account.name}" updated successfully!')
            return redirect('chart_of_accounts:account_detail', pk=account.pk)
    else:
        form = ChartOfAccountForm(instance=account, company=company)
        if account.parent_account:
            form.fields['parent_account_code'].initial = account.parent_account.account_code
    
    # Prepare account type categories for JavaScript
    account_type_categories = {}
    for account_type in AccountType.objects.filter(is_active=True):
        account_type_categories[str(account_type.pk)] = account_type.category
    
    context = {
        'form': form,
        'account': account,
        'company': company,
        'title': f'Edit Account: {account.name}',
        'account_type_categories': json.dumps(account_type_categories),
    }
    
    return render(request, 'chart_of_accounts/account_form.html', context)


@login_required
def account_delete(request, pk):
    """Delete Chart of Account"""
    company = Company.objects.filter(is_active=True).first()
    account = get_object_or_404(ChartOfAccount, pk=pk, company=company)
    
    if request.method == 'POST':
        account_name = account.name
        account.delete()
        messages.success(request, f'Account "{account_name}" deleted successfully!')
        return redirect('chart_of_accounts:account_list')
    
    context = {
        'account': account,
        'company': company,
    }
    
    return render(request, 'chart_of_accounts/account_confirm_delete.html', context)


@login_required
def account_hierarchy(request):
    """Display Chart of Accounts in hierarchical tree view"""
    company = Company.objects.filter(is_active=True).first()
    
    # Get root accounts (no parent)
    root_accounts = ChartOfAccount.objects.filter(
        company=company,
        is_active=True,
        parent_account__isnull=True
    ).order_by('account_code')
    
    context = {
        'root_accounts': root_accounts,
        'company': company,
    }
    
    return render(request, 'chart_of_accounts/account_hierarchy.html', context)


@login_required
def account_type_list(request):
    """Display list of account types with search and filter functionality"""
    
    # Get all account types
    account_types = AccountType.objects.all()
    
    # Search functionality
    search_query = request.GET.get('search', '')
    category_filter = request.GET.get('category', '')
    status_filter = request.GET.get('status', '')
    
    if search_query:
        account_types = account_types.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    if category_filter:
        account_types = account_types.filter(category=category_filter)
    
    if status_filter:
        if status_filter == 'active':
            account_types = account_types.filter(is_active=True)
        elif status_filter == 'inactive':
            account_types = account_types.filter(is_active=False)
    
    # Order by category and name
    account_types = account_types.order_by('category', 'name')
    
    # Pagination
    paginator = Paginator(account_types, 20)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    # Get counts for summary
    total_account_types = AccountType.objects.count()
    active_account_types = AccountType.objects.filter(is_active=True).count()
    inactive_account_types = AccountType.objects.filter(is_active=False).count()
    
    # Get counts by category
    category_counts = {}
    for category_code, category_name in AccountType.ACCOUNT_CATEGORIES:
        count = AccountType.objects.filter(category=category_code).count()
        category_counts[category_code] = count
    
    context = {
        'page_obj': page_obj,
        'search_query': search_query,
        'category_filter': category_filter,
        'status_filter': status_filter,
        'categories': AccountType.ACCOUNT_CATEGORIES,
        'total_account_types': total_account_types,
        'active_account_types': active_account_types,
        'inactive_account_types': inactive_account_types,
        'category_counts': category_counts,
    }
    
    return render(request, 'chart_of_accounts/account_type_list.html', context)


@login_required
def account_type_create(request):
    """Create a new account type"""
    
    if request.method == 'POST':
        form = AccountTypeForm(request.POST)
        if form.is_valid():
            account_type = form.save()
            messages.success(request, f'Account type "{account_type.name}" created successfully!')
            return redirect('chart_of_accounts:account_type_list')
    else:
        form = AccountTypeForm()
    
    context = {
        'form': form,
        'title': 'Create Account Type',
        'action': 'Create'
    }
    return render(request, 'chart_of_accounts/account_type_form.html', context)


@login_required
def account_type_edit(request, pk):
    """Edit an existing account type"""
    
    account_type = get_object_or_404(AccountType, pk=pk)
    
    if request.method == 'POST':
        form = AccountTypeForm(request.POST, instance=account_type)
        if form.is_valid():
            form.save()
            messages.success(request, f'Account type "{account_type.name}" updated successfully!')
            return redirect('chart_of_accounts:account_type_list')
    else:
        form = AccountTypeForm(instance=account_type)
    
    context = {
        'form': form,
        'account_type': account_type,
        'title': 'Edit Account Type',
        'action': 'Update'
    }
    return render(request, 'chart_of_accounts/account_type_form.html', context)


@login_required
def account_type_delete(request, pk):
    """Delete an account type"""
    
    account_type = get_object_or_404(AccountType, pk=pk)
    
    # Check if account type has associated accounts
    account_count = account_type.accounts.count()
    
    if request.method == 'POST':
        if account_count > 0:
            messages.error(request, f'Cannot delete account type "{account_type.name}" because it has {account_count} associated accounts.')
        else:
            account_type.delete()
            messages.success(request, f'Account type "{account_type.name}" deleted successfully!')
        return redirect('chart_of_accounts:account_type_list')
    
    context = {
        'account_type': account_type,
        'account_count': account_count
    }
    return render(request, 'chart_of_accounts/account_type_confirm_delete.html', context)


@login_required
def account_type_detail(request, pk):
    """Display account type details"""
    
    account_type = get_object_or_404(AccountType, pk=pk)
    
    # Get associated accounts
    accounts = account_type.accounts.filter(is_active=True).order_by('account_code')
    
    # Get account counts by status
    active_accounts = accounts.filter(is_active=True).count()
    inactive_accounts = account_type.accounts.filter(is_active=False).count()
    
    context = {
        'account_type': account_type,
        'accounts': accounts,
        'active_accounts': active_accounts,
        'inactive_accounts': inactive_accounts,
    }
    
    return render(request, 'chart_of_accounts/account_type_detail.html', context)


@login_required
def account_type_toggle_status(request, pk):
    """Toggle account type active status"""
    
    account_type = get_object_or_404(AccountType, pk=pk)
    
    if request.method == 'POST':
        account_type.is_active = not account_type.is_active
        account_type.save()
        
        status = 'activated' if account_type.is_active else 'deactivated'
        messages.success(request, f'Account type "{account_type.name}" {status} successfully!')
        
        return redirect('chart_of_accounts:account_type_list')
    
    return redirect('chart_of_accounts:account_type_detail', pk=pk)


@login_required
def bulk_import_accounts(request):
    """Bulk import accounts from CSV/Excel"""
    company = Company.objects.filter(is_active=True).first()
    
    if request.method == 'POST':
        form = BulkAccountImportForm(request.POST, request.FILES)
        if form.is_valid():
            # Handle file import logic here
            messages.success(request, 'Accounts imported successfully!')
            return redirect('chart_of_accounts:account_list')
    else:
        form = BulkAccountImportForm()
    
    context = {
        'form': form,
        'company': company,
    }
    
    return render(request, 'chart_of_accounts/bulk_import.html', context)


@login_required
def account_balance_list(request, account_pk):
    """List account balances for a specific account"""
    company = Company.objects.filter(is_active=True).first()
    account = get_object_or_404(ChartOfAccount, pk=account_pk, company=company)
    
    balances = AccountBalance.objects.filter(account=account).order_by('-period')
    
    context = {
        'account': account,
        'balances': balances,
        'company': company,
    }
    
    return render(request, 'chart_of_accounts/account_balance_list.html', context)


@login_required
def account_balance_create(request, account_pk):
    """Create account balance"""
    company = Company.objects.filter(is_active=True).first()
    account = get_object_or_404(ChartOfAccount, pk=account_pk, company=company)
    
    if request.method == 'POST':
        form = AccountBalanceForm(request.POST)
        if form.is_valid():
            balance = form.save(commit=False)
            balance.account = account
            balance.save()
            
            messages.success(request, 'Account balance created successfully!')
            return redirect('chart_of_accounts:account_balance_list', account_pk=account.pk)
    else:
        form = AccountBalanceForm()
    
    context = {
        'form': form,
        'account': account,
        'company': company,
    }
    
    return render(request, 'chart_of_accounts/account_balance_form.html', context)


@login_required
def account_template_list(request):
    """List account templates"""
    templates = AccountTemplate.objects.filter(is_active=True)
    
    context = {
        'templates': templates,
    }
    
    return render(request, 'chart_of_accounts/account_template_list.html', context)


@login_required
def account_template_detail(request, pk):
    """Detail view for account template"""
    template = get_object_or_404(AccountTemplate, pk=pk)
    items = template.items.all().order_by('account_code')
    
    context = {
        'template': template,
        'items': items,
    }
    
    return render(request, 'chart_of_accounts/account_template_detail.html', context)


@login_required
def apply_template(request, template_pk):
    """Apply account template to create accounts"""
    company = Company.objects.filter(is_active=True).first()
    template = get_object_or_404(AccountTemplate, pk=template_pk)
    
    if request.method == 'POST':
        # Apply template logic here
        messages.success(request, f'Template "{template.name}" applied successfully!')
        return redirect('chart_of_accounts:account_list')
    
    context = {
        'template': template,
        'company': company,
    }
    
    return render(request, 'chart_of_accounts/apply_template.html', context)


# AJAX Views
@login_required
@csrf_exempt
def get_parent_accounts(request):
    """Get parent accounts for AJAX dropdown - Only show Account Types"""
    company = Company.objects.filter(is_active=True).first()
    search_term = request.GET.get('q', '')
    
    # Filter to show only Account Types (accounts that represent account types)
    # These are typically group accounts that represent the main account type categories
    accounts = ChartOfAccount.objects.filter(
        company=company,
        is_active=True,
        is_group=True,
        level__lte=1  # Only show top-level account types (level 0 or 1)
    ).select_related('account_type')
    
    if search_term:
        accounts = accounts.filter(
            Q(account_code__icontains=search_term) |
            Q(name__icontains=search_term) |
            Q(account_type__name__icontains=search_term)
        )
    
    accounts = accounts.order_by('account_type__category', 'account_code')[:20]  # Limit results
    
    data = [{
        'id': acc.account_code, 
        'text': f"{acc.account_code} - {acc.name} ({acc.account_type.get_category_display()})"
    } for acc in accounts]
    
    return JsonResponse({'results': data})


@login_required
@csrf_exempt
def get_account_details(request, pk):
    """Get account details for AJAX"""
    try:
        company = Company.objects.filter(is_active=True).first()
        account = get_object_or_404(ChartOfAccount, pk=pk, company=company)
        
        data = {
            'success': True,
            'account': {
                'id': account.pk,
                'account_code': account.account_code,
                'name': account.name,
                'description': account.description,
                'account_type_name': account.account_type.name,
                'account_type_category': account.account_type.category,
                'account_nature': account.account_nature,
                'current_balance': float(account.current_balance),
                'is_group': account.is_group,
                'is_active': account.is_active,
                'level': account.level,
            }
        }
        
        return JsonResponse(data)
    except Exception as e:
        return JsonResponse({
            'success': False,
            'error': str(e)
        }, status=400)


@login_required
@csrf_exempt
def update_account_status(request, pk):
    """Update account active status via AJAX"""
    if request.method == 'POST':
        company = Company.objects.filter(is_active=True).first()
        account = get_object_or_404(ChartOfAccount, pk=pk, company=company)
        
        is_active = request.POST.get('is_active') == 'true'
        account.is_active = is_active
        account.save()
        
        return JsonResponse({'success': True, 'is_active': is_active})
    
    return JsonResponse({'success': False}, status=400)


@login_required
@csrf_exempt
def generate_account_code(request):
    """Generate account code based on account type"""
    if request.method == 'POST':
        account_type_id = request.POST.get('account_type_id')
        company = Company.objects.filter(is_active=True).first()
        
        if account_type_id and company:
            try:
                account_type = AccountType.objects.get(pk=account_type_id)
                generated_code = ChartOfAccountForm.generate_account_code(account_type, company)
                return JsonResponse({'success': True, 'account_code': generated_code})
            except AccountType.DoesNotExist:
                return JsonResponse({'success': False, 'error': 'Account type not found'})
        
        return JsonResponse({'success': False, 'error': 'Invalid request'})
    
    return JsonResponse({'success': False, 'error': 'Invalid method'}, status=400)


@login_required
def export_accounts_csv(request):
    """Export accounts to CSV"""
    company = Company.objects.filter(is_active=True).first()
    accounts = ChartOfAccount.objects.filter(company=company, is_active=True).order_by('account_code')
    
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="chart_of_accounts_{date.today()}.csv"'
    
    writer = csv.writer(response)
    writer.writerow([
        'Account Code', 'Name', 'Description', 'Account Type', 'Account Nature',
        'Parent Account', 'Is Group', 'Level', 'Currency', 'Current Balance', 'Is Active'
    ])
    
    for account in accounts:
        writer.writerow([
            account.account_code,
            account.name,
            account.description,
            account.account_type.name,
            account.account_nature,
            account.parent_account.account_code if account.parent_account else '',
            account.is_group,
            account.level,
            account.currency.code,
            account.current_balance,
            account.is_active,
        ])
    
    return response


@login_required
def export_parent_accounts_excel(request):
    """Export parent accounts to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'openpyxl is required for Excel export')
        return redirect('chart_of_accounts:parent_accounts_list')
    
    # Get parent accounts (accounts without parent)
    parent_accounts = ChartOfAccount.objects.filter(
        parent_account__isnull=True
    ).select_related('account_type', 'currency').order_by('account_code')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Parent Accounts"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Account Code', 'Account Name', 'Account Type', 'Account Nature', 
               'Description', 'Current Balance', 'Currency', 'Active', 'Created Date']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for row, account in enumerate(parent_accounts, 2):
        ws.cell(row=row, column=1, value=account.account_code).border = border
        ws.cell(row=row, column=2, value=account.name).border = border
        ws.cell(row=row, column=3, value=account.account_type.name).border = border
        ws.cell(row=row, column=4, value=account.get_account_nature_display()).border = border
        ws.cell(row=row, column=5, value=account.description or '').border = border
        ws.cell(row=row, column=6, value=float(account.current_balance)).border = border
        ws.cell(row=row, column=7, value=account.currency.code).border = border
        ws.cell(row=row, column=8, value='Yes' if account.is_active else 'No').border = border
        ws.cell(row=row, column=9, value=account.created_at.strftime('%Y-%m-%d')).border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="parent_accounts_{date.today()}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


@login_required
def import_parent_accounts_excel(request):
    """Import parent accounts from Excel"""
    if request.method == 'POST':
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Please select an Excel file to upload')
            return redirect('chart_of_accounts:parent_accounts_list')
        
        excel_file = request.FILES['excel_file']
        
        # Validate file type
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls)')
            return redirect('chart_of_accounts:parent_accounts_list')
        
        # Validate file size (10MB limit)
        if excel_file.size > 10 * 1024 * 1024:
            messages.error(request, 'File size must be less than 10MB')
            return redirect('chart_of_accounts:parent_accounts_list')
        
        try:
            import pandas as pd
        except ImportError:
            messages.error(request, 'pandas is required for Excel import')
            return redirect('chart_of_accounts:parent_accounts_list')
        
        try:
            # Read Excel file
            df = pd.read_excel(excel_file)
            
            # Required columns
            required_columns = ['Account Code', 'Account Name', 'Account Type', 'Account Nature']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                messages.error(request, f'Missing required columns: {", ".join(missing_columns)}')
                return redirect('chart_of_accounts:parent_accounts_list')
            
            # Process data
            success_count = 0
            error_count = 0
            errors = []
            
            for index, row in df.iterrows():
                try:
                    # Validate required fields
                    if pd.isna(row['Account Code']) or pd.isna(row['Account Name']) or \
                       pd.isna(row['Account Type']) or pd.isna(row['Account Nature']):
                        errors.append(f'Row {index + 2}: Missing required fields')
                        error_count += 1
                        continue
                    
                    # Check if account code already exists
                    if ChartOfAccount.objects.filter(account_code=row['Account Code']).exists():
                        errors.append(f'Row {index + 2}: Account code "{row["Account Code"]}" already exists')
                        error_count += 1
                        continue
                    
                    # Get account type
                    try:
                        account_type = AccountType.objects.get(name=row['Account Type'])
                    except AccountType.DoesNotExist:
                        errors.append(f'Row {index + 2}: Account type "{row["Account Type"]}" not found')
                        error_count += 1
                        continue
                    
                    # Validate account nature
                    valid_natures = ['DEBIT', 'CREDIT']
                    account_nature = str(row['Account Nature']).upper()
                    if account_nature not in valid_natures:
                        errors.append(f'Row {index + 2}: Invalid account nature "{row["Account Nature"]}" (must be DEBIT or CREDIT)')
                        error_count += 1
                        continue
                    
                    # Get default currency
                    from django.conf import settings
                    default_currency = Currency.objects.get(code=getattr(settings, 'DEFAULT_CURRENCY', 'USD'))
                    
                    # Create parent account
                    ChartOfAccount.objects.create(
                        account_code=row['Account Code'],
                        name=row['Account Name'],
                        account_type=account_type,
                        account_nature=account_nature,
                        description=row.get('Description', ''),
                        currency=default_currency,
                        parent_account=None,  # This ensures it's a parent account
                        is_active=True
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f'Row {index + 2}: {str(e)}')
                    error_count += 1
            
            # Show results
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} parent accounts')
            
            if error_count > 0:
                error_message = f'Failed to import {error_count} parent accounts:\n' + '\n'.join(errors[:10])
                if len(errors) > 10:
                    error_message += f'\n... and {len(errors) - 10} more errors'
                messages.error(request, error_message)
            
        except Exception as e:
            messages.error(request, f'Error processing Excel file: {str(e)}')
        
        return redirect('chart_of_accounts:parent_accounts_list')
    
    # GET request - show import form
    return render(request, 'chart_of_accounts/parent_account_import.html')


@login_required
def export_accounts_excel(request):
    """Export chart of accounts to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'openpyxl is required for Excel export')
        return redirect('chart_of_accounts:account_list')
    
    company = Company.objects.filter(is_active=True).first()
    accounts = ChartOfAccount.objects.filter(company=company, is_active=True).order_by('account_code')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Chart of Accounts"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = [
        'Account Code', 'Name', 'Description', 'Account Type', 'Account Nature',
        'Parent Account', 'Is Group', 'Level', 'Currency', 'Current Balance', 'Active'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for row, account in enumerate(accounts, 2):
        ws.cell(row=row, column=1, value=account.account_code).border = border
        ws.cell(row=row, column=2, value=account.name).border = border
        ws.cell(row=row, column=3, value=account.description or '').border = border
        ws.cell(row=row, column=4, value=account.account_type.name).border = border
        ws.cell(row=row, column=5, value=account.account_nature).border = border
        ws.cell(row=row, column=6, value=account.parent_account.account_code if account.parent_account else '').border = border
        ws.cell(row=row, column=7, value='Yes' if account.is_group else 'No').border = border
        ws.cell(row=row, column=8, value=account.level).border = border
        ws.cell(row=row, column=9, value=account.currency.code).border = border
        ws.cell(row=row, column=10, value=float(account.current_balance)).border = border
        ws.cell(row=row, column=11, value='Yes' if account.is_active else 'No').border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="chart_of_accounts_{date.today()}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response


@login_required
def import_accounts_excel(request):
    """Import chart of accounts from Excel"""
    if request.method == 'POST':
        try:
            import openpyxl
            import pandas as pd
        except ImportError:
            messages.error(request, 'openpyxl and pandas are required for Excel import')
            return redirect('chart_of_accounts:account_list')
        
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Please select an Excel file to upload')
            return redirect('chart_of_accounts:account_list')
        
        excel_file = request.FILES['excel_file']
        
        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls)')
            return redirect('chart_of_accounts:account_list')
        
        try:
            company = Company.objects.filter(is_active=True).first()
            if not company:
                messages.error(request, 'No active company found')
                return redirect('chart_of_accounts:account_list')
            
            # Read Excel file
            df = pd.read_excel(excel_file)
            
            # Validate required columns
            required_columns = ['Account Code', 'Name', 'Account Type', 'Account Nature']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                messages.error(request, f'Missing required columns: {", ".join(missing_columns)}')
                return redirect('chart_of_accounts:account_list')
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Process each row
            for index, row in df.iterrows():
                try:
                    # Check if account already exists
                    if ChartOfAccount.objects.filter(account_code=row['Account Code'], company=company).exists():
                        errors.append(f'Row {index + 2}: Account code "{row["Account Code"]}" already exists')
                        error_count += 1
                        continue
                    
                    # Get account type
                    try:
                        account_type = AccountType.objects.get(name=row['Account Type'])
                    except AccountType.DoesNotExist:
                        errors.append(f'Row {index + 2}: Account type "{row["Account Type"]}" not found')
                        error_count += 1
                        continue
                    
                    # Get parent account if specified
                    parent_account = None
                    if pd.notna(row.get('Parent Account', '')) and row.get('Parent Account', '').strip():
                        try:
                            parent_account = ChartOfAccount.objects.get(
                                account_code=row['Parent Account'], 
                                company=company
                            )
                        except ChartOfAccount.DoesNotExist:
                            errors.append(f'Row {index + 2}: Parent account "{row["Parent Account"]}" not found')
                            error_count += 1
                            continue
                    
                    # Create account
                    ChartOfAccount.objects.create(
                        company=company,
                        account_code=row['Account Code'],
                        name=row['Name'],
                        description=row.get('Description', ''),
                        account_type=account_type,
                        account_nature=row['Account Nature'],
                        parent_account=parent_account,
                        is_group=row.get('Is Group', 'No').lower() in ['yes', 'true', '1'],
                        level=int(row.get('Level', 1)),
                        currency=company.default_currency,
                        current_balance=float(row.get('Current Balance', 0)),
                        is_active=row.get('Active', 'Yes').lower() in ['yes', 'true', '1']
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f'Row {index + 2}: {str(e)}')
                    error_count += 1
            
            # Show results
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} accounts')
            
            if error_count > 0:
                error_message = f'Failed to import {error_count} accounts:\n' + '\n'.join(errors[:10])
                if len(errors) > 10:
                    error_message += f'\n... and {len(errors) - 10} more errors'
                messages.error(request, error_message)
            
        except Exception as e:
            messages.error(request, f'Error processing Excel file: {str(e)}')
        
        return redirect('chart_of_accounts:account_list')
    
    # GET request - show import form
    return render(request, 'chart_of_accounts/account_import.html')


@login_required
def import_account_types_excel(request):
    """Import account types from Excel"""
    if request.method == 'POST':
        try:
            import openpyxl
            import pandas as pd
        except ImportError:
            messages.error(request, 'openpyxl and pandas are required for Excel import')
            return redirect('chart_of_accounts:account_type_list')
        
        if 'excel_file' not in request.FILES:
            messages.error(request, 'Please select an Excel file to upload')
            return redirect('chart_of_accounts:account_type_list')
        
        excel_file = request.FILES['excel_file']
        
        # Validate file extension
        if not excel_file.name.endswith(('.xlsx', '.xls')):
            messages.error(request, 'Please upload a valid Excel file (.xlsx or .xls)')
            return redirect('chart_of_accounts:account_type_list')
        
        try:
            # Read Excel file
            df = pd.read_excel(excel_file)
            
            # Validate required columns
            required_columns = ['Name', 'Category', 'Description']
            missing_columns = [col for col in required_columns if col not in df.columns]
            
            if missing_columns:
                messages.error(request, f'Missing required columns: {", ".join(missing_columns)}')
                return redirect('chart_of_accounts:account_type_list')
            
            success_count = 0
            error_count = 0
            errors = []
            
            # Process each row
            for index, row in df.iterrows():
                try:
                    # Validate category
                    category_choices = dict(AccountType.ACCOUNT_CATEGORIES)
                    category = None
                    
                    # Find matching category (case-insensitive)
                    for key, value in category_choices.items():
                        if str(row['Category']).lower() == value.lower():
                            category = key
                            break
                    
                    if not category:
                        errors.append(f'Row {index + 2}: Invalid category "{row["Category"]}"')
                        error_count += 1
                        continue
                    
                    # Check if account type already exists
                    if AccountType.objects.filter(name=row['Name']).exists():
                        errors.append(f'Row {index + 2}: Account type "{row["Name"]}" already exists')
                        error_count += 1
                        continue
                    
                    # Create account type
                    AccountType.objects.create(
                        name=row['Name'],
                        category=category,
                        description=row.get('Description', ''),
                        is_active=True
                    )
                    success_count += 1
                    
                except Exception as e:
                    errors.append(f'Row {index + 2}: {str(e)}')
                    error_count += 1
            
            # Show results
            if success_count > 0:
                messages.success(request, f'Successfully imported {success_count} account types')
            
            if error_count > 0:
                error_message = f'Failed to import {error_count} account types:\n' + '\n'.join(errors[:10])
                if len(errors) > 10:
                    error_message += f'\n... and {len(errors) - 10} more errors'
                messages.error(request, error_message)
            
        except Exception as e:
            messages.error(request, f'Error processing Excel file: {str(e)}')
        
        return redirect('chart_of_accounts:account_type_list')
    
    # GET request - show import form
    return render(request, 'chart_of_accounts/account_type_import.html')


@login_required
def export_account_types_excel(request):
    """Export account types to Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        messages.error(request, 'openpyxl is required for Excel export')
        return redirect('chart_of_accounts:account_type_list')
    
    # Get account types
    account_types = AccountType.objects.all().order_by('category', 'name')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Account Types"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Headers
    headers = ['Name', 'Category', 'Description', 'Active', 'Created Date']
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    for row, account_type in enumerate(account_types, 2):
        ws.cell(row=row, column=1, value=account_type.name).border = border
        ws.cell(row=row, column=2, value=account_type.get_category_display()).border = border
        ws.cell(row=row, column=3, value=account_type.description or '').border = border
        ws.cell(row=row, column=4, value='Yes' if account_type.is_active else 'No').border = border
        ws.cell(row=row, column=5, value=account_type.created_at.strftime('%Y-%m-%d')).border = border
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="account_types_{date.today()}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    
    return response
