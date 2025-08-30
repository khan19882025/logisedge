from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.utils import timezone
from django.db.models import Q, Sum, Case, When, DecimalField
from django.contrib import messages
from decimal import Decimal
from datetime import datetime, timedelta
import json
import csv
import io
from reportlab.lib import colors
from reportlab.lib.pagesizes import letter, A4, landscape
from reportlab.lib.units import inch
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
from reportlab.lib.units import inch
from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
from django.core.mail import EmailMessage

from customer.models import Customer
from salesman.models import Salesman
from multi_currency.models import CurrencySettings
from .models import AccountsReceivableAgingReport, CustomerInvoiceAging
from .forms import AgingReportForm, AgingReportExportForm

def aging_report(request):
    """
    Main view for accounts receivable aging report
    """
    # Initialize forms
    form = AgingReportForm(request.GET if request.GET else None)
    export_form = AgingReportExportForm()
    
    # Initialize default values
    report_type = 'summary'
    
    aging_data = []
    summary_data = {
        'total_current': Decimal('0.00'),
        'total_1_30': Decimal('0.00'),
        'total_31_60': Decimal('0.00'),
        'total_61_90': Decimal('0.00'),
        'total_over_90': Decimal('0.00'),
        'grand_total': Decimal('0.00'),
        'customer_count': 0
    }
    
    # Process form if it has data and is valid, or if no data is provided (show default view)
    if not request.GET:
        # No parameters provided, show default view with sample data
        as_of_date = timezone.now().date()
        aging_data, summary_data = generate_aging_data(
            as_of_date=as_of_date,
            customer_filter=None,
            salesman_filter=None,
            customer_code_filter='',
            min_amount=Decimal('0.00'),
            aging_bucket_filter='',
            show_zero_balances=False,
            report_type='summary'
        )
    elif form.is_valid():
        # Form has data and is valid
        as_of_date = form.cleaned_data.get('as_of_date', timezone.now().date())
        customer_filter = form.cleaned_data.get('customer', None)
        salesman_filter = form.cleaned_data.get('salesman', None)
        customer_code_filter = form.cleaned_data.get('customer_code', '')
        min_amount = form.cleaned_data.get('min_amount', Decimal('0.00'))
        aging_bucket_filter = form.cleaned_data.get('aging_bucket', '')
        show_zero_balances = form.cleaned_data.get('show_zero_balances', False)
        report_type = form.cleaned_data.get('report_type', 'summary')
        
        # Generate aging data
        aging_data, summary_data = generate_aging_data(
            as_of_date=as_of_date,
            customer_filter=customer_filter,
            salesman_filter=salesman_filter,
            customer_code_filter=customer_code_filter,
            min_amount=min_amount,
            aging_bucket_filter=aging_bucket_filter,
            show_zero_balances=show_zero_balances,
            report_type=report_type
        )
    else:
        # Form has data but is not valid, show errors
        as_of_date = timezone.now().date()
        aging_data, summary_data = generate_aging_data(
            as_of_date=as_of_date,
            customer_filter=None,
            salesman_filter=None,
            customer_code_filter='',
            min_amount=Decimal('0.00'),
            aging_bucket_filter='',
            show_zero_balances=False,
            report_type='summary'
        )
    
    # Add customers and salesmen for template dropdowns
    customers = Customer.objects.filter(is_active=True).order_by('customer_name')
    salesmen = Salesman.objects.filter(status='active').order_by('first_name', 'last_name')
    
    # Get default currency
    currency_settings = CurrencySettings.objects.first()
    default_currency = currency_settings.default_currency.code if currency_settings and currency_settings.default_currency else 'USD'
    
    context = {
        'form': form,
        'export_form': export_form,
        'aging_data': aging_data,
        'summary_data': summary_data,
        'customers': customers,
        'salesmen': salesmen,
        'as_of_date': form.cleaned_data.get('as_of_date', timezone.now().date()) if form.is_valid() else timezone.now().date(),
        'default_currency': default_currency,
        'report_type': report_type,
    }
    
    return render(request, 'accounts_receivable_aging/aging_report.html', context)


def generate_aging_data(as_of_date, customer_filter=None, salesman_filter=None, customer_code_filter='', 
                       min_amount=Decimal('0.00'), aging_bucket_filter='', show_zero_balances=False, report_type='summary'):
    """
    Generate aging data based on actual invoice and payment data
    """
    from invoice.models import Invoice
    from customer_payments.models import CustomerPaymentInvoice
    
    aging_data = []
    summary_data = {
        'total_current': Decimal('0.00'),
        'total_1_30': Decimal('0.00'),
        'total_31_60': Decimal('0.00'),
        'total_61_90': Decimal('0.00'),
        'total_over_90': Decimal('0.00'),
        'grand_total': Decimal('0.00'),
        'customer_count': 0
    }
    
    # Get customers based on filters
    customers = Customer.objects.filter(is_active=True)
    
    if customer_filter:
        customers = customers.filter(id=customer_filter.id)
    
    if salesman_filter:
        customers = customers.filter(salesman=salesman_filter)
    
    if customer_code_filter:
        customers = customers.filter(customer_code__icontains=customer_code_filter)
    
    # Calculate real aging data for each customer
    for customer in customers.order_by('customer_name'):
        # Get all invoices for this customer that are not fully paid
        invoices = Invoice.objects.filter(
            customer=customer,
            status__in=['draft', 'sent', 'partial']  # Include draft, sent, and partially paid invoices
        ).exclude(status='paid')
        
        # Initialize aging buckets
        current_amount = Decimal('0.00')
        days_1_30 = Decimal('0.00')
        days_31_60 = Decimal('0.00')
        days_61_90 = Decimal('0.00')
        days_over_90 = Decimal('0.00')
        
        for invoice in invoices:
            # Calculate outstanding amount for this invoice
            invoice_total = invoice.total_sale or Decimal('0.00')
            
            # Get total payments for this invoice
            payments = CustomerPaymentInvoice.objects.filter(invoice=invoice)
            total_paid = sum(payment.amount_received for payment in payments)
            total_discount = sum(payment.discount_amount for payment in payments)
            
            outstanding_amount = invoice_total - total_paid - total_discount
            
            # Skip if no outstanding amount
            if outstanding_amount <= 0:
                continue
            
            # Calculate days outstanding from due date or invoice date
            reference_date = invoice.due_date if invoice.due_date else invoice.invoice_date
            days_outstanding = (as_of_date - reference_date).days
            
            # Categorize into aging buckets
            if days_outstanding <= 0:
                current_amount += outstanding_amount
            elif days_outstanding <= 30:
                days_1_30 += outstanding_amount
            elif days_outstanding <= 60:
                days_31_60 += outstanding_amount
            elif days_outstanding <= 90:
                days_61_90 += outstanding_amount
            else:
                days_over_90 += outstanding_amount
        
        total_outstanding = current_amount + days_1_30 + days_31_60 + days_61_90 + days_over_90
        
        # Apply minimum amount filter
        if min_amount is not None and min_amount != '':
            try:
                min_amount_decimal = Decimal(str(min_amount))
                if total_outstanding < min_amount_decimal:
                    continue
            except (ValueError, TypeError):
                pass  # Skip filter if min_amount is invalid
            
        # Apply zero balance filter
        if not show_zero_balances and total_outstanding == 0:
            continue
        
        # Only include customers with outstanding amounts or if show_zero_balances is True
        if total_outstanding > 0 or show_zero_balances:
            customer_data = {
                'customer_name': customer.customer_name,
                'customer_code': getattr(customer, 'customer_code', ''),
                'current_amount': current_amount,
                'days_1_30': days_1_30,
                'days_31_60': days_31_60,
                'days_61_90': days_61_90,
                'days_over_90': days_over_90,
                'total_outstanding': total_outstanding,
            }
            
            # Apply aging bucket filter
            if aging_bucket_filter:
                bucket_amount = customer_data.get(aging_bucket_filter.replace('-', '_'), Decimal('0.00'))
                if bucket_amount == 0:
                    continue
            
            aging_data.append(customer_data)
            
            # Update summary
            summary_data['total_current'] += current_amount
            summary_data['total_1_30'] += days_1_30
            summary_data['total_31_60'] += days_31_60
            summary_data['total_61_90'] += days_61_90
            summary_data['total_over_90'] += days_over_90
            summary_data['grand_total'] += total_outstanding
            summary_data['customer_count'] += 1
    
    # Handle different report types
    if report_type == 'details':
        # For details view, add invoice-level information to each customer
        detailed_aging_data = []
        for customer_data in aging_data:
            customer = Customer.objects.get(customer_name=customer_data['customer_name'])
            invoices = Invoice.objects.filter(
                customer=customer,
                status__in=['draft', 'sent', 'partial']
            ).exclude(status='paid')
            
            customer_data['invoices'] = []
            for invoice in invoices:
                invoice_total = invoice.total_sale or Decimal('0.00')
                payments = CustomerPaymentInvoice.objects.filter(invoice=invoice)
                total_paid = sum(payment.amount_received for payment in payments)
                total_discount = sum(payment.discount_amount for payment in payments)
                outstanding_amount = invoice_total - total_paid - total_discount
                
                if outstanding_amount > 0:
                    reference_date = invoice.due_date if invoice.due_date else invoice.invoice_date
                    days_outstanding = (as_of_date - reference_date).days
                    
                    customer_data['invoices'].append({
                        'invoice_number': invoice.invoice_number,
                        'invoice_date': invoice.invoice_date,
                        'due_date': invoice.due_date,
                        'total_amount': invoice_total,
                        'outstanding_amount': outstanding_amount,
                        'days_outstanding': days_outstanding
                    })
            
            detailed_aging_data.append(customer_data)
        aging_data = detailed_aging_data
    
    elif report_type == 'summary_with_advance_payment':
         # For summary with advance payment, add advance payment information
         from customer_payments.models import CustomerPayment
         
         total_advance_payment = Decimal('0.00')
         total_net_outstanding = Decimal('0.00')
         
         for customer_data in aging_data:
             customer = Customer.objects.get(customer_name=customer_data['customer_name'])
             
             # Get advance payments (payments without invoice allocation)
             advance_payments = CustomerPayment.objects.filter(
                 customer=customer,
                 payment_date__lte=as_of_date
             ).exclude(
                 customerpaymentinvoice__isnull=False
             )
             
             total_advance = sum(payment.amount_received for payment in advance_payments)
             customer_data['advance_payment'] = total_advance
             customer_data['net_outstanding'] = customer_data['total_outstanding'] - total_advance
             
             total_advance_payment += total_advance
             total_net_outstanding += customer_data['net_outstanding']
         
         # Add totals to summary_data
         summary_data['total_advance_payment'] = total_advance_payment
         summary_data['total_net_outstanding'] = total_net_outstanding
    
    return aging_data, summary_data


def export_aging_report(request):
    """
    Export aging report in various formats
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    # Process export request - be more lenient with form validation for export
    form = AgingReportForm(request.POST or None)
    export_form = AgingReportExportForm(request.POST)
    
    # Initialize default values
    report_type = 'summary'
    
    # For export, we'll be more lenient - only require export form to be valid
    if not export_form.is_valid():
        return JsonResponse({'error': 'Invalid export form data', 'export_form_errors': export_form.errors}, status=400)
    
    # Get form data - use cleaned_data if form is valid, otherwise use defaults
    if form.is_valid():
        as_of_date = form.cleaned_data.get('as_of_date', timezone.now().date())
        customer_filter = form.cleaned_data.get('customer', None)
        salesman_filter = form.cleaned_data.get('salesman', None)
        customer_code_filter = form.cleaned_data.get('customer_code', '')
        min_amount = form.cleaned_data.get('min_amount', Decimal('0.00'))
        aging_bucket_filter = form.cleaned_data.get('aging_bucket', '')
        show_zero_balances = form.cleaned_data.get('show_zero_balances', False)
    else:
        # Use default values if form is not valid
        as_of_date_str = request.POST.get('as_of_date', '')
        as_of_date = timezone.now().date()
        if as_of_date_str:
            try:
                as_of_date = datetime.strptime(as_of_date_str, '%Y-%m-%d').date()
            except ValueError:
                pass
        
        customer_filter = None
        salesman_filter = None
        customer_code_filter = request.POST.get('customer_code', '')
        min_amount_str = request.POST.get('min_amount', '')
        min_amount = Decimal('0.00')
        if min_amount_str:
            try:
                min_amount = Decimal(min_amount_str)
            except (ValueError, TypeError):
                pass
        aging_bucket_filter = request.POST.get('aging_bucket', '')
        show_zero_balances = request.POST.get('show_zero_balances', 'false').lower() == 'true'
        report_type = request.POST.get('report_type', 'summary')
    
    export_format = export_form.cleaned_data.get('export_format', 'pdf')
    include_details = export_form.cleaned_data.get('include_details', True)
    
    # Generate aging data
    aging_data, summary_data = generate_aging_data(
        as_of_date=as_of_date,
        customer_filter=customer_filter,
        salesman_filter=salesman_filter,
        customer_code_filter=customer_code_filter,
        min_amount=min_amount,
        aging_bucket_filter=aging_bucket_filter,
        show_zero_balances=show_zero_balances,
        report_type=report_type
    )
    
    if export_format == 'pdf':
        return export_pdf(aging_data, summary_data, as_of_date, report_type, include_details)
    elif export_format == 'csv':
        return export_csv(aging_data, summary_data, as_of_date)
    elif export_format == 'excel':
        return export_excel(aging_data, summary_data, as_of_date)
    else:
        return JsonResponse({'error': 'Invalid export format'}, status=400)


def export_pdf(aging_data, summary_data, as_of_date, report_type='summary', include_details=True):
    """
    Export aging report as PDF
    """
    response = HttpResponse(content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="accounts_receivable_aging_{as_of_date}.pdf"'
    
    # Create PDF document
    doc = SimpleDocTemplate(
        response,
        pagesize=landscape(A4),
        rightMargin=50,
        leftMargin=50,
        topMargin=20,
        bottomMargin=50
    )
    
    # Container for PDF elements
    elements = []
    styles = getSampleStyleSheet()
    
    # Custom styles
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=16,
        spaceAfter=30,
        alignment=TA_CENTER,
        textColor=colors.black
    )
    
    header_style = ParagraphStyle(
        'CustomHeader',
        parent=styles['Normal'],
        fontSize=12,
        spaceAfter=20,
        alignment=TA_CENTER
    )
    
    # Header table with company info and report details
    header_data = [
        [
            Paragraph("<b>Adirai Freight Service LLC (Br)<br/>Accounts Receivable Aging Report</b>", header_style),
            Paragraph(f"<b>As of Date:</b> {as_of_date}<br/><b>Generated:</b> {timezone.now().strftime('%Y-%m-%d %H:%M')}", header_style)
        ]
    ]
    
    header_table = Table(header_data, colWidths=[4*inch, 3*inch])
    header_table.setStyle(TableStyle([
        ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
        ('VALIGN', (0, 0), (-1, -1), 'TOP'),
        ('FONTNAME', (0, 0), (-1, -1), 'Helvetica'),
        ('FONTSIZE', (0, 0), (-1, -1), 10),
        ('TOPPADDING', (0, 0), (-1, -1), 0),
        ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
    ]))
    
    elements.append(header_table)
    elements.append(Spacer(1, 20))
    
    # Get currency symbol - use AED for dirham
    currency_symbol = 'AED '
    
    # Aging report table - adjust headers based on report type
    if report_type == 'details':
        table_data = [
            ['Invoice Number', 'Invoice Date', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days', 'Total Outstanding']
        ]
        col_widths = [1.2*inch, 1*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 0.8*inch, 1*inch]
    elif report_type == 'summary_with_advance_payment':
        table_data = [
            ['Customer', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days', 'Advance Payment', 'Net Outstanding']
        ]
        col_widths = [1.8*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.7*inch, 0.8*inch, 0.8*inch]
    else:
        table_data = [
            ['Customer', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days', 'Total']
        ]
        col_widths = [2.2*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 1*inch]
    
    for row in aging_data:
        if report_type == 'details':
            # For details report, show invoice-level information directly
            if 'invoices' in row:
                for invoice in row['invoices']:
                    # Format invoice date
                    invoice_date = invoice.get('invoice_date', '')
                    if invoice_date:
                        try:
                            from datetime import datetime
                            if isinstance(invoice_date, str):
                                invoice_date = datetime.strptime(invoice_date, '%Y-%m-%d').strftime('%d/%m/%Y')
                            else:
                                invoice_date = invoice_date.strftime('%d/%m/%Y')
                        except:
                            invoice_date = str(invoice_date)
                    
                    table_data.append([
                        invoice['invoice_number'],
                        invoice_date,
                        f"{currency_symbol}{invoice.get('current_amount', 0):,.2f}",
                        f"{currency_symbol}{invoice.get('days_1_30', 0):,.2f}",
                        f"{currency_symbol}{invoice.get('days_31_60', 0):,.2f}",
                        f"{currency_symbol}{invoice.get('days_61_90', 0):,.2f}",
                        f"{currency_symbol}{invoice.get('days_over_90', 0):,.2f}",
                        f"{currency_symbol}{invoice['outstanding_amount']:,.2f}"
                    ])
        else:
            # Summary or summary_with_advance_payment
            if report_type == 'summary_with_advance_payment':
                table_data.append([
                    row['customer_name'],
                    f"{currency_symbol}{row['current_amount']:,.2f}",
                    f"{currency_symbol}{row['days_1_30']:,.2f}",
                    f"{currency_symbol}{row['days_31_60']:,.2f}",
                    f"{currency_symbol}{row['days_61_90']:,.2f}",
                    f"{currency_symbol}{row['days_over_90']:,.2f}",
                    f"{currency_symbol}{row.get('advance_payment', 0):,.2f}",
                    f"{currency_symbol}{row.get('net_outstanding', row['total_outstanding']):,.2f}"
                ])
            else:
                table_data.append([
                    row['customer_name'],
                    f"{currency_symbol}{row['current_amount']:,.2f}",
                    f"{currency_symbol}{row['days_1_30']:,.2f}",
                    f"{currency_symbol}{row['days_31_60']:,.2f}",
                    f"{currency_symbol}{row['days_61_90']:,.2f}",
                    f"{currency_symbol}{row['days_over_90']:,.2f}",
                    f"{currency_symbol}{row['total_outstanding']:,.2f}"
                ])
    
    # Summary row
    if report_type == 'details':
        table_data.append([
            'TOTAL',
            '',  # Empty invoice date column
            f"{currency_symbol}{summary_data['total_current']:,.2f}",
            f"{currency_symbol}{summary_data['total_1_30']:,.2f}",
            f"{currency_symbol}{summary_data['total_31_60']:,.2f}",
            f"{currency_symbol}{summary_data['total_61_90']:,.2f}",
            f"{currency_symbol}{summary_data['total_over_90']:,.2f}",
            f"{currency_symbol}{summary_data['grand_total']:,.2f}"
        ])
    elif report_type == 'summary_with_advance_payment':
        table_data.append([
            'TOTAL',
            f"{currency_symbol}{summary_data['total_current']:,.2f}",
            f"{currency_symbol}{summary_data['total_1_30']:,.2f}",
            f"{currency_symbol}{summary_data['total_31_60']:,.2f}",
            f"{currency_symbol}{summary_data['total_61_90']:,.2f}",
            f"{currency_symbol}{summary_data['total_over_90']:,.2f}",
            f"{currency_symbol}{summary_data.get('total_advance_payment', 0):,.2f}",
            f"{currency_symbol}{summary_data.get('total_net_outstanding', summary_data['grand_total']):,.2f}"
        ])
    else:
        table_data.append([
            'TOTAL',
            f"{currency_symbol}{summary_data['total_current']:,.2f}",
            f"{currency_symbol}{summary_data['total_1_30']:,.2f}",
            f"{currency_symbol}{summary_data['total_31_60']:,.2f}",
            f"{currency_symbol}{summary_data['total_61_90']:,.2f}",
            f"{currency_symbol}{summary_data['total_over_90']:,.2f}",
            f"{currency_symbol}{summary_data['grand_total']:,.2f}"
        ])
    
    # Create table with dynamic column widths
    table = Table(table_data, colWidths=col_widths)
    table.setStyle(TableStyle([
        # Header row
        ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
        ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
        ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
        ('ALIGN', (0, 1), (0, -1), 'LEFT'),  # Customer names left-aligned
        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
        ('FONTSIZE', (0, 0), (-1, 0), 10),
        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
        
        # Data rows
        ('FONTNAME', (0, 1), (-1, -2), 'Helvetica'),
        ('FONTSIZE', (0, 1), (-1, -2), 9),
        ('ROWBACKGROUNDS', (0, 1), (-1, -2), [colors.beige, colors.white]),
        
        # Total row
        ('BACKGROUND', (0, -1), (-1, -1), colors.lightgrey),
        ('FONTNAME', (0, -1), (-1, -1), 'Helvetica-Bold'),
        ('FONTSIZE', (0, -1), (-1, -1), 10),
        
        # Borders
        ('GRID', (0, 0), (-1, -1), 1, colors.black),
        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.black),
        ('LINEABOVE', (0, -1), (-1, -1), 2, colors.black),
    ]))
    
    elements.append(table)
    
    # Build PDF
    doc.build(elements)
    return response


def export_csv(aging_data, summary_data, as_of_date):
    """
    Export aging report as CSV
    """
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="accounts_receivable_aging_{as_of_date}.csv"'
    
    writer = csv.writer(response)
    
    # Header
    writer.writerow(['Accounts Receivable Aging Report'])
    writer.writerow([f'As of Date: {as_of_date}'])
    writer.writerow([''])
    
    # Column headers
    writer.writerow(['Customer', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days', 'Total'])
    
    # Data rows
    for row in aging_data:
        writer.writerow([
            row['customer_name'],
            f"{row['current_amount']:.2f}",
            f"{row['days_1_30']:.2f}",
            f"{row['days_31_60']:.2f}",
            f"{row['days_61_90']:.2f}",
            f"{row['days_over_90']:.2f}",
            f"{row['total_outstanding']:.2f}"
        ])
    
    # Summary row
    writer.writerow([
        'TOTAL',
        f"{summary_data['total_current']:.2f}",
        f"{summary_data['total_1_30']:.2f}",
        f"{summary_data['total_31_60']:.2f}",
        f"{summary_data['total_61_90']:.2f}",
        f"{summary_data['total_over_90']:.2f}",
        f"{summary_data['grand_total']:.2f}"
    ])
    
    return response


def export_excel(aging_data, summary_data, as_of_date):
    """
    Export aging report as Excel
    """
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
        from openpyxl.utils import get_column_letter
    except ImportError:
        # Fallback to CSV if openpyxl is not available
        return export_csv(aging_data, summary_data, as_of_date)
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Aging Report"
    
    # Get currency symbol - use AED for dirham
    currency_symbol = 'AED '
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="D3D3D3", end_color="D3D3D3", fill_type="solid")
    
    currency_alignment = Alignment(horizontal="right")
    border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin")
    )
    
    # Report header
    ws.merge_cells('A1:G1')
    ws['A1'] = "Adirai Freight Service LLC (Br) - Accounts Receivable Aging Report"
    ws['A1'].font = Font(bold=True, size=14)
    ws['A1'].alignment = Alignment(horizontal="center")
    
    ws.merge_cells('A2:G2')
    ws['A2'] = f"As of Date: {as_of_date}"
    ws['A2'].font = Font(bold=True)
    ws['A2'].alignment = Alignment(horizontal="center")
    
    # Column headers
    headers = ['Customer', 'Current', '1-30 Days', '31-60 Days', '61-90 Days', 'Over 90 Days', 'Total']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=4, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    # Data rows
    row_num = 5
    for row_data in aging_data:
        ws.cell(row=row_num, column=1, value=row_data['customer_name']).border = border
        ws.cell(row=row_num, column=2, value=float(row_data['current_amount'])).border = border
        ws.cell(row=row_num, column=3, value=float(row_data['days_1_30'])).border = border
        ws.cell(row=row_num, column=4, value=float(row_data['days_31_60'])).border = border
        ws.cell(row=row_num, column=5, value=float(row_data['days_61_90'])).border = border
        ws.cell(row=row_num, column=6, value=float(row_data['days_over_90'])).border = border
        ws.cell(row=row_num, column=7, value=float(row_data['total_outstanding'])).border = border
        
        # Format currency columns
        for col in range(2, 8):
            cell = ws.cell(row=row_num, column=col)
            cell.number_format = f'{currency_symbol}#,##0.00'
            cell.alignment = currency_alignment
        
        row_num += 1
    
    # Total row
    ws.cell(row=row_num, column=1, value="TOTAL").font = total_font
    ws.cell(row=row_num, column=1).fill = total_fill
    ws.cell(row=row_num, column=1).border = border
    
    totals = [
        float(summary_data['total_current']),
        float(summary_data['total_1_30']),
        float(summary_data['total_31_60']),
        float(summary_data['total_61_90']),
        float(summary_data['total_over_90']),
        float(summary_data['grand_total'])
    ]
    
    for col, total in enumerate(totals, 2):
        cell = ws.cell(row=row_num, column=col, value=total)
        cell.font = total_font
        cell.fill = total_fill
        cell.border = border
        cell.number_format = f'{currency_symbol}#,##0.00'
        cell.alignment = currency_alignment
    
    # Adjust column widths
    ws.column_dimensions['A'].width = 25
    for col in range(2, 8):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="accounts_receivable_aging_{as_of_date}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    return response


def get_aging_summary(request):
    """
    AJAX endpoint to get aging summary data
    """
    if request.method != 'GET':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    form = AgingReportForm(request.GET if request.GET else None)
    
    # Handle case where no form data is provided or form is invalid
    if not request.GET or not form.is_valid():
        # Use default values when no data provided or form is invalid
        as_of_date = timezone.now().date()
        customer_filter = None
        customer_code_filter = ''
        min_amount = Decimal('0.00')
        aging_bucket_filter = ''
        show_zero_balances = False
    else:
        # Get form data when form is valid
        as_of_date = form.cleaned_data.get('as_of_date', timezone.now().date())
        customer_filter = form.cleaned_data.get('customer', None)
        customer_code_filter = form.cleaned_data.get('customer_code', '')
        min_amount = form.cleaned_data.get('min_amount', Decimal('0.00'))
        aging_bucket_filter = form.cleaned_data.get('aging_bucket', '')
        show_zero_balances = form.cleaned_data.get('show_zero_balances', False)
    
    # Generate aging data
    aging_data, summary_data = generate_aging_data(
        as_of_date=as_of_date,
        customer_filter=customer_filter,
        customer_code_filter=customer_code_filter,
        min_amount=min_amount,
        aging_bucket_filter=aging_bucket_filter,
        show_zero_balances=show_zero_balances
    )
    
    # Convert Decimal to float for JSON serialization
    for key, value in summary_data.items():
        if isinstance(value, Decimal):
            summary_data[key] = float(value)
    
    # Get default currency
    currency_settings = CurrencySettings.objects.first()
    default_currency = currency_settings.default_currency.code if currency_settings and currency_settings.default_currency else 'USD'
    
    return JsonResponse({
        'summary': summary_data,
        'record_count': len(aging_data),
        'default_currency': default_currency
    })


def send_aging_report_email(request):
    """
    Send aging report via email
    """
    if request.method != 'POST':
        return JsonResponse({'error': 'Invalid request method'}, status=405)
    
    try:
        # Get form data
        recipient_email = request.POST.get('recipient_email', '').strip()
        cc_email = request.POST.get('cc_email', '').strip()
        subject = request.POST.get('subject', 'Accounts Receivable Aging Report').strip()
        message = request.POST.get('message', '').strip()
        export_format = request.POST.get('export_format', 'pdf')
        
        if not recipient_email:
            return JsonResponse({'error': 'Recipient email is required'}, status=400)
        
        # Get aging report data using the same filters
        form = AgingReportForm(request.POST)
        if not form.is_valid():
            return JsonResponse({'error': 'Invalid form data'}, status=400)
        
        # Get form data
        as_of_date = form.cleaned_data.get('as_of_date', timezone.now().date())
        customer_filter = form.cleaned_data.get('customer', '')
        customer_code_filter = form.cleaned_data.get('customer_code', '')
        min_amount = form.cleaned_data.get('min_amount', Decimal('0.00'))
        aging_bucket_filter = form.cleaned_data.get('aging_bucket', '')
        show_zero_balances = form.cleaned_data.get('show_zero_balances', False)
        
        # Generate aging data
        aging_data, summary_data = generate_aging_data(
            as_of_date=as_of_date,
            customer_filter=customer_filter,
            customer_code_filter=customer_code_filter,
            min_amount=min_amount,
            aging_bucket_filter=aging_bucket_filter,
            show_zero_balances=show_zero_balances
        )
        
        # Generate the report file based on format
        if export_format == 'excel':
            file_response = export_excel(aging_data, summary_data, as_of_date)
            filename = f'accounts_receivable_aging_{as_of_date}.xlsx'
            content_type = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
        elif export_format == 'csv':
            file_response = export_csv(aging_data, summary_data, as_of_date)
            filename = f'accounts_receivable_aging_{as_of_date}.csv'
            content_type = 'text/csv'
        else:  # Default to PDF
            file_response = export_pdf(aging_data, summary_data, as_of_date)
            filename = f'accounts_receivable_aging_{as_of_date}.pdf'
            content_type = 'application/pdf'
        
        # Create email
        email_subject = subject or f'Accounts Receivable Aging Report - {as_of_date}'
        email_message = message or f'Please find attached the Accounts Receivable Aging Report as of {as_of_date}.'
        
        # Prepare recipient list
        recipients = [recipient_email]
        cc_list = []
        if cc_email:
            cc_list = [email.strip() for email in cc_email.split(',') if email.strip()]
        
        # Create email with attachment
        email = EmailMessage(
            subject=email_subject,
            body=email_message,
            from_email=None,  # Use default from email
            to=recipients,
            cc=cc_list,
        )
        
        # Attach the report file
        email.attach(filename, file_response.content, content_type)
        
        # Send email
        email.send()
        
        return JsonResponse({
            'success': True,
            'message': f'Aging report sent successfully to {recipient_email}'
        })
        
    except Exception as e:
        return JsonResponse({
            'error': f'Failed to send email: {str(e)}'
        }, status=500)
