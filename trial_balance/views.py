from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, Count
from django.utils import timezone
from django.views.decorators.http import require_POST, require_GET
from django.template.loader import render_to_string
from django.conf import settings
from datetime import date, datetime, timedelta
from decimal import Decimal
import json
import csv
import io

from .forms import TrialBalanceFilterForm, ExportForm
from multi_currency.models import Currency, CurrencySettings


@login_required
def trial_balance_report(request):
    """Main trial balance report view with filters and data"""
    
    # Initialize filter form
    filter_form = TrialBalanceFilterForm(request.GET or None)
    
    # Get filter parameters
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    company_id = request.GET.get('company')
    include_zero_balances = request.GET.get('include_zero_balances')
    account_type = request.GET.get('account_type')
    
    # Only generate data if form has been submitted or dates are provided
    trial_balance_data = []
    if request.GET and (from_date or to_date or request.GET.get('generate')):
        # Set default dates if not provided
        if not from_date:
            from_date = date.today().replace(day=1).isoformat()
        if not to_date:
            to_date = date.today().isoformat()
        
        # Get trial balance data
        trial_balance_data = get_trial_balance_data(
            from_date=from_date,
            to_date=to_date,
            company_id=company_id,
            include_zero_balances=include_zero_balances is not None,
            account_type=account_type
        )
    
    # Calculate totals
    total_debit = sum(entry['closing_debit'] for entry in trial_balance_data) if trial_balance_data else Decimal('0.00')
    total_credit = sum(entry['closing_credit'] for entry in trial_balance_data) if trial_balance_data else Decimal('0.00')
    difference = total_debit - total_credit
    
    # Get default currency
    default_currency = None
    try:
        currency_settings = CurrencySettings.objects.first()
        if currency_settings and currency_settings.default_currency:
            default_currency = currency_settings.default_currency
        else:
            # Fallback to AED or first active currency
            default_currency = Currency.objects.filter(code='AED').first() or Currency.objects.filter(is_active=True).first()
    except:
        # Fallback to AED or first active currency
        default_currency = Currency.objects.filter(code='AED').first() or Currency.objects.filter(is_active=True).first()
    
    # Export form
    export_form = ExportForm()
    
    context = {
        'filter_form': filter_form,
        'export_form': export_form,
        'trial_balance_data': trial_balance_data,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'difference': difference,
        'is_balanced': abs(difference) < Decimal('0.01'),
        'from_date': from_date,
        'to_date': to_date,
        'company_id': company_id,
        'include_zero_balances': include_zero_balances,
        'account_type': account_type,
        'default_currency': default_currency,
    }
    
    return render(request, 'trial_balance/trial_balance_report.html', context)


def get_trial_balance_data(from_date, to_date, company_id=None, 
                          include_zero_balances=True, account_type=None):
    """Get trial balance data based on filters"""
    
    try:
        from_date = datetime.strptime(from_date, '%Y-%m-%d').date()
        to_date = datetime.strptime(to_date, '%Y-%m-%d').date()
    except:
        from_date = date.today().replace(day=1)
        to_date = date.today()
    
    # Get accounts
    try:
        from chart_of_accounts.models import ChartOfAccount
        accounts = ChartOfAccount.objects.filter(is_active=True)
        
        if company_id:
            accounts = accounts.filter(company_id=company_id)
        
        if account_type:
            accounts = accounts.filter(account_type_id=account_type)
        
        accounts = accounts.order_by('account_code')
        
    except Exception as e:
        return []
    
    trial_balance_data = []
    
    for account in accounts:
        # Get ledger entries for the period
        try:
            from ledger.models import Ledger
            
            # Period entries
            period_entries = Ledger.objects.filter(
                account=account,
                entry_date__gte=from_date,
                entry_date__lte=to_date
            )
            
            period_debit = period_entries.filter(entry_type='DR').aggregate(
                total=Sum('amount'))['total'] or Decimal('0.00')
            period_credit = period_entries.filter(entry_type='CR').aggregate(
                total=Sum('amount'))['total'] or Decimal('0.00')
            
            # Opening balance (entries before from_date)
            opening_entries = Ledger.objects.filter(
                account=account,
                entry_date__lt=from_date
            )
            
            opening_debit = opening_entries.filter(entry_type='DR').aggregate(
                total=Sum('amount'))['total'] or Decimal('0.00')
            opening_credit = opening_entries.filter(entry_type='CR').aggregate(
                total=Sum('amount'))['total'] or Decimal('0.00')
            
            # Calculate closing balances
            closing_debit = opening_debit + period_debit
            closing_credit = opening_credit + period_credit
            
            # Calculate running balance
            if closing_debit > closing_credit:
                running_balance = closing_debit - closing_credit
            else:
                running_balance = -(closing_credit - closing_debit)
            
            # Skip accounts with zero or insignificant balances if not included
            if not include_zero_balances:
                # Check if all values are effectively zero (handling floating point precision)
                tolerance = 0.01  # Consider values less than 0.01 as zero
                if (
                    abs(running_balance) < tolerance and 
                    opening_debit < tolerance and 
                    opening_credit < tolerance and 
                    period_debit < tolerance and 
                    period_credit < tolerance and 
                    closing_debit < tolerance and 
                    closing_credit < tolerance
                ):
                    continue
            
            trial_balance_data.append({
                'account_code': account.account_code,
                'account_name': account.name,
                'account_type': account.account_type.name if account.account_type else '',
                'opening_debit': opening_debit,
                'opening_credit': opening_credit,
                'period_debit': period_debit,
                'period_credit': period_credit,
                'closing_debit': closing_debit,
                'closing_credit': closing_credit,
                'running_balance': running_balance,
            })
            
        except Exception as e:
            # If ledger model doesn't exist, use account balances
            trial_balance_data.append({
                'account_code': account.account_code,
                'account_name': account.name,
                'account_type': account.account_type.name if account.account_type else '',
                'opening_debit': Decimal('0.00'),
                'opening_credit': Decimal('0.00'),
                'period_debit': Decimal('0.00'),
                'period_credit': Decimal('0.00'),
                'closing_debit': Decimal('0.00'),
                'closing_credit': Decimal('0.00'),
                'running_balance': Decimal('0.00'),
            })
    
    return trial_balance_data


@login_required
@require_POST
def export_trial_balance(request):
    """Export trial balance data to various formats"""
    
    # Get export parameters
    export_format = request.POST.get('format', 'csv')
    include_headers = request.POST.get('include_headers', 'on') == 'on'
    include_totals = request.POST.get('include_totals', 'on') == 'on'
    include_running_balance = request.POST.get('include_running_balance', 'on') == 'on'
    
    # Get filter parameters
    from_date = request.POST.get('from_date')
    to_date = request.POST.get('to_date')
    company_id = request.POST.get('company')
    include_zero_balances = request.POST.get('include_zero_balances', 'on') == 'on'
    account_type = request.POST.get('account_type')
    
    # Get trial balance data
    trial_balance_data = get_trial_balance_data(
        from_date=from_date,
        to_date=to_date,
        company_id=company_id,
        include_zero_balances=include_zero_balances,
        account_type=account_type
    )
    
    # Calculate totals
    total_debit = sum(entry['closing_debit'] for entry in trial_balance_data)
    total_credit = sum(entry['closing_credit'] for entry in trial_balance_data)
    difference = total_debit - total_credit
    
    # Get default currency
    try:
        currency_settings = CurrencySettings.objects.first()
        default_currency = currency_settings.default_currency if currency_settings else None
        if not default_currency:
            default_currency = Currency.objects.filter(code='AED').first()
        if not default_currency:
            default_currency = Currency.objects.filter(is_active=True).first()
    except:
        default_currency = None
    
    # Generate filename
    filename = f"trial_balance_{from_date}_to_{to_date}"
    
    if export_format == 'csv':
        response = export_to_csv(
            trial_balance_data, total_debit, total_credit, difference,
            include_headers, include_totals, include_running_balance,
            from_date, to_date, default_currency
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}.csv"'
        return response
    
    elif export_format == 'excel':
        try:
            response = export_to_excel(
                trial_balance_data, total_debit, total_credit, difference,
                include_headers, include_totals, include_running_balance,
                from_date, to_date, default_currency
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.xlsx"'
            return response
        except ImportError as e:
            messages.error(request, str(e))
            return redirect('trial_balance:trial_balance_report')
    
    elif export_format == 'pdf':
        try:
            response = export_to_pdf(
                trial_balance_data, total_debit, total_credit, difference,
                include_headers, include_totals, include_running_balance,
                from_date, to_date, default_currency
            )
            response['Content-Disposition'] = f'attachment; filename="{filename}.pdf"'
            return response
        except ImportError as e:
            messages.error(request, str(e))
            return redirect('trial_balance:trial_balance_report')
    
    else:
        messages.error(request, 'Invalid export format')
        return redirect('trial_balance:trial_balance_report')


def export_to_csv(trial_balance_data, total_debit, total_credit, difference,
                  include_headers, include_totals, include_running_balance,
                  from_date, to_date, default_currency=None):
    """Export trial balance data to CSV format"""
    
    response = HttpResponse(content_type='text/csv')
    writer = csv.writer(response)
    
    if include_headers:
        headers = ['Account Code', 'Account Name', 'Account Type', 'Opening Debit', 'Opening Credit']
        if include_running_balance:
            headers.extend(['Period Debit', 'Period Credit', 'Closing Debit', 'Closing Credit', 'Running Balance'])
        else:
            headers.extend(['Period Debit', 'Period Credit', 'Closing Debit', 'Closing Credit'])
        writer.writerow(headers)
    
    for entry in trial_balance_data:
        row = [
            entry['account_code'],
            entry['account_name'],
            entry['account_type'],
            entry['opening_debit'],
            entry['opening_credit'],
            entry['period_debit'],
            entry['period_credit'],
            entry['closing_debit'],
            entry['closing_credit'],
        ]
        if include_running_balance:
            row.append(entry['running_balance'])
        writer.writerow(row)
    
    if include_totals:
        currency_symbol = default_currency.symbol if default_currency else ''
        formatted_total_debit = f"{currency_symbol} {total_debit}" if currency_symbol else str(total_debit)
        formatted_total_credit = f"{currency_symbol} {total_credit}" if currency_symbol else str(total_credit)
        formatted_difference = f"{currency_symbol} {difference}" if currency_symbol else str(difference)
        
        writer.writerow([])
        writer.writerow(['TOTALS', '', '', formatted_total_debit, formatted_total_credit, '', '', formatted_total_debit, formatted_total_credit])
        if include_running_balance:
            writer.writerow(['', '', '', '', '', '', '', '', '', formatted_difference])
        else:
            writer.writerow(['', '', '', '', '', '', '', '', formatted_difference])
    
    return response


def export_to_excel(trial_balance_data, total_debit, total_credit, difference,
                    include_headers, include_totals, include_running_balance,
                    from_date, to_date, default_currency=None):
    """Export trial balance data to Excel format"""
    
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill
    except ImportError:
        raise ImportError('openpyxl is required for Excel export')
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Trial Balance"
    
    # Styles
    header_font = Font(bold=True)
    header_fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
    total_font = Font(bold=True)
    total_fill = PatternFill(start_color="E6E6E6", end_color="E6E6E6", fill_type="solid")
    
    current_row = 1
    
    # Add title
    ws.merge_cells(f'A{current_row}:I{current_row}')
    ws[f'A{current_row}'] = f"Trial Balance Report - {from_date} to {to_date}"
    ws[f'A{current_row}'].font = Font(bold=True, size=14)
    ws[f'A{current_row}'].alignment = Alignment(horizontal='center')
    current_row += 2
    
    # Add headers
    if include_headers:
        headers = ['Account Code', 'Account Name', 'Account Type', 'Opening Debit', 'Opening Credit']
        if include_running_balance:
            headers.extend(['Period Debit', 'Period Credit', 'Closing Debit', 'Closing Credit', 'Running Balance'])
        else:
            headers.extend(['Period Debit', 'Period Credit', 'Closing Debit', 'Closing Credit'])
        
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=current_row, column=col, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
    
    # Add data
    for entry in trial_balance_data:
        row_data = [
            entry['account_code'],
            entry['account_name'],
            entry['account_type'],
            entry['opening_debit'],
            entry['opening_credit'],
            entry['period_debit'],
            entry['period_credit'],
            entry['closing_debit'],
            entry['closing_credit'],
        ]
        if include_running_balance:
            row_data.append(entry['running_balance'])
        
        for col, value in enumerate(row_data, 1):
            ws.cell(row=current_row, column=col, value=value)
        
        current_row += 1
    
    # Add totals
    if include_totals:
        current_row += 1
        total_row = current_row
        
        # Format totals with currency symbol
        currency_symbol = default_currency.symbol if default_currency else ''
        formatted_total_debit = f"{currency_symbol} {total_debit}" if currency_symbol else str(total_debit)
        formatted_total_credit = f"{currency_symbol} {total_credit}" if currency_symbol else str(total_credit)
        formatted_difference = f"{currency_symbol} {difference}" if currency_symbol else str(difference)
        
        # Total row
        ws.cell(row=total_row, column=1, value="TOTALS").font = total_font
        ws.cell(row=total_row, column=4, value=formatted_total_debit).font = total_font
        ws.cell(row=total_row, column=5, value=formatted_total_credit).font = total_font
        ws.cell(row=total_row, column=8, value=formatted_total_debit).font = total_font
        ws.cell(row=total_row, column=9, value=formatted_total_credit).font = total_font
        
        if include_running_balance:
            ws.cell(row=total_row, column=10, value=formatted_difference).font = total_font
        
        # Apply fill to total row
        for col in range(1, len(headers) + 1):
            ws.cell(row=total_row, column=col).fill = total_fill
    
    # Auto-adjust column widths
    from openpyxl.cell.cell import MergedCell
    for column in ws.columns:
        max_length = 0
        column_letter = None
        
        # Find the first non-merged cell to get column letter
        for cell in column:
            if not isinstance(cell, MergedCell):
                column_letter = cell.column_letter
                break
        
        if column_letter is None:
            continue
            
        for cell in column:
            try:
                if not isinstance(cell, MergedCell) and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    wb.save(response)
    
    return response


def export_to_pdf(trial_balance_data, total_debit, total_credit, difference,
                  include_headers, include_totals, include_running_balance,
                  from_date, to_date, default_currency=None):
    """Export trial balance data to PDF format"""
    
    try:
        from reportlab.lib.pagesizes import letter, A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch
        from reportlab.lib import colors
    except ImportError:
        raise ImportError('reportlab is required for PDF export')
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    
    # Create PDF document with landscape orientation and smaller margins
    doc = SimpleDocTemplate(
        response, 
        pagesize=landscape(A4),
        leftMargin=0.5*inch,
        rightMargin=0.5*inch,
        topMargin=0.5*inch,
        bottomMargin=0.5*inch
    )
    elements = []
    
    # Styles
    styles = getSampleStyleSheet()
    title_style = ParagraphStyle(
        'CustomTitle',
        parent=styles['Heading1'],
        fontSize=12,
        spaceAfter=15,
        alignment=1  # Center alignment
    )
    
    # Add title
    title = Paragraph(f"Trial Balance Report<br/>Period: {from_date} to {to_date}", title_style)
    elements.append(title)
    elements.append(Spacer(1, 10))
    
    # Prepare table data
    table_data = []
    
    # Add headers
    if include_headers:
        headers = ['Account Code', 'Account Name', 'Account Type', 'Opening Debit', 'Opening Credit']
        if include_running_balance:
            headers.extend(['Period Debit', 'Period Credit', 'Closing Debit', 'Closing Credit', 'Running Balance'])
        else:
            headers.extend(['Period Debit', 'Period Credit', 'Closing Debit', 'Closing Credit'])
        table_data.append(headers)
    
    # Add data rows
    for entry in trial_balance_data:
        row = [
            entry['account_code'],
            entry['account_name'],
            entry['account_type'],
            f"{entry['opening_debit']:.2f}",
            f"{entry['opening_credit']:.2f}",
            f"{entry['period_debit']:.2f}",
            f"{entry['period_credit']:.2f}",
            f"{entry['closing_debit']:.2f}",
            f"{entry['closing_credit']:.2f}",
        ]
        if include_running_balance:
            row.append(f"{entry['running_balance']:.2f}")
        table_data.append(row)
    
    # Add totals
    if include_totals:
        currency_symbol = default_currency.symbol if default_currency else ''
        formatted_total_debit = f"{currency_symbol} {total_debit:.2f}" if currency_symbol else f"{total_debit:.2f}"
        formatted_total_credit = f"{currency_symbol} {total_credit:.2f}" if currency_symbol else f"{total_credit:.2f}"
        formatted_difference = f"{currency_symbol} {difference:.2f}" if currency_symbol else f"{difference:.2f}"
        
        table_data.append([])  # Empty row
        total_row = ["TOTALS", "", "", formatted_total_debit, formatted_total_credit, "", "", formatted_total_debit, formatted_total_credit]
        if include_running_balance:
            total_row.append(formatted_difference)
        table_data.append(total_row)
    
    # Create table with optimized column widths
    if table_data:
        # Define column widths based on content type
        if include_running_balance:
            col_widths = [0.8*inch, 2.2*inch, 1.0*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch, 0.9*inch]
        else:
            col_widths = [0.9*inch, 2.5*inch, 1.1*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.0*inch, 1.0*inch]
        
        table = Table(table_data, repeatRows=1, colWidths=col_widths)
        
        # Table style with reduced font sizes and optimized spacing
        style = TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 8),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 6),
            ('TOPPADDING', (0, 0), (-1, 0), 6),
            
            # Data rows styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 7),
            ('BOTTOMPADDING', (0, 1), (-1, -1), 3),
            ('TOPPADDING', (0, 1), (-1, -1), 3),
            
            # Alignment
            ('ALIGN', (0, 0), (2, -1), 'LEFT'),  # Left align text columns
            ('ALIGN', (3, 0), (-1, -1), 'RIGHT'),  # Right align number columns
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Totals row styling
            ('BACKGROUND', (0, -2), (-1, -1), colors.lightgrey),
            ('FONTNAME', (0, -2), (-1, -1), 'Helvetica-Bold'),
            ('FONTSIZE', (0, -2), (-1, -1), 8),
            
            # Grid and borders
            ('GRID', (0, 0), (-1, -1), 0.5, colors.black),
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.black),
            ('LINEABOVE', (0, -2), (-1, -2), 1, colors.black),
        ])
        
        table.setStyle(style)
        elements.append(table)
    
    # Build PDF
    doc.build(elements)
    
    return response