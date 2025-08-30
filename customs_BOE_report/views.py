from django.shortcuts import render
from django.http import HttpResponse, JsonResponse
from django.db.models import Q, Sum
from django.utils import timezone
from django.contrib import messages
from datetime import datetime, date
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from weasyprint import HTML, CSS
from django.template.loader import render_to_string
from .models import BOETransaction
from job.models import Job, JobContainer
from delivery_order.models import DeliveryOrder, DeliveryOrderItem
from documentation.models import Documentation, DocumentationCargo
from crossstuffing.models import CrossStuffing, CrossStuffingCargo
import io

def get_additional_goods_out_for_boe(boe_number):
    """
    Calculate actual goods out from documentation and cross stuffing modules
    for BOE number 30200521211225 - only real transactions, no demo data
    """
    target_boe = "30200521211225"
    if boe_number != target_boe:
        return {'qty_out': 0, 'wt_out': 0, 'value_out': 0}
    
    total_qty_out = 0
    total_wt_out = 0
    total_value_out = 0
    
    # Get actual data from Documentation module - only processed/completed records
    documentation_records = Documentation.objects.filter(
        boe=target_boe
    ).exclude(status__in=['draft', 'cancelled', 'demo'])
    
    for doc in documentation_records:
        for cargo_item in doc.cargo_items.all():
            # Only count items that have been actually processed (not demo/placeholder)
            if cargo_item.quantity and cargo_item.quantity > 0:
                total_qty_out += float(cargo_item.quantity)
                total_wt_out += float(cargo_item.gross_weight or cargo_item.net_weight or 0)
                total_value_out += float(cargo_item.amount or 0)
    
    # Get actual data from Cross Stuffing module - only completed cross stuffing
    crossstuffing_records = CrossStuffing.objects.filter(
        boe=target_boe
    ).exclude(status__in=['draft', 'cancelled', 'demo'])
    
    for cs in crossstuffing_records:
        for cargo_item in cs.cargo_items.all():
            # Only count items that have been actually cross stuffed
            if cargo_item.quantity and cargo_item.quantity > 0:
                total_qty_out += float(cargo_item.quantity)
                total_wt_out += float(cargo_item.gross_weight or cargo_item.net_weight or 0)
                total_value_out += float(cargo_item.amount or 0)
    
    return {
        'qty_out': total_qty_out,
        'wt_out': total_wt_out,
        'value_out': total_value_out
    }

def customs_boe_report(request):
    """Main view for Customs BOE Stock report"""
    filter_type = request.GET.get('filter', 'all')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    declaration_no = request.GET.get('declaration_no', '').strip()
    hs_code = request.GET.get('hs_code', '').strip()
    particulars = request.GET.get('particulars', '').strip()
    cog = request.GET.get('cog', '').strip()
    
    # Get data from Job containers and related GRN items
    containers = JobContainer.objects.filter(
        Q(ed_number__isnull=False) | Q(m1_number__isnull=False)
    ).exclude(
        ed_number='', m1_number=''
    ).select_related('job')
    
    # Set default dates if not provided
    if not from_date:
        # Get the earliest declaration date
        earliest_container = containers.order_by('job__created_at').first()
        if earliest_container:
            from_date = earliest_container.job.created_at.date().strftime('%Y-%m-%d')
        else:
            from_date = date.today().strftime('%Y-%m-%d')
    
    if not to_date:
        # Always set to current date
        to_date = date.today().strftime('%Y-%m-%d')
    
    # Apply date filtering if provided
    if from_date:
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            containers = containers.filter(job__created_at__date__gte=from_date_obj)
        except ValueError:
            messages.error(request, 'Invalid from date format. Please use YYYY-MM-DD.')
    
    if to_date:
        try:
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
            containers = containers.filter(job__created_at__date__lte=to_date_obj)
        except ValueError:
            messages.error(request, 'Invalid to date format. Please use YYYY-MM-DD.')
    
    # Build transaction data from containers and their cargo details
    transactions = []
    processed_boes = set()  # Track processed BOEs to avoid duplication
    
    for container in containers:
        # Skip BOE 30200521211225 as it's handled separately in the additional section
        target_boe = "30200521211225"
        if container.ed_number == target_boe:
            continue
            
        # Get cargo items for this container's job
        cargo_items = container.job.cargo_items.all()
        
        if cargo_items.exists():
            # Create transaction for each cargo item
            for cargo in cargo_items:
                # Calculate outbound quantities from delivery orders
                qty_out = 0
                wt_out = 0
                value_out = 0
                
                # Find delivery orders for this container/BOE
                delivery_orders = DeliveryOrder.objects.filter(
                    Q(boe=container.ed_number) | Q(container=container.container_number)
                ).exclude(status='cancelled')
                
                for do in delivery_orders:
                    # Get delivery order items that match this cargo item
                    do_items = do.items.filter(item=cargo.item) if cargo.item else []
                    for do_item in do_items:
                        qty_out += float(do_item.shipped_qty or 0)
                        # Calculate weight proportionally based on shipped quantity
                        if cargo.quantity and cargo.quantity > 0:
                            weight_ratio = float(do_item.shipped_qty or 0) / float(cargo.quantity)
                            wt_out += float(cargo.gross_weight or cargo.net_weight or 0) * weight_ratio
                            value_out += float(cargo.amount or 0) * weight_ratio
                
                transaction_data = {
                    'declaration_no': container.ed_number or '',
                    'bill_no': container.m1_number or '',
                    'date': container.job.created_at.date(),
                    'hs_code': cargo.hs_code or '',
                    'particulars': cargo.item.item_name if cargo.item else (cargo.item_code or f'Container {container.container_number or "N/A"}'),
                    'cog': cargo.coo or '',
                    'pkg_type': cargo.unit or '',
                    'qty_in': float(cargo.quantity or 0),
                    'wt_in': float(cargo.gross_weight or cargo.net_weight or 0),
                    'value_in': float(cargo.amount or 0),
                    'qty_out': qty_out,
                    'wt_out': wt_out,
                    'value_out': value_out,
                    'duty': 0,
                    'total_dues': 0,
                    'job_code': container.job.job_code,
                    'container_number': container.container_number or '',
                }
                transactions.append(transaction_data)
        else:
            # Create placeholder transaction if no cargo items
            # Calculate outbound quantities from delivery orders for this container
            qty_out = 0
            wt_out = 0
            value_out = 0
            
            # Find delivery orders for this container/BOE
            delivery_orders = DeliveryOrder.objects.filter(
                Q(boe=container.ed_number) | Q(container=container.container_number)
            ).exclude(status='cancelled')
            
            for do in delivery_orders:
                # Sum all shipped quantities for this delivery order
                for do_item in do.items.all():
                    qty_out += float(do_item.shipped_qty or 0)
                    # Use item weights if available
                    if do_item.item:
                        wt_out += float(do_item.item.gross_weight or do_item.item.net_weight or 0) * float(do_item.shipped_qty or 0)
                        if do_item.unit_price:
                            value_out += float(do_item.unit_price) * float(do_item.shipped_qty or 0)
            

            
            # Get cargo items from the job to populate HS CODE, PARTICULARS, and COG
            cargo_items = container.job.cargo_items.all()
            hs_codes = []
            particulars = []
            cogs = []
            pkg_types = []
            
            for cargo in cargo_items:
                if cargo.hs_code:
                    hs_codes.append(cargo.hs_code)
                if cargo.item and cargo.item.item_name:
                    particulars.append(cargo.item.item_name)
                elif cargo.item_code:
                    particulars.append(cargo.item_code)
                if cargo.coo:
                    cogs.append(cargo.coo)
                if cargo.unit:
                    pkg_types.append(cargo.unit)
            
            # Join multiple values with commas or use container info as fallback
            hs_code_display = ', '.join(hs_codes) if hs_codes else ''
            particulars_display = ', '.join(particulars) if particulars else f'Container {container.container_number or "N/A"} - {container.job.job_code}'
            cog_display = ', '.join(cogs) if cogs else ''
            pkg_type_display = ', '.join(pkg_types) if pkg_types else container.container_size or ''
            
            transaction_data = {
                'declaration_no': container.ed_number or '',
                'bill_no': container.m1_number or '',
                'date': container.job.created_at.date(),
                'hs_code': hs_code_display,
                'particulars': particulars_display,
                'cog': cog_display,
                'pkg_type': pkg_type_display,
                'qty_in': 1,
                'wt_in': 0,
                'value_in': 0,
                'qty_out': qty_out,
                'wt_out': wt_out,
                'value_out': value_out,
                'duty': 0,
                'total_dues': 0,
                'job_code': container.job.job_code,
                'container_number': container.container_number or '',
            }
            transactions.append(transaction_data)
    
    # Add additional transactions from documentation and cross stuffing modules
    # Get all documentation records with BOE numbers
    documentation_records = Documentation.objects.filter(
        boe__isnull=False
    ).exclude(boe='')
        
    for doc in documentation_records:
        # Process each unique BOE only once to avoid duplication
        if doc.boe not in processed_boes:
            processed_boes.add(doc.boe)
            
            for cargo_item in doc.cargo_items.all():
                if cargo_item.quantity and cargo_item.quantity > 0:
                    transaction_data = {
                        'declaration_no': doc.boe,
                        'bill_no': doc.document_no or '',
                        'date': doc.created_at.date() if hasattr(doc, 'created_at') else timezone.now().date(),
                        'hs_code': cargo_item.hs_code or '',
                        'particulars': cargo_item.item_name or '',
                        'cog': cargo_item.coo or '',
                        'pkg_type': cargo_item.unit or '',
                        'qty_in': 0,  # Documentation is outbound only
                        'wt_in': 0,
                        'value_in': 0,
                        'qty_out': float(cargo_item.quantity),
                        'wt_out': float(cargo_item.gross_weight or cargo_item.net_weight or 0),
                        'value_out': float(cargo_item.amount or 0),
                        'duty': 0,
                        'total_dues': 0,
                        'job_code': doc.document_no or '',
                        'container_number': '',
                    }
                    transactions.append(transaction_data)
        
    # Get all cross stuffing records with BOE numbers
    crossstuffing_records = CrossStuffing.objects.filter(
        boe__isnull=False
    ).exclude(boe='').exclude(status__in=['draft', 'cancelled', 'demo'])
    
    for cs in crossstuffing_records:
        # Process each unique BOE only once to avoid duplication
        if cs.boe not in processed_boes:
            processed_boes.add(cs.boe)
            
            for cargo_item in cs.cargo_items.all():
                if cargo_item.quantity and cargo_item.quantity > 0:
                    transaction_data = {
                        'declaration_no': cs.boe,
                        'bill_no': cs.cs_number or '',
                        'date': cs.created_at.date() if hasattr(cs, 'created_at') else timezone.now().date(),
                        'hs_code': cargo_item.hs_code or '',
                        'particulars': cargo_item.item_name or '',
                        'cog': cargo_item.coo or '',
                        'pkg_type': cargo_item.unit or '',
                        'qty_in': 0,  # Cross stuffing is outbound only
                        'wt_in': 0,
                        'value_in': 0,
                        'qty_out': float(cargo_item.quantity),
                        'wt_out': float(cargo_item.gross_weight or cargo_item.net_weight or 0),
                        'value_out': float(cargo_item.amount or 0),
                        'duty': 0,
                        'total_dues': 0,
                        'job_code': cs.cs_number or '',
                        'container_number': '',
                    }
                    transactions.append(transaction_data)
    
    # Apply text-based filters
    if declaration_no:
        transactions = [t for t in transactions if declaration_no.lower() in str(t['declaration_no']).lower()]
    
    if hs_code:
        transactions = [t for t in transactions if hs_code.lower() in str(t['hs_code']).lower()]
    
    if particulars:
        transactions = [t for t in transactions if particulars.lower() in str(t['particulars']).lower()]
    
    if cog:
        transactions = [t for t in transactions if cog.lower() in str(t['cog']).lower()]
    
    # Apply type-based filters
    if filter_type == 'in_boe':
        transactions = [t for t in transactions if t['qty_in'] > 0]
    elif filter_type == 'out_boe':
        transactions = [t for t in transactions if t['qty_out'] > 0]
    elif filter_type == 'balance_boe':
        transactions = [t for t in transactions if t['qty_in'] > 0 or t['qty_out'] > 0]
    
    # Sort transactions by date with current date always last
    current_date = timezone.now().date()
    
    def date_sort_key(transaction):
        trans_date = transaction['date']
        # If transaction date is current date, assign a high sort value to put it last
        if trans_date == current_date:
            return (1, trans_date)  # Current date transactions go last
        else:
            return (0, trans_date)  # Other dates sorted normally
    
    transactions.sort(key=date_sort_key)
    
    # Calculate dynamic date range for display
    if from_date and to_date:
        date_range = f"{from_date} to {to_date}"
    elif from_date:
        date_range = f"From {from_date}"
    elif to_date:
        date_range = f"Up to {to_date}"
    elif transactions:
        # Calculate from actual transaction dates
        transaction_dates = [t['date'] for t in transactions]
        min_date = min(transaction_dates).strftime('%d/%m/%Y')
        max_date = max(transaction_dates).strftime('%d/%m/%Y')
        date_range = f"{min_date} to {max_date}"
    else:
        date_range = "No data available"
    
    context = {
        'transactions': transactions,
        'filter_type': filter_type,
        'total_transactions': len(transactions),
        'from_date': from_date,
        'to_date': to_date,
        'declaration_no': declaration_no,
        'hs_code': hs_code,
        'particulars': particulars,
        'cog': cog,
        'date_range': date_range,
        'today': timezone.now().date(),
    }
    
    return render(request, 'customs_BOE_report/report.html', context)

def export_to_excel(request):
    """Export BOE transactions to Excel with specific formatting"""
    filter_type = request.GET.get('filter', 'all')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    declaration_no = request.GET.get('declaration_no', '').strip()
    hs_code = request.GET.get('hs_code', '').strip()
    particulars = request.GET.get('particulars', '').strip()
    cog = request.GET.get('cog', '').strip()
    
    # Get data from Job containers and related GRN items (same logic as main view)
    containers = JobContainer.objects.filter(
        Q(ed_number__isnull=False) | Q(m1_number__isnull=False)
    ).exclude(
        ed_number='', m1_number=''
    ).select_related('job')
    
    # Set default dates if not provided
    if not from_date:
        # Get the earliest declaration date
        earliest_container = containers.order_by('job__created_at').first()
        if earliest_container:
            from_date = earliest_container.job.created_at.date().strftime('%Y-%m-%d')
        else:
            from_date = date.today().strftime('%Y-%m-%d')
    
    if not to_date:
        # Always set to current date
        to_date = date.today().strftime('%Y-%m-%d')
    
    # Apply date filtering if provided
    if from_date:
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            containers = containers.filter(job__created_at__date__gte=from_date_obj)
        except ValueError:
            pass  # Skip invalid dates in export
    
    if to_date:
        try:
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
            containers = containers.filter(job__created_at__date__lte=to_date_obj)
        except ValueError:
            pass  # Skip invalid dates in export
    
    # Build transaction data from containers and their cargo details
    transactions = []
    for container in containers:
        # Get cargo items for this container's job
        cargo_items = container.job.cargo_items.all()
        
        if cargo_items.exists():
            # Create transaction for each cargo item
            for cargo in cargo_items:
                # Calculate outbound quantities from delivery orders
                qty_out = 0
                wt_out = 0
                value_out = 0
                
                # Find delivery orders for this container/BOE
                delivery_orders = DeliveryOrder.objects.filter(
                    Q(boe=container.ed_number) | Q(container=container.container_number)
                ).exclude(status='cancelled')
                
                for do in delivery_orders:
                    # Get delivery order items that match this cargo item
                    do_items = do.items.filter(item=cargo.item) if cargo.item else []
                    for do_item in do_items:
                        qty_out += float(do_item.shipped_qty or 0)
                        # Calculate weight proportionally based on shipped quantity
                        if cargo.quantity and cargo.quantity > 0:
                            weight_ratio = float(do_item.shipped_qty or 0) / float(cargo.quantity)
                            wt_out += float(cargo.gross_weight or cargo.net_weight or 0) * weight_ratio
                            value_out += float(cargo.amount or 0) * weight_ratio
                
                transaction_data = {
                    'declaration_no': container.ed_number or '',
                    'bill_no': container.m1_number or '',
                    'date': container.job.created_at.date(),
                    'hs_code': cargo.hs_code or '',
                    'particulars': cargo.item.item_name if cargo.item else (cargo.item_code or f'Container {container.container_number or "N/A"}'),
                    'cog': cargo.coo or '',
                    'pkg_type': cargo.unit or '',
                    'qty_in': float(cargo.quantity or 0),
                    'wt_in': float(cargo.gross_weight or cargo.net_weight or 0),
                    'value_in': float(cargo.amount or 0),
                    'qty_out': qty_out,
                    'wt_out': wt_out,
                    'value_out': value_out,
                    'duty': 0,
                    'total_dues': 0,
                    'job_code': container.job.job_code,
                    'container_number': container.container_number or '',
                }
                transactions.append(transaction_data)
        else:
            # Create placeholder transaction if no cargo items
            # Calculate outbound quantities from delivery orders for this container
            qty_out = 0
            wt_out = 0
            value_out = 0
            
            # Find delivery orders for this container/BOE
            delivery_orders = DeliveryOrder.objects.filter(
                Q(boe=container.ed_number) | Q(container=container.container_number)
            ).exclude(status='cancelled')
            
            for do in delivery_orders:
                # Sum all shipped quantities for this delivery order
                for do_item in do.items.all():
                    qty_out += float(do_item.shipped_qty or 0)
                    # Use item weights if available
                    if do_item.item:
                        wt_out += float(do_item.item.gross_weight or do_item.item.net_weight or 0) * float(do_item.shipped_qty or 0)
                        if do_item.unit_price:
                            value_out += float(do_item.unit_price) * float(do_item.shipped_qty or 0)
            
            transaction_data = {
                'declaration_no': container.ed_number or '',
                'bill_no': container.m1_number or '',
                'date': container.job.created_at.date(),
                'hs_code': '',
                'particulars': f'Container {container.container_number or "N/A"} - {container.job.job_code}',
                'cog': '',
                'pkg_type': container.container_size or '',
                'qty_in': 1,
                'wt_in': 0,
                'value_in': 0,
                'qty_out': qty_out,
                'wt_out': wt_out,
                'value_out': value_out,
                'duty': 0,
                'total_dues': 0,
                'job_code': container.job.job_code,
                'container_number': container.container_number or '',
            }
            transactions.append(transaction_data)
    
    # Apply text-based filters
    if declaration_no:
        transactions = [t for t in transactions if declaration_no.lower() in str(t['declaration_no']).lower()]
    
    if hs_code:
        transactions = [t for t in transactions if hs_code.lower() in str(t['hs_code']).lower()]
    
    if particulars:
        transactions = [t for t in transactions if particulars.lower() in str(t['particulars']).lower()]
    
    if cog:
        transactions = [t for t in transactions if cog.lower() in str(t['cog']).lower()]
    
    # Apply type-based filters
    if filter_type == 'in_boe':
        transactions = [t for t in transactions if t['qty_in'] > 0]
    elif filter_type == 'out_boe':
        transactions = [t for t in transactions if t['qty_out'] > 0]
    elif filter_type == 'balance_boe':
        transactions = [t for t in transactions if t['qty_in'] > 0 or t['qty_out'] > 0]
    
    # Sort transactions by date with current date always last
    current_date = timezone.now().date()
    
    def date_sort_key(transaction):
        trans_date = transaction['date']
        # If transaction date is current date, assign a high sort value to put it last
        if trans_date == current_date:
            return (1, trans_date)  # Current date transactions go last
        else:
            return (0, trans_date)  # Other dates sorted normally
    
    transactions.sort(key=date_sort_key)
    
    # Calculate date range for display
    if from_date and to_date:
        try:
            from_date_display = datetime.strptime(from_date, '%Y-%m-%d').strftime('%d/%m/%Y')
            to_date_display = datetime.strptime(to_date, '%Y-%m-%d').strftime('%d/%m/%Y')
            date_range = f"{from_date_display} to {to_date_display}"
        except ValueError:
            date_range = "Invalid date range"
    elif transactions:
        # Use actual transaction date range
        dates = [t['date'] for t in transactions]
        min_date = min(dates).strftime('%d/%m/%Y')
        max_date = max(dates).strftime('%d/%m/%Y')
        date_range = f"{min_date} to {max_date}"
    else:
        date_range = "No data available"
    
    # Create workbook and worksheet
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Customs BOE Stock Report"
    
    # Define styles
    header_font = Font(bold=True, size=12)
    title_font = Font(bold=True, size=14)
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    center_alignment = Alignment(horizontal='center', vertical='center')
    
    # First row: Company header
    ws.merge_cells('A1:P1')
    ws['A1'] = f'CODE NO:AE-1153161 | COMPANY NAME: ADIRAI FREIGHT SERVICES LLC | DETAILS OF TRANSACTION FROM {date_range}'
    ws['A1'].font = title_font
    ws['A1'].alignment = center_alignment
    
    # Second row: Activity
    ws.merge_cells('A2:P2')
    ws['A2'] = f'Activity: {filter_type.replace("_", " ").title()}'
    ws['A2'].font = header_font
    ws['A2'].alignment = center_alignment
    
    # Third row: Headers
    headers = [
        'Declaration No', 'BILL NOS', 'DATE', 'HS CODE', 'PARTICULARS', 'COG', 'PKG_TYPE',
        'QTY IN', 'WT IN', 'VALUE IN',
        'QTY OUT', 'WT OUT', 'VALUE OUT',
        'AVAILABLE QTY', 'AVAILABLE WT', 'AVAILABLE VALUE',
        'BALANCE QTY', 'BALANCE WT', 'BALANCE VALUE',
        'DUTY @5%', 'TOTAL DUES'
    ]
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.alignment = center_alignment
        cell.border = border
    
    # Data rows
    for row, transaction in enumerate(transactions, 4):
        # Calculate derived values
        available_qty = transaction['qty_in'] - transaction['qty_out']
        available_wt = transaction['wt_in'] - transaction['wt_out']
        available_value = transaction['value_in'] - transaction['value_out']
        balance_qty = available_qty
        balance_wt = available_wt
        balance_value = available_value
        duty_percentage = (transaction['duty'] / transaction['value_in'] * 100) if transaction['value_in'] > 0 else 0
        
        data = [
            transaction['declaration_no'],
            transaction['bill_no'],
            transaction['date'].strftime('%d/%m/%Y') if transaction['date'] else '',
            transaction['hs_code'],
            transaction['particulars'],
            transaction['cog'],
            transaction['pkg_type'],
            transaction['qty_in'],
            transaction['wt_in'],
            transaction['value_in'],
            transaction['qty_out'],
            transaction['wt_out'],
            transaction['value_out'],
            available_qty,
            available_wt,
            available_value,
            balance_qty,
            balance_wt,
            balance_value,
            duty_percentage,
            transaction['total_dues']
        ]
        
        for col, value in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=value)
            cell.border = border
            if col >= 8:  # Numeric columns
                cell.alignment = Alignment(horizontal='right')
    
    # Auto-adjust column widths
    for col in range(1, len(headers) + 1):
        ws.column_dimensions[get_column_letter(col)].width = 15
    
    # Create response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="customs_boe_stock_report_{filter_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
    
    # Save workbook to response
    wb.save(response)
    return response

def export_to_pdf(request):
    """Export BOE transactions to PDF"""
    filter_type = request.GET.get('filter', 'all')
    from_date = request.GET.get('from_date')
    to_date = request.GET.get('to_date')
    declaration_no = request.GET.get('declaration_no', '').strip()
    hs_code = request.GET.get('hs_code', '').strip()
    particulars = request.GET.get('particulars', '').strip()
    cog = request.GET.get('cog', '').strip()
    
    # Get data from Job containers and related GRN items (same logic as main view)
    containers = JobContainer.objects.filter(
        Q(ed_number__isnull=False) | Q(m1_number__isnull=False)
    ).exclude(
        ed_number='', m1_number=''
    ).select_related('job')
    
    # Set default dates if not provided
    if not from_date:
        # Get the earliest declaration date
        earliest_container = containers.order_by('job__created_at').first()
        if earliest_container:
            from_date = earliest_container.job.created_at.date().strftime('%Y-%m-%d')
        else:
            from_date = date.today().strftime('%Y-%m-%d')
    
    if not to_date:
        # Always set to current date
        to_date = date.today().strftime('%Y-%m-%d')
    
    # Apply date filtering if provided
    if from_date:
        try:
            from_date_obj = datetime.strptime(from_date, '%Y-%m-%d').date()
            containers = containers.filter(job__created_at__date__gte=from_date_obj)
        except ValueError:
            pass  # Skip invalid dates in export
    
    if to_date:
        try:
            to_date_obj = datetime.strptime(to_date, '%Y-%m-%d').date()
            containers = containers.filter(job__created_at__date__lte=to_date_obj)
        except ValueError:
            pass  # Skip invalid dates in export
    
    # Build transaction data from containers and their cargo details
    transactions = []
    for container in containers:
        # Get cargo items for this container's job
        cargo_items = container.job.cargo_items.all()
        
        if cargo_items.exists():
            # Create transaction for each cargo item
            for cargo in cargo_items:
                # Calculate outbound quantities from delivery orders
                qty_out = 0
                wt_out = 0
                value_out = 0
                
                # Find delivery orders for this container/BOE
                delivery_orders = DeliveryOrder.objects.filter(
                    Q(boe=container.ed_number) | Q(container=container.container_number)
                ).exclude(status='cancelled')
                
                for do in delivery_orders:
                    # Get delivery order items that match this cargo item
                    do_items = do.items.filter(item=cargo.item) if cargo.item else []
                    for do_item in do_items:
                        qty_out += float(do_item.shipped_qty or 0)
                        # Calculate weight proportionally based on shipped quantity
                        if cargo.quantity and cargo.quantity > 0:
                            weight_ratio = float(do_item.shipped_qty or 0) / float(cargo.quantity)
                            wt_out += float(cargo.gross_weight or cargo.net_weight or 0) * weight_ratio
                            value_out += float(cargo.amount or 0) * weight_ratio
                
                transaction_data = {
                    'declaration_no': container.ed_number or '',
                    'bill_no': container.m1_number or '',
                    'date': container.job.created_at.date(),
                    'hs_code': cargo.hs_code or '',
                    'particulars': cargo.item.item_name if cargo.item else (cargo.item_code or f'Container {container.container_number or "N/A"}'),
                    'cog': cargo.coo or '',
                    'pkg_type': cargo.unit or '',
                    'qty_in': float(cargo.quantity or 0),
                    'wt_in': float(cargo.gross_weight or cargo.net_weight or 0),
                    'value_in': float(cargo.amount or 0),
                    'qty_out': qty_out,
                    'wt_out': wt_out,
                    'value_out': value_out,
                    'duty': 0,
                    'total_dues': 0,
                    'job_code': container.job.job_code,
                    'container_number': container.container_number or '',
                }
                transactions.append(transaction_data)
        else:
            # Create placeholder transaction if no cargo items
            # Calculate outbound quantities from delivery orders for this container
            qty_out = 0
            wt_out = 0
            value_out = 0
            
            # Find delivery orders for this container/BOE
            delivery_orders = DeliveryOrder.objects.filter(
                Q(boe=container.ed_number) | Q(container=container.container_number)
            ).exclude(status='cancelled')
            
            for do in delivery_orders:
                # Sum all shipped quantities for this delivery order
                for do_item in do.items.all():
                    qty_out += float(do_item.shipped_qty or 0)
                    # Use item weights if available
                    if do_item.item:
                        wt_out += float(do_item.item.gross_weight or do_item.item.net_weight or 0) * float(do_item.shipped_qty or 0)
                        if do_item.unit_price:
                            value_out += float(do_item.unit_price) * float(do_item.shipped_qty or 0)
            
            transaction_data = {
                'declaration_no': container.ed_number or '',
                'bill_no': container.m1_number or '',
                'date': container.job.created_at.date(),
                'hs_code': '',
                'particulars': f'Container {container.container_number or "N/A"} - {container.job.job_code}',
                'cog': '',
                'pkg_type': container.container_size or '',
                'qty_in': 1,
                'wt_in': 0,
                'value_in': 0,
                'qty_out': qty_out,
                'wt_out': wt_out,
                'value_out': value_out,
                'duty': 0,
                'total_dues': 0,
                'job_code': container.job.job_code,
                'container_number': container.container_number or '',
            }
            transactions.append(transaction_data)
    
    # Apply text-based filters
    if declaration_no:
        transactions = [t for t in transactions if declaration_no.lower() in str(t['declaration_no']).lower()]
    
    if hs_code:
        transactions = [t for t in transactions if hs_code.lower() in str(t['hs_code']).lower()]
    
    if particulars:
        transactions = [t for t in transactions if particulars.lower() in str(t['particulars']).lower()]
    
    if cog:
        transactions = [t for t in transactions if cog.lower() in str(t['cog']).lower()]
    
    # Apply type-based filters
    if filter_type == 'in_boe':
        transactions = [t for t in transactions if t['qty_in'] > 0]
    elif filter_type == 'out_boe':
        transactions = [t for t in transactions if t['qty_out'] > 0]
    elif filter_type == 'balance_boe':
        transactions = [t for t in transactions if t['qty_in'] > 0 or t['qty_out'] > 0]
    
    # Create date range string based on actual data or filters
    if from_date and to_date:
        date_range = f"{datetime.strptime(from_date, '%Y-%m-%d').strftime('%d/%m/%Y')} to {datetime.strptime(to_date, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    elif from_date:
        date_range = f"From {datetime.strptime(from_date, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    elif to_date:
        date_range = f"Up to {datetime.strptime(to_date, '%Y-%m-%d').strftime('%d/%m/%Y')}"
    else:
        # Use actual date range from transactions
        if transactions:
            dates = [t['date'] for t in transactions]
            min_date = min(dates).strftime('%d/%m/%Y')
            max_date = max(dates).strftime('%d/%m/%Y')
            date_range = f"{min_date} to {max_date}"
        else:
            date_range = 'No data available'
    
    context = {
        'transactions': transactions,
        'filter_type': filter_type.replace('_', ' ').title(),
        'company_name': 'ADIRAI FREIGHT SERVICES LLC',
        'code_no': 'AE-1153161',
        'date_range': date_range,
        'generated_date': datetime.now().strftime('%d/%m/%Y %H:%M:%S')
    }
    
    # Render HTML template
    html_string = render_to_string('customs_BOE_report/pdf_template.html', context)
    
    # Create PDF
    html = HTML(string=html_string)
    pdf = html.write_pdf()
    
    # Create response
    response = HttpResponse(pdf, content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="customs_boe_stock_report_{filter_type}_{datetime.now().strftime("%Y%m%d_%H%M%S")}.pdf"'
    
    return response
