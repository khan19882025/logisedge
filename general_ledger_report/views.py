from django.shortcuts import render, redirect, get_object_or_404
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.http import JsonResponse, HttpResponse
from django.core.paginator import Paginator
from django.db.models import Q, Sum, DecimalField
from django.db.models.functions import Coalesce
from django.utils import timezone
from django.template.loader import render_to_string
from django.views.decorators.csrf import csrf_exempt
from django.contrib.auth.models import User
from chart_of_accounts.models import ChartOfAccount
from company.company_model import Company
from fiscal_year.models import FiscalYear
from ledger.models import Ledger
from .models import GeneralLedgerReport, ReportTemplate
from .forms import GeneralLedgerReportForm, QuickReportForm, ReportTemplateForm
from decimal import Decimal
import json
from datetime import datetime, timedelta


@login_required
def general_ledger_report_list(request):
    """List all general ledger reports"""
    reports = GeneralLedgerReport.objects.filter(created_by=request.user).order_by('-created_at')
    
    # Search functionality
    search_query = request.GET.get('search', '')
    if search_query:
        reports = reports.filter(
            Q(name__icontains=search_query) |
            Q(description__icontains=search_query)
        )
    
    # Pagination
    paginator = Paginator(reports, 10)
    page_number = request.GET.get('page')
    page_obj = paginator.get_page(page_number)
    
    return render(request, 'general_ledger_report/list.html', {
        'page_obj': page_obj,
        'search_query': search_query,
    })


@login_required
def general_ledger_report_create(request):
    """Create a new general ledger report"""
    # Get user's company and active fiscal year
    try:
        company = Company.objects.filter(is_active=True).first()
        if not company:
            messages.error(request, "No company found. Please set up your company first.")
            return redirect('dashboard:dashboard')
        
        active_fiscal_year = FiscalYear.objects.filter(is_current=True).first()
        if not active_fiscal_year:
            messages.error(request, "No active fiscal year found. Please set up your fiscal year first.")
            return redirect('dashboard:dashboard')
    except Exception as e:
        messages.error(request, f"Error loading company/fiscal year: {str(e)}")
        return redirect('dashboard:dashboard')
    
    if request.method == 'POST':
        form = GeneralLedgerReportForm(request.POST, company=company)
        if form.is_valid():
            report = form.save(commit=False)
            report.created_by = request.user
            report.company = company
            report.fiscal_year = active_fiscal_year
            report.save()
            messages.success(request, 'General Ledger Report created successfully!')
            return redirect('general_ledger_report:report_detail', pk=report.pk)
    else:
        form = GeneralLedgerReportForm(company=company)
    
    return render(request, 'general_ledger_report/create.html', {
        'form': form,
        'title': 'Create General Ledger Report',
        'company': company,
        'fiscal_year': active_fiscal_year
    })


@login_required
def general_ledger_report_detail(request, pk):
    """View general ledger report details and generate report"""
    report = get_object_or_404(GeneralLedgerReport, pk=pk, created_by=request.user)
    
    # Generate report data
    report_data = generate_ledger_report_data(report)
    
    return render(request, 'general_ledger_report/detail.html', {
        'report': report,
        'report_data': report_data,
    })


@login_required
def general_ledger_report_quick(request):
    """Quick report generation without saving"""
    if request.method == 'POST':
        form = QuickReportForm(request.POST)
        if form.is_valid():
            # Create temporary report object
            temp_report = GeneralLedgerReport(
                name=form.cleaned_data['name'],
                from_date=form.cleaned_data['from_date'],
                to_date=form.cleaned_data['to_date'],
                account=form.cleaned_data.get('account'),
                report_type=form.cleaned_data['report_type'],
                include_opening_balance=form.cleaned_data['include_opening_balance'],
                include_closing_balance=form.cleaned_data['include_closing_balance'],
                created_by=request.user
            )
            
            # Generate report data
            report_data = generate_ledger_report_data(temp_report)
            
            # Handle export requests
            export_format = request.POST.get('export_format')
            if export_format == 'pdf':
                return export_report_pdf(temp_report, report_data)
            elif export_format == 'excel':
                return export_report_excel(temp_report, report_data)
            elif export_format == 'csv':
                return export_report_csv(temp_report, report_data)
            
            return render(request, 'general_ledger_report/quick_result.html', {
                'report': temp_report,
                'report_data': report_data,
            })
    else:
        form = QuickReportForm()
    
    return render(request, 'general_ledger_report/quick.html', {
        'form': form,
    })


@login_required
def general_ledger_report_export(request, pk):
    """Export general ledger report in various formats"""
    report = get_object_or_404(GeneralLedgerReport, pk=pk, created_by=request.user)
    report_data = generate_ledger_report_data(report)
    
    export_format = request.GET.get('format', 'pdf')
    
    if export_format == 'pdf':
        return export_report_pdf(report, report_data)
    elif export_format == 'excel':
        return export_report_excel(report, report_data)
    elif export_format == 'csv':
        return export_report_csv(report, report_data)
    else:
        messages.error(request, 'Invalid export format')
        return redirect('general_ledger_report:report_detail', pk=pk)


@login_required
def general_ledger_report_save(request, pk):
    """Save a temporary report"""
    report = get_object_or_404(GeneralLedgerReport, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        # Handle saving logic here if needed
        messages.success(request, 'Report saved successfully!')
        return redirect('general_ledger_report:report_detail', pk=pk)
    
    # For GET requests, redirect back to the report detail page
    return redirect('general_ledger_report:report_detail', pk=pk)


@login_required
def general_ledger_report_delete(request, pk):
    """Delete a general ledger report"""
    report = get_object_or_404(GeneralLedgerReport, pk=pk, created_by=request.user)
    
    if request.method == 'POST':
        report.delete()
        messages.success(request, 'Report deleted successfully!')
        return redirect('general_ledger_report:report_list')
    
    return render(request, 'general_ledger_report/delete.html', {
        'report': report,
    })


@login_required
def report_template_list(request):
    """List all report templates"""
    templates = ReportTemplate.objects.filter(created_by=request.user).order_by('-created_at')
    
    return render(request, 'general_ledger_report/template_list.html', {
        'templates': templates,
    })


@login_required
def report_template_create(request):
    """Create a new report template"""
    if request.method == 'POST':
        form = ReportTemplateForm(request.POST)
        if form.is_valid():
            template = form.save(commit=False)
            template.created_by = request.user
            template.save()
            messages.success(request, 'Report template created successfully!')
            return redirect('general_ledger_report:template_list')
    else:
        form = ReportTemplateForm()
    
    return render(request, 'general_ledger_report/template_create.html', {
        'form': form,
    })


@login_required
def report_template_use(request, pk):
    """Use a template to create a new report"""
    template = get_object_or_404(ReportTemplate, pk=pk, created_by=request.user)
    
    # Create form with template data
    initial_data = {
        'name': template.name,
        'from_date': template.default_from_date,
        'to_date': template.default_to_date,
        'include_sub_accounts': template.include_sub_accounts,
        'include_opening_balance': template.include_opening_balance,
        'include_closing_balance': template.include_closing_balance,
        'export_format': template.default_export_format,
    }
    
    form = GeneralLedgerReportForm(initial=initial_data)
    
    return render(request, 'general_ledger_report/create.html', {
        'form': form,
        'title': f'Create Report from Template: {template.name}',
        'template': template,
    })


@csrf_exempt
@login_required
def ajax_get_accounts(request):
    """AJAX endpoint to get accounts based on search term"""
    search_term = request.GET.get('term', '')
    accounts = ChartOfAccount.objects.filter(
        Q(name__icontains=search_term) | Q(account_code__icontains=search_term)
    )[:10]
    
    results = [
        {
            'id': account.id,
            'text': f'{account.account_code} - {account.name}'
        }
        for account in accounts
    ]
    
    return JsonResponse({
        'results': results
    })


def generate_ledger_report_data(report):
    """Generate the actual report data"""
    # Base queryset
    entries = Ledger.objects.filter(
        entry_date__gte=report.from_date,
        entry_date__lte=report.to_date
    ).order_by('entry_date', 'id')
    
    # Filter by account if specified
    if report.account:
        entries = entries.filter(account=report.account)
    
    # Calculate running balance
    running_balance = Decimal('0.00')
    if report.include_opening_balance:
        running_balance = calculate_opening_balance(report)
    
    # Process entries and calculate running balance
    processed_entries = []
    for entry in entries:
        # Handle None amounts by treating them as 0
        amount = entry.amount or Decimal('0.00')
        if entry.entry_type == 'DR':
            running_balance += amount
        else:  # CR
            running_balance -= amount
        
        # Add calculated balance to entry
        entry.running_balance = running_balance
        entry.calculated_balance = running_balance
        processed_entries.append(entry)
    
    # Calculate totals
    total_debit = entries.filter(entry_type='DR').aggregate(
        total=Coalesce(Sum('amount'), Decimal('0.00'), output_field=DecimalField())
    )['total']
    
    total_credit = entries.filter(entry_type='CR').aggregate(
        total=Coalesce(Sum('amount'), Decimal('0.00'), output_field=DecimalField())
    )['total']
    
    net_movement = total_debit - total_credit
    
    opening_balance = Decimal('0.00')
    if report.include_opening_balance:
        opening_balance = calculate_opening_balance(report)
    
    closing_balance = opening_balance + net_movement
    
    return {
        'entries': processed_entries,
        'total_debit': total_debit,
        'total_credit': total_credit,
        'net_movement': net_movement,
        'opening_balance': opening_balance,
        'closing_balance': closing_balance,
        'entry_count': len(processed_entries),
    }


def calculate_opening_balance(report):
    """Calculate opening balance for the report period"""
    opening_balance = Decimal('0.00')
    
    # Get all entries before the report start date
    previous_entries = Ledger.objects.filter(
        entry_date__lt=report.from_date
    )
    
    if report.account:
        previous_entries = previous_entries.filter(account=report.account)
    
    # Calculate the balance
    for entry in previous_entries:
        # Handle None amounts by treating them as 0
        amount = entry.amount or Decimal('0.00')
        if entry.entry_type == 'DR':
            opening_balance += amount
        else:  # CR
            opening_balance -= amount
    
    return opening_balance


def get_user_company(user):
    """Get the company associated with the user"""
    try:
        return Company.objects.get(created_by=user)
    except Company.DoesNotExist:
        return None


def get_current_fiscal_year(company):
    """Get the current fiscal year for the company"""
    try:
        return FiscalYear.objects.get(company=company, is_active=True)
    except FiscalYear.DoesNotExist:
        return None


def export_report_pdf(report, report_data):
    """Export General Ledger Report as professional PDF"""
    
    try:
        from reportlab.lib.pagesizes import letter, A4, landscape
        from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
        from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
        from reportlab.lib.units import inch, cm
        from reportlab.lib import colors
        from datetime import datetime
    except ImportError:
        raise ImportError('reportlab is required for PDF export')
    
    # Create response
    response = HttpResponse(content_type='application/pdf')
    filename = f"General_Ledger_Report_{report.from_date}_to_{report.to_date}.pdf"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    try:
        # Create PDF document with landscape orientation and smaller margins
        doc = SimpleDocTemplate(
            response, 
            pagesize=landscape(A4),
            rightMargin=1*cm,
            leftMargin=1*cm,
            topMargin=1.5*cm,
            bottomMargin=1.5*cm
        )
        elements = []
        
        # Professional Styles
        styles = getSampleStyleSheet()
        
        # Company header style - left aligned
        company_style = ParagraphStyle(
            'CompanyHeader',
            parent=styles['Normal'],
            fontSize=14,
            spaceAfter=3,
            alignment=0,  # Left alignment
            textColor=colors.HexColor('#1a365d'),
            fontName='Helvetica-Bold'
        )
        
        # Company details style - left aligned
        company_details_style = ParagraphStyle(
            'CompanyDetails',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=2,
            alignment=0,  # Left alignment
            textColor=colors.HexColor('#4a5568')
        )
        
        # Report title style - left aligned
        title_style = ParagraphStyle(
            'ReportTitle',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=5,
            spaceBefore=15,
            alignment=0,  # Left alignment
            textColor=colors.HexColor('#2d3748'),
            fontName='Helvetica-Bold'
        )
        
        # Period style - left aligned
        period_style = ParagraphStyle(
            'Period',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=3,
            alignment=0,  # Left alignment
            textColor=colors.HexColor('#4a5568'),
            fontName='Helvetica-Bold'
        )
        
        # Report info style - left aligned
        info_style = ParagraphStyle(
            'ReportInfo',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=15,
            alignment=0,  # Left alignment
            textColor=colors.HexColor('#4a5568')
        )
        
        # Right-aligned styles for report details
        info_right_style = ParagraphStyle(
            'ReportInfoRight',
            parent=styles['Normal'],
            fontSize=10,
            spaceAfter=3,
            alignment=2,  # Right alignment
            textColor=colors.HexColor('#4a5568')
        )
        
        title_right_style = ParagraphStyle(
            'ReportTitleRight',
            parent=styles['Heading1'],
            fontSize=16,
            spaceAfter=5,
            alignment=2,  # Right alignment
            textColor=colors.HexColor('#2d3748'),
            fontName='Helvetica-Bold'
        )
        
        period_right_style = ParagraphStyle(
            'PeriodRight',
            parent=styles['Normal'],
            fontSize=11,
            spaceAfter=3,
            alignment=2,  # Right alignment
            textColor=colors.HexColor('#4a5568'),
            fontName='Helvetica-Bold'
        )
        
        # Company information (left side)
        company_info = [
            Paragraph("<b>ADIRAI FREIGHT SERVICE LLC (BR)</b>", company_style),
            Paragraph("JAFZA SOUTH2, DUBAI.UAE", company_details_style),
            Paragraph("Tel: +971 4 8808477 | Email: info@adiraifreight.com", company_details_style)
        ]
        
        # Report details (right side)
        report_details = [
            Paragraph("<b>GENERAL LEDGER REPORT</b>", title_right_style),
            Paragraph(f"For the period from {report.from_date} to {report.to_date}", period_right_style),
            Paragraph(f"<b>Report Name:</b> {report.name}", info_right_style),
            Paragraph(f"<b>Report Type:</b> {report.get_report_type_display()}", info_right_style)
        ]
        
        if report.account:
            report_details.append(Paragraph(f"<b>Account:</b> {report.account.account_code} - {report.account.name}", info_right_style))
        else:
            report_details.append(Paragraph(f"<b>Account:</b> All Accounts", info_right_style))
        
        # Create two-column header table
        header_table_data = []
        max_rows = max(len(company_info), len(report_details))
        
        for i in range(max_rows):
            left_cell = company_info[i] if i < len(company_info) else ""
            right_cell = report_details[i] if i < len(report_details) else ""
            header_table_data.append([left_cell, right_cell])
        
        # Create and style the header table
        header_table = Table(header_table_data, colWidths=[4*inch, 4*inch])
        header_table.setStyle(TableStyle([
            ('VALIGN', (0, 0), (-1, -1), 'TOP'),
            ('ALIGN', (0, 0), (0, -1), 'LEFT'),   # Left column left-aligned
            ('ALIGN', (1, 0), (1, -1), 'RIGHT'),  # Right column right-aligned
            ('LEFTPADDING', (0, 0), (-1, -1), 0),
            ('RIGHTPADDING', (0, 0), (-1, -1), 0),
            ('TOPPADDING', (0, 0), (-1, -1), 0),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 2),
        ]))
        
        elements.append(header_table)
        
        # Add spacing after header
        elements.append(Spacer(1, 20))
        
        # Generation timestamp
        timestamp_style = ParagraphStyle(
            'Timestamp',
            parent=styles['Normal'],
            fontSize=8,
            spaceAfter=15,
            alignment=0,  # Left alignment
            textColor=colors.HexColor('#718096')
        )
        timestamp = Paragraph(f"Generated on: {datetime.now().strftime('%B %d, %Y at %I:%M %p')}", timestamp_style)
        elements.append(timestamp)
        
        # Add spacing before table
        elements.append(Spacer(1, 10))
        
        # Helper function to format amounts
        def format_amount(amount):
            if amount is None:
                return '0.00'
            return f"{amount:,.2f}"
        
        # Group entries by account for separate tables
        from collections import defaultdict
        entries_by_account = defaultdict(list)
        
        for entry in report_data['entries']:
            try:
                account_key = entry.account.id if entry.account else 'no_account'
                entries_by_account[account_key].append(entry)
            except AttributeError:
                entries_by_account['no_account'].append(entry)
        
        # If no specific account is selected and we have multiple accounts, create separate tables
        if not report.account and len(entries_by_account) > 1:
            # Create separate table for each account
            for account_key, account_entries in entries_by_account.items():
                if not account_entries:
                    continue
                    
                # Get account information
                first_entry = account_entries[0]
                try:
                    account_code = first_entry.account.account_code if first_entry.account else 'N/A'
                    account_name = first_entry.account.name if first_entry.account else 'No Account'
                except AttributeError:
                    account_code = 'N/A'
                    account_name = 'No Account'
                
                # Add account title
                account_title_style = ParagraphStyle(
                    'AccountTitle',
                    parent=styles['Normal'],
                    fontSize=14,
                    spaceAfter=10,
                    spaceBefore=20,
                    alignment=0,  # Left alignment
                    textColor=colors.HexColor('#2d3748'),
                    fontName='Helvetica-Bold'
                )
                
                account_title = Paragraph(f"<b>Account: {account_code} - {account_name}</b>", account_title_style)
                elements.append(account_title)
                
                # Create table for this account
                table_data = []
                
                # Table headers
                headers = ['Date', 'Ledger Code', 'Account Name', 'Voucher No.', 'Description', 'Debit (AED)', 'Credit (AED)', 'Balance (AED)']
                table_data.append(headers)
                
                # Calculate opening balance for this account
                account_opening_balance = Decimal('0.00')
                if report.include_opening_balance:
                    # Calculate opening balance for this specific account
                    opening_entries = Ledger.objects.filter(
                        account=first_entry.account,
                        entry_date__lt=report.from_date
                    ).order_by('entry_date')
                    
                    for opening_entry in opening_entries:
                        amount = opening_entry.amount or Decimal('0.00')
                        if opening_entry.entry_type == 'DR':
                            account_opening_balance += amount
                        else:
                            account_opening_balance -= amount
                
                # Add opening balance row if enabled and non-zero
                if report.include_opening_balance and account_opening_balance != 0:
                    opening_row = [
                        '',
                        '',
                        '',
                        '',
                        'Opening Balance',
                        '',
                        '',
                        format_amount(account_opening_balance)
                    ]
                    table_data.append(opening_row)
                
                # Calculate running balance for this account
                running_balance = account_opening_balance
                
                # Add data rows for this account
                for entry in account_entries:
                    debit_amount = format_amount(entry.amount) if entry.entry_type == 'DR' else ''
                    credit_amount = format_amount(entry.amount) if entry.entry_type == 'CR' else ''
                    
                    # Update running balance
                    amount = entry.amount or Decimal('0.00')
                    if entry.entry_type == 'DR':
                        running_balance += amount
                    else:
                        running_balance -= amount
                    
                    row = [
                         entry.entry_date.strftime('%d/%m/%Y'),
                         account_code,
                         account_name,
                         entry.ledger_number or '',
                         entry.description[:40] + '...' if len(entry.description) > 40 else entry.description,
                         debit_amount,
                         credit_amount,
                         format_amount(running_balance)
                     ]
                    table_data.append(row)
                
                # Add closing balance row if enabled
                if report.include_closing_balance:
                    closing_row = [
                        '',
                        '',
                        '',
                        '',
                        'Closing Balance',
                        '',
                        '',
                        format_amount(running_balance)
                    ]
                    table_data.append(closing_row)
                
                # Create and style the table for this account
                if table_data:
                    col_widths = [0.8*inch, 0.8*inch, 1.5*inch, 1*inch, 2*inch, 1*inch, 1*inch, 1*inch]
                    
                    table = Table(table_data, colWidths=col_widths, repeatRows=1)
                    
                    # Enhanced table styling
                    style_commands = [
                        # Header styling
                        ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
                        ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                        ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                        ('FONTSIZE', (0, 0), (-1, 0), 10),
                        ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                        ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                        ('TOPPADDING', (0, 0), (-1, 0), 12),
                        
                        # Data rows styling
                        ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                        ('FONTSIZE', (0, 1), (-1, -1), 9),
                        ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                        ('ALIGN', (5, 1), (-1, -1), 'RIGHT'),  # Right align amounts
                        ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                        ('LEFTPADDING', (0, 0), (-1, -1), 6),
                        ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                        ('TOPPADDING', (0, 1), (-1, -1), 4),
                        ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                        
                        # Grid and borders
                        ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2d3748')),
                        ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                    ]
                    
                    # Special styling for opening/closing balance rows
                    for i, row in enumerate(table_data):
                        if i > 0 and len(row) > 4:  # Skip header row
                            if 'Opening Balance' in row[4] or 'Closing Balance' in row[4]:
                                style_commands.extend([
                                    ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f7fafc')),
                                    ('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'),
                                    ('LINEABOVE', (0, i), (-1, i), 1, colors.HexColor('#4a5568')),
                                    ('LINEBELOW', (0, i), (-1, i), 1, colors.HexColor('#4a5568')),
                                ])
                    
                    # Apply styling
                    table.setStyle(TableStyle(style_commands))
                    elements.append(table)
                    
                    # Add spacing between tables
                    elements.append(Spacer(1, 20))
        else:
            # Single table for single account or all accounts combined
            table_data = []
            
            # Table headers
            headers = ['Date', 'Ledger Code', 'Account Name', 'Voucher No.', 'Description', 'Debit (AED)', 'Credit (AED)', 'Balance (AED)']
            table_data.append(headers)
            
            # Add opening balance row if enabled
            if report.include_opening_balance and report_data['opening_balance'] != 0:
                opening_row = [
                    '',
                    '',
                    '',
                    '',
                    'Opening Balance',
                    '',
                    '',
                    format_amount(report_data['opening_balance'])
                ]
                table_data.append(opening_row)
            
            # Add data rows
            for entry in report_data['entries']:
                debit_amount = format_amount(entry.amount) if entry.entry_type == 'DR' else ''
                credit_amount = format_amount(entry.amount) if entry.entry_type == 'CR' else ''
                
                # Safely access account data
                try:
                    account_code = entry.account.account_code if entry.account else '-'
                    account_name = entry.account.name if entry.account else '-'
                except AttributeError:
                    account_code = '-'
                    account_name = '-'
                
                row = [
                     entry.entry_date.strftime('%d/%m/%Y'),
                     account_code,
                     account_name,
                     entry.ledger_number or '',
                     entry.description[:40] + '...' if len(entry.description) > 40 else entry.description,
                     debit_amount,
                     credit_amount,
                     format_amount(entry.running_balance)
                 ]
                table_data.append(row)
            
            # Add closing balance row if enabled
            if report.include_closing_balance:
                closing_row = [
                    '',
                    '',
                    '',
                    '',
                    'Closing Balance',
                    '',
                    '',
                    format_amount(report_data['closing_balance'])
                ]
                table_data.append(closing_row)
            
            # Create and style the main table
            if table_data:
                # Adjust column widths for landscape orientation with new columns
                col_widths = [0.8*inch, 0.8*inch, 1.5*inch, 1*inch, 2*inch, 1*inch, 1*inch, 1*inch]
                
                table = Table(table_data, colWidths=col_widths, repeatRows=1)
                
                # Enhanced table styling
                style_commands = [
                    # Header styling
                    ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
                    ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
                    ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                    ('FONTSIZE', (0, 0), (-1, 0), 10),
                    ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
                    ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                    ('TOPPADDING', (0, 0), (-1, 0), 12),
                    
                    # Data rows styling
                    ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
                    ('FONTSIZE', (0, 1), (-1, -1), 9),
                    ('ALIGN', (0, 1), (-1, -1), 'LEFT'),
                    ('ALIGN', (5, 1), (-1, -1), 'RIGHT'),  # Right align amounts (Debit, Credit, Balance)
                    ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
                    ('LEFTPADDING', (0, 0), (-1, -1), 6),
                    ('RIGHTPADDING', (0, 0), (-1, -1), 6),
                    ('TOPPADDING', (0, 1), (-1, -1), 4),
                    ('BOTTOMPADDING', (0, 1), (-1, -1), 4),
                    
                    # Grid and borders
                    ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2d3748')),
                    ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
                ]
                
                # Special styling for opening/closing balance rows
                for i, row in enumerate(table_data):
                    if i > 0 and len(row) > 4:  # Skip header row
                        if 'Opening Balance' in row[4] or 'Closing Balance' in row[4]:
                            style_commands.extend([
                                ('BACKGROUND', (0, i), (-1, i), colors.HexColor('#f7fafc')),
                                ('FONTNAME', (0, i), (-1, i), 'Helvetica-Bold'),
                                ('LINEABOVE', (0, i), (-1, i), 1, colors.HexColor('#4a5568')),
                                ('LINEBELOW', (0, i), (-1, i), 1, colors.HexColor('#4a5568')),
                            ])
                
                # Apply styling
                table.setStyle(TableStyle(style_commands))
                elements.append(table)
        
        # Add spacing before summary
        elements.append(Spacer(1, 25))
        
        # Enhanced Summary Section with two columns
        summary_data = [
            ['FINANCIAL SUMMARY', '', 'REPORT STATISTICS', ''],
        ]
        
        # Financial metrics (left side)
        financial_metrics = []
        if report.include_opening_balance:
            financial_metrics.append(['Opening Balance:', f"{format_amount(report_data['opening_balance'])} AED"])
        
        financial_metrics.extend([
            ['Total Debit:', f"{format_amount(report_data['total_debit'])} AED"],
            ['Total Credit:', f"{format_amount(report_data['total_credit'])} AED"],
            ['Net Movement:', f"{format_amount(report_data['net_movement'])} AED"]
        ])
        
        if report.include_closing_balance:
            financial_metrics.append(['Closing Balance:', f"{format_amount(report_data['closing_balance'])} AED"])
        
        # Report statistics (right side)
        report_stats = [
            ['Total Entries:', str(report_data['entry_count'])],
            ['Report Period:', f"{(report.to_date - report.from_date).days + 1} days"],
            ['Account Type:', 'Specific Account' if report.account else 'All Accounts'],
            ['Generated:', datetime.now().strftime('%d/%m/%Y %H:%M')]
        ]
        
        # Combine data into table format
        max_rows = max(len(financial_metrics), len(report_stats))
        for i in range(max_rows):
            left_label = financial_metrics[i][0] if i < len(financial_metrics) else ''
            left_value = financial_metrics[i][1] if i < len(financial_metrics) else ''
            right_label = report_stats[i][0] if i < len(report_stats) else ''
            right_value = report_stats[i][1] if i < len(report_stats) else ''
            summary_data.append([left_label, left_value, right_label, right_value])
        
        summary_table = Table(summary_data, colWidths=[2.5*inch, 2.5*inch, 2.5*inch, 2.5*inch])
        summary_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (1, 0), colors.HexColor('#2d3748')),
            ('BACKGROUND', (2, 0), (3, 0), colors.HexColor('#4a5568')),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 11),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            
            # Data styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 10),
            ('ALIGN', (0, 1), (0, -1), 'LEFT'),    # Financial labels
            ('ALIGN', (1, 1), (1, -1), 'RIGHT'),   # Financial values
            ('ALIGN', (2, 1), (2, -1), 'LEFT'),    # Stats labels
            ('ALIGN', (3, 1), (3, -1), 'RIGHT'),   # Stats values
            
            # Borders and spacing
            ('GRID', (0, 0), (-1, -1), 0.5, colors.HexColor('#e2e8f0')),
            ('LINEBELOW', (0, 0), (-1, 0), 2, colors.HexColor('#2d3748')),
            ('TOPPADDING', (0, 0), (-1, -1), 10),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 10),
            ('LEFTPADDING', (0, 0), (-1, -1), 12),
            ('RIGHTPADDING', (0, 0), (-1, -1), 12),
            
            # Highlight net movement row
            ('BACKGROUND', (0, -2), (1, -1), colors.HexColor('#f7fafc')),  # Net movement row
            ('FONTNAME', (0, -2), (1, -1), 'Helvetica-Bold'),
        ]))
        
        elements.append(summary_table)
        
        # Add spacing before signature
        elements.append(Spacer(1, 30))
        
        # Enhanced Signature Section
        auth_header_style = ParagraphStyle(
            'AuthHeader',
            parent=styles['Normal'],
            fontSize=12,
            spaceAfter=15,
            alignment=1,  # Center alignment
            textColor=colors.HexColor('#2d3748'),
            fontName='Helvetica-Bold'
        )
        
        auth_header = Paragraph("<b>REPORT AUTHORIZATION</b>", auth_header_style)
        elements.append(auth_header)
        
        signature_table_data = [
            ['PREPARED BY', '', 'REVIEWED BY', '', 'APPROVED BY'],
            ['', '', '', '', ''],
            ['', '', '', '', ''],
            ['_____________________', '', '_____________________', '', '_____________________'],
            ['Name & Signature', '', 'Name & Signature', '', 'Name & Signature'],
            ['', '', '', '', ''],
            ['Date: _______________', '', 'Date: _______________', '', 'Date: _______________'],
            ['', '', '', '', ''],
            ['Position: ____________', '', 'Position: ____________', '', 'Position: ____________']
        ]
        
        signature_table = Table(signature_table_data, colWidths=[2.2*inch, 0.6*inch, 2.2*inch, 0.6*inch, 2.2*inch])
        signature_table.setStyle(TableStyle([
            # Header styling
            ('BACKGROUND', (0, 0), (-1, 0), colors.HexColor('#f7fafc')),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 10),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.HexColor('#2d3748')),
            
            # Content styling
            ('FONTNAME', (0, 1), (-1, -1), 'Helvetica'),
            ('FONTSIZE', (0, 1), (-1, -1), 9),
            ('ALIGN', (0, 1), (-1, -1), 'CENTER'),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
            
            # Borders and spacing
            ('LINEBELOW', (0, 0), (-1, 0), 1, colors.HexColor('#2d3748')),
            ('TOPPADDING', (0, 0), (-1, -1), 8),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 8),
            ('LEFTPADDING', (0, 0), (-1, -1), 5),
            ('RIGHTPADDING', (0, 0), (-1, -1), 5),
        ]))
        
        elements.append(signature_table)
        
        # Build PDF
        doc.build(elements)
        return response
    
    except Exception as e:
        # Handle any errors gracefully
        response = HttpResponse(content_type='application/pdf')
        response['Content-Disposition'] = 'attachment; filename="error_report.pdf"'
        return response


def export_report_excel(report, report_data):
    """Export General Ledger Report as professional Excel file"""
    
    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        from datetime import datetime
        import io
    except ImportError:
        raise ImportError('openpyxl is required for Excel export')
    
    # Create workbook and worksheet
    wb = Workbook()
    ws = wb.active
    ws.title = "General Ledger Report"
    
    # Define styles
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="2D3748", end_color="2D3748", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    title_font = Font(bold=True, size=16)
    subtitle_font = Font(bold=True, size=12)
    normal_font = Font(size=10)
    
    border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # Helper function to format amounts
    def format_amount(amount):
        if amount is None:
            return 0.00
        return float(amount)
    
    # Add title and company information
    current_row = 1
    
    # Company name and title
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = f"{report.company.name if report.company else 'Company Name'}"
    ws[f'A{current_row}'].font = title_font
    ws[f'A{current_row}'].alignment = Alignment(horizontal="center")
    current_row += 1
    
    ws.merge_cells(f'A{current_row}:F{current_row}')
    ws[f'A{current_row}'] = "General Ledger Report"
    ws[f'A{current_row}'].font = subtitle_font
    ws[f'A{current_row}'].alignment = Alignment(horizontal="center")
    current_row += 2
    
    # Report details
    ws[f'A{current_row}'] = "Report Name:"
    ws[f'A{current_row}'].font = Font(bold=True)
    ws[f'B{current_row}'] = report.name
    current_row += 1
    
    ws[f'A{current_row}'] = "Period:"
    ws[f'A{current_row}'].font = Font(bold=True)
    ws[f'B{current_row}'] = f"{report.from_date.strftime('%d/%m/%Y')} to {report.to_date.strftime('%d/%m/%Y')}"
    current_row += 1
    
    if report.account:
        ws[f'A{current_row}'] = "Account:"
        ws[f'A{current_row}'].font = Font(bold=True)
        ws[f'B{current_row}'] = f"{report.account.account_code} - {report.account.name}"
        current_row += 1
    
    ws[f'A{current_row}'] = "Generated:"
    ws[f'A{current_row}'].font = Font(bold=True)
    ws[f'B{current_row}'] = datetime.now().strftime('%d/%m/%Y %I:%M %p')
    current_row += 2
    
    # Table headers
    headers = ['Date', 'Ledger Code', 'Account Name', 'Voucher No.', 'Description', 'Debit (AED)', 'Credit (AED)', 'Balance (AED)']
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=current_row, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
    
    current_row += 1
    
    # Add opening balance if enabled
    if report.include_opening_balance and report_data['opening_balance'] != 0:
        ws.cell(row=current_row, column=1, value="")
        ws.cell(row=current_row, column=2, value="")
        ws.cell(row=current_row, column=3, value="")
        ws.cell(row=current_row, column=4, value="")
        ws.cell(row=current_row, column=5, value="Opening Balance").font = Font(bold=True)
        ws.cell(row=current_row, column=6, value="")
        ws.cell(row=current_row, column=7, value="")
        ws.cell(row=current_row, column=8, value=format_amount(report_data['opening_balance']))
        
        # Apply borders
        for col in range(1, 9):
            ws.cell(row=current_row, column=col).border = border
        
        current_row += 1
    
    # Add data rows
    for entry in report_data['entries']:
        debit_amount = format_amount(entry.amount) if entry.entry_type == 'DR' else ""
        credit_amount = format_amount(entry.amount) if entry.entry_type == 'CR' else ""
        
        ws.cell(row=current_row, column=1, value=entry.entry_date.strftime('%d/%m/%Y'))
        ws.cell(row=current_row, column=2, value=entry.account.account_code if entry.account else "")
        ws.cell(row=current_row, column=3, value=entry.account.name if entry.account else "")
        ws.cell(row=current_row, column=4, value=entry.ledger_number or "")
        ws.cell(row=current_row, column=5, value=entry.description)
        ws.cell(row=current_row, column=6, value=debit_amount)
        ws.cell(row=current_row, column=7, value=credit_amount)
        ws.cell(row=current_row, column=8, value=format_amount(entry.running_balance))
        
        # Apply borders and formatting
        for col in range(1, 9):
            cell = ws.cell(row=current_row, column=col)
            cell.border = border
            cell.font = normal_font
            
            # Right align numeric columns
            if col in [6, 7, 8]:
                cell.alignment = Alignment(horizontal="right")
        
        current_row += 1
    
    # Add closing balance if enabled
    if report.include_closing_balance:
        ws.cell(row=current_row, column=1, value="")
        ws.cell(row=current_row, column=2, value="")
        ws.cell(row=current_row, column=3, value="")
        ws.cell(row=current_row, column=4, value="")
        ws.cell(row=current_row, column=5, value="Closing Balance").font = Font(bold=True)
        ws.cell(row=current_row, column=6, value="")
        ws.cell(row=current_row, column=7, value="")
        ws.cell(row=current_row, column=8, value=format_amount(report_data['closing_balance']))
        
        # Apply borders
        for col in range(1, 9):
            ws.cell(row=current_row, column=col).border = border
    
    # Auto-adjust column widths
    column_widths = [12, 12, 25, 15, 30, 15, 15, 15]  # Approximate widths
    for col, width in enumerate(column_widths, 1):
        ws.column_dimensions[get_column_letter(col)].width = width
    
    # Create response
    response = HttpResponse(content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')
    filename = f"General_Ledger_Report_{report.from_date}_to_{report.to_date}.xlsx"
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    excel_file = io.BytesIO()
    wb.save(excel_file)
    excel_file.seek(0)
    response.write(excel_file.getvalue())
    
    return response


def export_report_csv(report, report_data):
    """Export report as CSV"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="general_ledger_report_{report.pk}.csv"'
    
    response.write("Date,Ledger Code,Account Name,Voucher Number,Description,Debit,Credit,Balance\n")
    
    for entry in report_data['entries']:
        debit = entry.amount if entry.entry_type == 'DR' else ''
        credit = entry.amount if entry.entry_type == 'CR' else ''
        ledger_code = entry.account.account_code if entry.account else ''
        account_name = entry.account.name if entry.account else ''
        response.write(f"{entry.entry_date},{ledger_code},{account_name},{entry.ledger_number},{entry.description},{debit},{credit},{entry.calculated_balance}\n")
    
    return response
