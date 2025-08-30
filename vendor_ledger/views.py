from django.shortcuts import render, get_object_or_404
from django.http import JsonResponse, HttpResponse
from django.contrib.auth.decorators import login_required
from django.contrib import messages
from django.db.models import Q, Sum, F
from django.utils import timezone
from datetime import datetime, timedelta
import json
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side
from io import BytesIO

from customer.models import Customer
from invoice.models import Invoice
from supplier_payments.models import SupplierPayment, SupplierPaymentInvoice
from customer_payments.models import CustomerPayment, CustomerPaymentInvoice
from .forms import VendorLedgerFilterForm, QuickFilterForm
from .models import VendorLedgerReport


@login_required
def vendor_ledger_report(request):
    """Main view for vendor ledger report"""
    form = VendorLedgerFilterForm()
    quick_filter_form = QuickFilterForm()
    
    context = {
        'form': form,
        'quick_filter_form': quick_filter_form,
        'report_data': None,
        'total_summary': None,
        'page_title': 'Vendor Ledger Report',
        'breadcrumb': [
            {'name': 'Dashboard', 'url': '/'},
            {'name': 'Reports', 'url': '/reports/'},
            {'name': 'Vendor Ledger Report', 'url': None}
        ]
    }
    
    if request.method == 'POST':
        form = VendorLedgerFilterForm(request.POST)
        if form.is_valid():
            # Generate report data
            report_data, total_summary = generate_vendor_ledger_data(
                vendor=form.cleaned_data.get('vendor'),
                date_from=form.cleaned_data['date_from'],
                date_to=form.cleaned_data['date_to'],
                payment_status=form.cleaned_data.get('payment_status', 'all')
            )
            
            context.update({
                'report_data': report_data,
                'total_summary': total_summary,
                'form': form
            })
            
            if not report_data:
                messages.info(request, 'No data found for the selected criteria.')
        else:
            messages.error(request, 'Please correct the errors in the form.')
    
    return render(request, 'vendor_ledger/report.html', context)


def generate_vendor_ledger_data(vendor=None, date_from=None, date_to=None, payment_status='all'):
    """Generate vendor ledger report data"""
    
    # Get vendors based on filter
    if vendor:
        vendors = [vendor]
    else:
        # Filter vendors by checking customer_types field content
        vendors = Customer.objects.filter(
            is_active=True
        ).order_by('customer_name')
        # Filter vendors that have 'vendor' in their customer_types
        vendors = [v for v in vendors if v.customer_types and 'vendor' in v.customer_types.lower()]
    
    report_data = []
    total_debit = 0
    total_credit = 0
    
    for vendor_obj in vendors:
        vendor_data = {
            'vendor': vendor_obj,
            'transactions': [],
            'total_debit': 0,
            'total_credit': 0,
            'balance': 0
        }
        
        running_balance = 0
        
        # Get invoices where this vendor appears in invoice items
        invoices = Invoice.objects.filter(
            invoice_date__range=[date_from, date_to]
        ).order_by('invoice_date')
        
        for invoice in invoices:
            # Check if vendor appears in invoice items
            vendor_amount = 0
            vendor_code = vendor_obj.customer_code
            vendor_name = vendor_obj.customer_name
            
            for item in invoice.invoice_items:
                vendor_field = item.get('vendor', '')
                if vendor_code in vendor_field or vendor_name in vendor_field:
                    try:
                        item_total = float(item.get('total', 0))
                        vendor_amount += item_total
                    except (ValueError, TypeError):
                        continue
            
            if vendor_amount > 0:
                # Check payment status if filtering is required
                if payment_status != 'all':
                    # Get payments for this invoice
                    paid_amount = SupplierPaymentInvoice.objects.filter(
                        invoice=invoice
                    ).aggregate(total=Sum('allocated_amount'))['total'] or 0
                    
                    invoice_status = 'pending'
                    if paid_amount >= vendor_amount:
                        invoice_status = 'fully_paid'
                    elif paid_amount > 0:
                        invoice_status = 'partially_paid'
                    
                    if payment_status != invoice_status:
                        continue
                
                running_balance += vendor_amount
                vendor_data['transactions'].append({
                    'date': invoice.invoice_date,
                    'description': f"Invoice {invoice.invoice_number} - Vendor Items",
                    'reference': invoice.invoice_number,
                    'debit': vendor_amount,
                    'credit': 0,
                    'balance': running_balance,
                    'type': 'invoice'
                })
                
                vendor_data['total_debit'] += vendor_amount
        
        # Get supplier payments for this vendor
        supplier_payments = SupplierPayment.objects.filter(
            supplier=vendor_obj,
            payment_date__range=[date_from, date_to]
        ).order_by('payment_date')
        
        for payment in supplier_payments:
            amount = payment.amount or 0
            running_balance -= amount
            
            vendor_data['transactions'].append({
                'date': payment.payment_date,
                'description': f"Payment - {payment.notes or 'Supplier Payment'}",
                'reference': payment.payment_id or f"SP-{payment.id}",
                'debit': 0,
                'credit': amount,
                'balance': running_balance,
                'type': 'payment'
            })
            
            vendor_data['total_credit'] += amount
        
        # Sort transactions by date
        vendor_data['transactions'].sort(key=lambda x: x['date'])
        
        # Calculate final balance
        vendor_data['balance'] = vendor_data['total_debit'] - vendor_data['total_credit']
        
        # Only include vendors with transactions
        if vendor_data['transactions']:
            report_data.append(vendor_data)
            total_debit += vendor_data['total_debit']
            total_credit += vendor_data['total_credit']
    
    total_summary = {
        'total_debit': total_debit,
        'total_credit': total_credit,
        'net_balance': total_debit - total_credit,
        'vendor_count': len(report_data)
    }
    
    return report_data, total_summary


@login_required
def export_excel(request):
    """Export vendor ledger report to Excel"""
    if request.method == 'POST':
        form = VendorLedgerFilterForm(request.POST)
        if form.is_valid():
            report_data, total_summary = generate_vendor_ledger_data(
                vendor=form.cleaned_data.get('vendor'),
                date_from=form.cleaned_data['date_from'],
                date_to=form.cleaned_data['date_to'],
                payment_status=form.cleaned_data.get('payment_status', 'all')
            )
            
            # Create Excel workbook
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = "Vendor Ledger Report"
            
            # Header styling
            header_font = Font(bold=True, size=12)
            header_alignment = Alignment(horizontal='center')
            border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Report title
            ws.merge_cells('A1:G1')
            ws['A1'] = 'Vendor Ledger Report'
            ws['A1'].font = Font(bold=True, size=16)
            ws['A1'].alignment = header_alignment
            
            # Date range
            ws.merge_cells('A2:G2')
            ws['A2'] = f"Period: {form.cleaned_data['date_from']} to {form.cleaned_data['date_to']}"
            ws['A2'].alignment = header_alignment
            
            row = 4
            
            for vendor_data in report_data:
                # Vendor header
                ws.merge_cells(f'A{row}:G{row}')
                ws[f'A{row}'] = f"Vendor: {vendor_data['vendor'].customer_name}"
                ws[f'A{row}'].font = header_font
                row += 1
                
                # Column headers
                headers = ['Date', 'Description', 'Reference', 'Debit', 'Credit', 'Balance', 'Type']
                for col, header in enumerate(headers, 1):
                    cell = ws.cell(row=row, column=col, value=header)
                    cell.font = header_font
                    cell.alignment = header_alignment
                    cell.border = border
                row += 1
                
                # Transaction data
                for transaction in vendor_data['transactions']:
                    ws.cell(row=row, column=1, value=transaction['date'])
                    ws.cell(row=row, column=2, value=transaction['description'])
                    ws.cell(row=row, column=3, value=transaction['reference'])
                    ws.cell(row=row, column=4, value=transaction['debit'])
                    ws.cell(row=row, column=5, value=transaction['credit'])
                    ws.cell(row=row, column=6, value=transaction['balance'])
                    ws.cell(row=row, column=7, value=transaction['type'])
                    row += 1
                
                # Vendor summary
                ws.cell(row=row, column=3, value='Total:')
                ws.cell(row=row, column=4, value=vendor_data['total_debit'])
                ws.cell(row=row, column=5, value=vendor_data['total_credit'])
                ws.cell(row=row, column=6, value=vendor_data['balance'])
                row += 2
            
            # Overall summary
            if total_summary:
                ws.merge_cells(f'A{row}:G{row}')
                ws[f'A{row}'] = 'Overall Summary'
                ws[f'A{row}'].font = header_font
                row += 1
                
                ws.cell(row=row, column=3, value='Grand Total:')
                ws.cell(row=row, column=4, value=total_summary['total_debit'])
                ws.cell(row=row, column=5, value=total_summary['total_credit'])
                ws.cell(row=row, column=6, value=total_summary['net_balance'])
            
            # Save to BytesIO
            output = BytesIO()
            wb.save(output)
            output.seek(0)
            
            # Create response
            response = HttpResponse(
                output.getvalue(),
                content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            )
            response['Content-Disposition'] = f'attachment; filename="vendor_ledger_report_{datetime.now().strftime("%Y%m%d_%H%M%S")}.xlsx"'
            
            return response
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def export_pdf(request):
    """Export vendor ledger report to PDF"""
    if request.method == 'POST':
        form = VendorLedgerFilterForm(request.POST)
        if form.is_valid():
            report_data, total_summary = generate_vendor_ledger_data(
                vendor=form.cleaned_data.get('vendor'),
                date_from=form.cleaned_data['date_from'],
                date_to=form.cleaned_data['date_to'],
                payment_status=form.cleaned_data.get('payment_status', 'all')
            )
            
            context = {
                'report_data': report_data,
                'total_summary': total_summary,
                'date_from': form.cleaned_data['date_from'],
                'date_to': form.cleaned_data['date_to'],
                'vendor': form.cleaned_data.get('vendor'),
                'payment_status': form.cleaned_data.get('payment_status', 'all'),
                'generated_at': timezone.now()
            }
            
            return render(request, 'vendor_ledger/report_pdf.html', context)
    
    return JsonResponse({'error': 'Invalid request'}, status=400)


@login_required
def ajax_quick_filter(request):
    """AJAX endpoint for quick date filters"""
    if request.method == 'POST':
        filter_type = request.POST.get('filter_type')
        today = timezone.now().date()
        
        if filter_type == 'today':
            date_from = date_to = today
        elif filter_type == 'yesterday':
            date_from = date_to = today - timedelta(days=1)
        elif filter_type == 'this_week':
            date_from = today - timedelta(days=today.weekday())
            date_to = today
        elif filter_type == 'last_week':
            date_from = today - timedelta(days=today.weekday() + 7)
            date_to = today - timedelta(days=today.weekday() + 1)
        elif filter_type == 'this_month':
            date_from = today.replace(day=1)
            date_to = today
        elif filter_type == 'last_month':
            if today.month == 1:
                date_from = today.replace(year=today.year-1, month=12, day=1)
                date_to = today.replace(day=1) - timedelta(days=1)
            else:
                date_from = today.replace(month=today.month-1, day=1)
                date_to = today.replace(day=1) - timedelta(days=1)
        else:
            return JsonResponse({'error': 'Invalid filter type'}, status=400)
        
        return JsonResponse({
            'date_from': date_from.strftime('%Y-%m-%d'),
            'date_to': date_to.strftime('%Y-%m-%d')
        })
    
    return JsonResponse({'error': 'Invalid request'}, status=400)