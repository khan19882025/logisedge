from django.shortcuts import render, redirect
from django.contrib.auth.decorators import login_required
from django.contrib.auth.mixins import LoginRequiredMixin, UserPassesTestMixin
from django.views.generic import ListView, DetailView, View
from django.http import HttpResponse, JsonResponse
from django.contrib import messages
from django.utils import timezone
from django.db.models import Q
from django.core.paginator import Paginator
import json
import csv
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
import io

from .models import ExportLog
from .forms import CustomerExportForm, ItemExportForm, TransactionExportForm
from customer.models import Customer
from items.models import Item
from invoice.models import Invoice  # Assuming this is the transaction model


class BulkExportDashboardView(LoginRequiredMixin, UserPassesTestMixin, View):
    """Dashboard view for bulk export functionality"""
    template_name = 'bulk_export/dashboard.html'
    
    def test_func(self):
        """Check if user has permission to access bulk export"""
        return self.request.user.has_perm('bulk_export.can_export_data')
    
    def get(self, request, *args, **kwargs):
        """Handle GET request"""
        context = self.get_context_data(**kwargs)
        return render(request, self.template_name, context)
    
    def get_context_data(self, **kwargs):
        context = {}
        context['recent_exports'] = ExportLog.objects.filter(
            user=self.request.user
        ).order_by('-created_at')[:10]
        context['export_stats'] = {
            'total_exports': ExportLog.objects.filter(user=self.request.user).count(),
            'customers_exports': ExportLog.objects.filter(
                user=self.request.user, export_type='customers'
            ).count(),
            'items_exports': ExportLog.objects.filter(
                user=self.request.user, export_type='items'
            ).count(),
            'transactions_exports': ExportLog.objects.filter(
                user=self.request.user, export_type='transactions'
            ).count(),
        }
        return context


@login_required
def customer_export_view(request):
    """View for customer export form and processing"""
    if not request.user.has_perm('bulk_export.can_export_data'):
        messages.error(request, "You don't have permission to export data.")
        return redirect('bulk_export:dashboard')
    
    if request.method == 'POST':
        form = CustomerExportForm(request.POST)
        if form.is_valid():
            return process_customer_export(request, form.cleaned_data)
    else:
        form = CustomerExportForm()
    
    return render(request, 'bulk_export/customer_export.html', {'form': form})


@login_required
def item_export_view(request):
    """View for item export form and processing"""
    if not request.user.has_perm('bulk_export.can_export_data'):
        messages.error(request, "You don't have permission to export data.")
        return redirect('bulk_export:dashboard')
    
    if request.method == 'POST':
        form = ItemExportForm(request.POST)
        if form.is_valid():
            return process_item_export(request, form.cleaned_data)
    else:
        form = ItemExportForm()
    
    return render(request, 'bulk_export/item_export.html', {'form': form})


@login_required
def transaction_export_view(request):
    """View for transaction export form and processing"""
    if not request.user.has_perm('bulk_export.can_export_data'):
        messages.error(request, "You don't have permission to export data.")
        return redirect('bulk_export:dashboard')
    
    if request.method == 'POST':
        form = TransactionExportForm(request.POST)
        if form.is_valid():
            return process_transaction_export(request, form.cleaned_data)
    else:
        form = TransactionExportForm()
    
    return render(request, 'bulk_export/transaction_export.html', {'form': form})


def process_customer_export(request, form_data):
    """Process customer export based on form data"""
    # Create export log entry
    export_log = ExportLog.objects.create(
        user=request.user,
        export_type='customers',
        export_format=form_data['export_format'],
        filename=f"customers_export_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.{form_data['export_format']}",
        filters_applied=form_data,
        status='processing'
    )
    
    try:
        # Build query based on filters
        queryset = Customer.objects.all()
        
        if form_data.get('customer_code'):
            queryset = queryset.filter(
                customer_code__icontains=form_data['customer_code']
            )
        
        if form_data.get('customer_name'):
            queryset = queryset.filter(
                customer_name__icontains=form_data['customer_name']
            )
        
        if form_data.get('region'):
            queryset = queryset.filter(
                region__icontains=form_data['region']
            )
        
        if form_data.get('account_status'):
            queryset = queryset.filter(
                account_status=form_data['account_status']
            )
        
        if form_data.get('registration_date_from'):
            queryset = queryset.filter(
                registration_date__gte=form_data['registration_date_from']
            )
        
        if form_data.get('registration_date_to'):
            queryset = queryset.filter(
                registration_date__lte=form_data['registration_date_to']
            )
        
        # Update export log with record count
        export_log.records_exported = queryset.count()
        export_log.save()
        
        # Generate export file
        if form_data['export_format'] == 'excel':
            response = generate_excel_export(
                queryset, 
                'customers', 
                export_log.filename
            )
        else:
            response = generate_csv_export(
                queryset, 
                'customers', 
                export_log.filename
            )
        
        # Update export log as completed
        export_log.status = 'completed'
        export_log.completed_at = timezone.now()
        export_log.save()
        
        return response
        
    except Exception as e:
        export_log.status = 'failed'
        export_log.error_message = str(e)
        export_log.save()
        messages.error(request, f"Export failed: {str(e)}")
        return redirect('bulk_export:customer_export')


def process_item_export(request, form_data):
    """Process item export based on form data"""
    # Create export log entry
    export_log = ExportLog.objects.create(
        user=request.user,
        export_type='items',
        export_format=form_data['export_format'],
        filename=f"items_export_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.{form_data['export_format']}",
        filters_applied=form_data,
        status='processing'
    )
    
    try:
        # Build query based on filters
        queryset = Item.objects.all()
        
        if form_data.get('item_code'):
            queryset = queryset.filter(
                item_code__icontains=form_data['item_code']
            )
        
        if form_data.get('item_name'):
            queryset = queryset.filter(
                item_name__icontains=form_data['item_name']
            )
        
        if form_data.get('category'):
            queryset = queryset.filter(
                category__icontains=form_data['category']
            )
        
        if form_data.get('brand'):
            queryset = queryset.filter(
                brand__icontains=form_data['brand']
            )
        
        if form_data.get('supplier'):
            queryset = queryset.filter(
                supplier__icontains=form_data['supplier']
            )
        
        if form_data.get('status'):
            queryset = queryset.filter(
                status=form_data['status']
            )
        
        # Update export log with record count
        export_log.records_exported = queryset.count()
        export_log.save()
        
        # Generate export file
        if form_data['export_format'] == 'excel':
            response = generate_excel_export(
                queryset, 
                'items', 
                export_log.filename
            )
        else:
            response = generate_csv_export(
                queryset, 
                'items', 
                export_log.filename
            )
        
        # Update export log as completed
        export_log.status = 'completed'
        export_log.completed_at = timezone.now()
        export_log.save()
        
        return response
        
    except Exception as e:
        export_log.status = 'failed'
        export_log.error_message = str(e)
        export_log.save()
        messages.error(request, f"Export failed: {str(e)}")
        return redirect('bulk_export:item_export')


def process_transaction_export(request, form_data):
    """Process transaction export based on form data"""
    # Create export log entry
    export_log = ExportLog.objects.create(
        user=request.user,
        export_type='transactions',
        export_format=form_data['export_format'],
        filename=f"transactions_export_{datetime.now().strftime('%Y-%m-%d_%H-%M')}.{form_data['export_format']}",
        filters_applied=form_data,
        status='processing'
    )
    
    try:
        # Build query based on filters
        queryset = Invoice.objects.all()  # Assuming Invoice is the transaction model
        
        if form_data.get('transaction_type'):
            queryset = queryset.filter(
                transaction_type=form_data['transaction_type']
            )
        
        if form_data.get('date_from'):
            queryset = queryset.filter(
                date__gte=form_data['date_from']
            )
        
        if form_data.get('date_to'):
            queryset = queryset.filter(
                date__lte=form_data['date_to']
            )
        
        if form_data.get('customer'):
            queryset = queryset.filter(
                customer=form_data['customer']
            )
        
        if form_data.get('item'):
            queryset = queryset.filter(
                items=form_data['item']
            )
        
        if form_data.get('payment_status'):
            queryset = queryset.filter(
                payment_status=form_data['payment_status']
            )
        
        if form_data.get('location'):
            queryset = queryset.filter(
                location__icontains=form_data['location']
            )
        
        # Update export log with record count
        export_log.records_exported = queryset.count()
        export_log.save()
        
        # Generate export file
        if form_data['export_format'] == 'excel':
            response = generate_excel_export(
                queryset, 
                'transactions', 
                export_log.filename
            )
        else:
            response = generate_csv_export(
                queryset, 
                'transactions', 
                export_log.filename
            )
        
        # Update export log as completed
        export_log.status = 'completed'
        export_log.completed_at = timezone.now()
        export_log.save()
        
        return response
        
    except Exception as e:
        export_log.status = 'failed'
        export_log.error_message = str(e)
        export_log.save()
        messages.error(request, f"Export failed: {str(e)}")
        return redirect('bulk_export:transaction_export')


def generate_excel_export(queryset, export_type, filename):
    """Generate Excel export file"""
    wb = openpyxl.Workbook()
    ws = wb.active
    
    # Define field mappings for human-readable headers
    field_mappings = get_field_mappings(export_type)
    
    # Write headers
    headers = list(field_mappings.values())
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = Font(bold=True)
        cell.fill = PatternFill(start_color="CCCCCC", end_color="CCCCCC", fill_type="solid")
        cell.alignment = Alignment(horizontal="center")
    
    # Write data
    for row, obj in enumerate(queryset, 2):
        for col, field_name in enumerate(field_mappings.keys(), 1):
            value = getattr(obj, field_name, '')
            if hasattr(value, 'strftime'):  # Handle date fields
                value = value.strftime('%Y-%m-%d')
            ws.cell(row=row, column=col, value=value)
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = get_column_letter(column[0].column)
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to response
    response = HttpResponse(
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
    )
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Save workbook to response
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    response.write(output.getvalue())
    
    return response


def generate_csv_export(queryset, export_type, filename):
    """Generate CSV export file"""
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = f'attachment; filename="{filename}"'
    
    # Define field mappings for human-readable headers
    field_mappings = get_field_mappings(export_type)
    
    writer = csv.writer(response)
    
    # Write headers
    writer.writerow(list(field_mappings.values()))
    
    # Write data
    for obj in queryset:
        row = []
        for field_name in field_mappings.keys():
            value = getattr(obj, field_name, '')
            if hasattr(value, 'strftime'):  # Handle date fields
                value = value.strftime('%Y-%m-%d')
            row.append(str(value) if value is not None else '')
        writer.writerow(row)
    
    return response


def get_field_mappings(export_type):
    """Get field mappings for different export types"""
    if export_type == 'customers':
        return {
            'customer_code': 'Customer Code',
            'customer_name': 'Customer Name',
            'email': 'Email',
            'phone': 'Phone',
            'address': 'Address',
            'region': 'Region',
            'account_status': 'Account Status',
            'registration_date': 'Registration Date',
            'credit_limit': 'Credit Limit',
            'payment_terms': 'Payment Terms',
        }
    elif export_type == 'items':
        return {
            'item_code': 'Item Code',
            'item_name': 'Item Name',
            'category': 'Category',
            'brand': 'Brand',
            'supplier': 'Supplier',
            'unit_price': 'Unit Price',
            'cost_price': 'Cost Price',
            'status': 'Status',
            'description': 'Description',
            'created_date': 'Created Date',
        }
    elif export_type == 'transactions':
        return {
            'invoice_number': 'Invoice Number',
            'transaction_type': 'Transaction Type',
            'date': 'Date',
            'customer': 'Customer',
            'total_amount': 'Total Amount',
            'payment_status': 'Payment Status',
            'location': 'Location',
            'items': 'Items',
            'created_by': 'Created By',
        }
    return {}


class ExportLogListView(LoginRequiredMixin, ListView):
    """View for listing export logs"""
    model = ExportLog
    template_name = 'bulk_export/export_logs.html'
    context_object_name = 'export_logs'
    paginate_by = 20
    
    def get_queryset(self):
        return ExportLog.objects.filter(user=self.request.user).order_by('-created_at')


class ExportLogDetailView(LoginRequiredMixin, DetailView):
    """View for showing export log details"""
    model = ExportLog
    template_name = 'bulk_export/export_log_detail.html'
    context_object_name = 'export_log'
    
    def get_queryset(self):
        return ExportLog.objects.filter(user=self.request.user)
