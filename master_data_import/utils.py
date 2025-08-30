import pandas as pd
import json
import re
import hashlib
from datetime import datetime
from django.utils import timezone
from django.db import transaction
from django.core.files.storage import default_storage
from django.core.files.base import ContentFile
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
import os
import tempfile

from .models import ImportJob, ImportDataError, ImportAuditLog


def get_client_ip(request):
    """Get client IP address from request"""
    x_forwarded_for = request.META.get('HTTP_X_FORWARDED_FOR')
    if x_forwarded_for:
        ip = x_forwarded_for.split(',')[0]
    else:
        ip = request.META.get('REMOTE_ADDR')
    return ip


def create_audit_log(import_job, action, message, request=None, details=None):
    """Helper function to create audit logs"""
    try:
        # Get request information if available
        ip_address = None
        user_agent = None
        
        if request:
            ip_address = get_client_ip(request)
            user_agent = request.META.get('HTTP_USER_AGENT', '')
        
        ImportAuditLog.objects.create(
            import_job=import_job,
            action=action,
            message=message,
            details=details or {},
            ip_address=ip_address,
            user_agent=user_agent
        )
    except Exception as e:
        # Log error but don't break the main operation
        print(f"Error creating audit log: {e}")


def validate_email(email):
    """Validate email format"""
    if not email:
        return False
    pattern = r'^[a-zA-Z0-9._%+-]+@[a-zA-Z0-9.-]+\.[a-zA-Z]{2,}$'
    return re.match(pattern, email) is not None


def validate_phone(phone):
    """Validate phone number format"""
    if not phone:
        return False
    # Remove all non-digit characters
    digits_only = re.sub(r'\D', '', phone)
    return len(digits_only) >= 10


def validate_date(date_str):
    """Validate date format"""
    if not date_str:
        return False
    try:
        datetime.strptime(date_str, '%Y-%m-%d')
        return True
    except ValueError:
        return False


def validate_number(value):
    """Validate number format"""
    if not value:
        return False
    try:
        float(value)
        return True
    except ValueError:
        return False


def validate_data(import_job, column_mapping, preview_data):
    """Validate data according to template rules"""
    validation_results = {
        'valid_rows': 0,
        'invalid_rows': 0,
        'errors': []
    }
    
    template = import_job.template
    validation_rules = template.validation_rules
    required_fields = template.required_fields
    
    for row_index, row_data in enumerate(preview_data, start=1):
        row_errors = []
        is_row_valid = True
        
        # Check required fields
        for field in required_fields:
            if field in column_mapping:
                mapped_column = column_mapping[field]
                if mapped_column in row_data:
                    value = row_data[mapped_column]
                    if not value or str(value).strip() == '':
                        row_errors.append(f"Required field '{field}' is empty")
                        is_row_valid = False
        
        # Apply validation rules
        for rule in validation_rules:
            field_name = rule.get('field')
            rule_type = rule.get('type')
            rule_config = rule.get('config', {})
            
            if field_name in column_mapping:
                mapped_column = column_mapping[field_name]
                if mapped_column in row_data:
                    value = row_data[mapped_column]
                    
                    if rule_type == 'email' and value:
                        if not validate_email(str(value)):
                            row_errors.append(f"Invalid email format for '{field_name}': {value}")
                            is_row_valid = False
                    
                    elif rule_type == 'phone' and value:
                        if not validate_phone(str(value)):
                            row_errors.append(f"Invalid phone format for '{field_name}': {value}")
                            is_row_valid = False
                    
                    elif rule_type == 'date' and value:
                        if not validate_date(str(value)):
                            row_errors.append(f"Invalid date format for '{field_name}': {value}")
                            is_row_valid = False
                    
                    elif rule_type == 'number' and value:
                        if not validate_number(str(value)):
                            row_errors.append(f"Invalid number format for '{field_name}': {value}")
                            is_row_valid = False
                    
                    elif rule_type == 'min_length' and value:
                        min_length = rule_config.get('min_length', 0)
                        if len(str(value)) < min_length:
                            row_errors.append(f"'{field_name}' must be at least {min_length} characters")
                            is_row_valid = False
                    
                    elif rule_type == 'max_length' and value:
                        max_length = rule_config.get('max_length', 255)
                        if len(str(value)) > max_length:
                            row_errors.append(f"'{field_name}' must be at most {max_length} characters")
                            is_row_valid = False
        
        if is_row_valid:
            validation_results['valid_rows'] += 1
        else:
            validation_results['invalid_rows'] += 1
            validation_results['errors'].extend([
                {
                    'row': row_index,
                    'error': error
                }
                for error in row_errors
            ])
    
    return validation_results


def process_import(import_job, column_mapping, skip_errors=True):
    """Process the actual import"""
    try:
        # Read the file
        import_file = import_job.import_file.file
        file_extension = os.path.splitext(import_job.file_name)[1].lower()
        
        if file_extension == '.csv':
            df = pd.read_csv(import_file, encoding='utf-8')
        elif file_extension in ['.xlsx', '.xls']:
            df = pd.read_excel(import_file)
        else:
            raise ValueError("Unsupported file format")
        
        successful_rows = 0
        failed_rows = 0
        
        # Process each row
        for index, row in df.iterrows():
            row_number = index + 2  # +2 because Excel/CSV is 1-indexed and has header
            
            try:
                # Map data according to column mapping
                mapped_data = {}
                for excel_column, db_field in column_mapping.items():
                    if excel_column in row:
                        mapped_data[db_field] = row[excel_column]
                
                # Import data based on template type
                if import_job.template.data_type == 'customers':
                    success = import_customer_data(mapped_data)
                elif import_job.template.data_type == 'vendors':
                    success = import_vendor_data(mapped_data)
                elif import_job.template.data_type == 'products':
                    success = import_product_data(mapped_data)
                elif import_job.template.data_type == 'chart_of_accounts':
                    success = import_chart_of_accounts_data(mapped_data)
                elif import_job.template.data_type == 'employees':
                    success = import_employee_data(mapped_data)
                elif import_job.template.data_type == 'inventory_items':
                    success = import_inventory_item_data(mapped_data)
                elif import_job.template.data_type == 'price_lists':
                    success = import_price_list_data(mapped_data)
                else:
                    success = False
                
                if success:
                    successful_rows += 1
                else:
                    failed_rows += 1
                    if not skip_errors:
                        raise Exception("Import failed for this row")
                
            except Exception as e:
                failed_rows += 1
                # Create error record
                ImportDataError.objects.create(
                    import_job=import_job,
                    row_number=row_number,
                    error_type='system',
                    error_message=str(e),
                    field_value=str(row.to_dict())
                )
                
                if not skip_errors:
                    raise e
            
            # Update progress
            import_job.processed_rows = index + 1
            import_job.successful_rows = successful_rows
            import_job.failed_rows = failed_rows
            import_job.save()
        
        # Mark job as completed
        import_job.status = 'completed'
        import_job.completed_at = timezone.now()
        import_job.save()
        
        # Create audit log
        create_audit_log(
            import_job=import_job,
            action='complete',
            message=f'Import completed. {successful_rows} successful, {failed_rows} failed.'
        )
        
    except Exception as e:
        # Mark job as failed
        import_job.status = 'failed'
        import_job.save()
        
        # Create audit log
        create_audit_log(
            import_job=import_job,
            action='error',
            message=f'Import failed: {str(e)}'
        )
        
        raise e


def import_customer_data(data):
    """Import customer data"""
    try:
        from customer.models import Customer
        
        customer = Customer.objects.create(
            name=data.get('name', ''),
            email=data.get('email', ''),
            phone=data.get('phone', ''),
            address=data.get('address', ''),
            # Add other fields as needed
        )
        return True
    except Exception as e:
        print(f"Error importing customer: {e}")
        return False


def import_vendor_data(data):
    """Import vendor data"""
    try:
        # Import vendor data based on your vendor model
        # This is a placeholder - implement based on your actual vendor model
        return True
    except Exception as e:
        print(f"Error importing vendor: {e}")
        return False


def import_product_data(data):
    """Import product data"""
    try:
        # Import product data based on your product model
        # This is a placeholder - implement based on your actual product model
        return True
    except Exception as e:
        print(f"Error importing product: {e}")
        return False


def import_chart_of_accounts_data(data):
    """Import chart of accounts data"""
    try:
        # Import chart of accounts data based on your model
        # This is a placeholder - implement based on your actual model
        return True
    except Exception as e:
        print(f"Error importing chart of accounts: {e}")
        return False


def import_employee_data(data):
    """Import employee data"""
    try:
        # Import employee data based on your employee model
        # This is a placeholder - implement based on your actual employee model
        return True
    except Exception as e:
        print(f"Error importing employee: {e}")
        return False


def import_inventory_item_data(data):
    """Import inventory item data"""
    try:
        # Import inventory item data based on your model
        # This is a placeholder - implement based on your actual model
        return True
    except Exception as e:
        print(f"Error importing inventory item: {e}")
        return False


def import_price_list_data(data):
    """Import price list data"""
    try:
        # Import price list data based on your model
        # This is a placeholder - implement based on your actual model
        return True
    except Exception as e:
        print(f"Error importing price list: {e}")
        return False


def generate_template_file(template):
    """Generate Excel template file with sample data"""
    wb = Workbook()
    ws = wb.active
    ws.title = "Import Template"
    
    # Style for headers
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    
    # Get column mappings
    column_mappings = template.column_mappings
    required_fields = template.required_fields
    
    # Write headers
    for col, field_name in enumerate(column_mappings.keys(), 1):
        cell = ws.cell(row=1, column=col, value=field_name)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        
        # Add required field indicator
        if field_name in required_fields:
            cell.value = f"{field_name} *"
    
    # Add sample data rows
    sample_data = get_sample_data(template.data_type)
    for row, data in enumerate(sample_data, 2):
        for col, field_name in enumerate(column_mappings.keys(), 1):
            ws.cell(row=row, column=col, value=data.get(field_name, ''))
    
    # Add instructions
    instruction_row = len(sample_data) + 3
    ws.cell(row=instruction_row, column=1, value="Instructions:")
    ws.cell(row=instruction_row + 1, column=1, value="1. Fields marked with * are required")
    ws.cell(row=instruction_row + 2, column=1, value="2. Replace sample data with your actual data")
    ws.cell(row=instruction_row + 3, column=1, value="3. Do not modify the header row")
    ws.cell(row=instruction_row + 4, column=1, value="4. Save as CSV or Excel format")
    
    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 50)
        ws.column_dimensions[column_letter].width = adjusted_width
    
    # Save to bytes
    from io import BytesIO
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return output.getvalue()


def get_sample_data(data_type):
    """Get sample data for different data types"""
    if data_type == 'customers':
        return [
            {
                'name': 'Sample Customer 1',
                'email': 'customer1@example.com',
                'phone': '+971501234567',
                'address': 'Dubai, UAE',
                'contact_person': 'John Doe',
                'credit_limit': '10000',
                'payment_terms': '30 days'
            },
            {
                'name': 'Sample Customer 2',
                'email': 'customer2@example.com',
                'phone': '+971507654321',
                'address': 'Abu Dhabi, UAE',
                'contact_person': 'Jane Smith',
                'credit_limit': '5000',
                'payment_terms': '15 days'
            }
        ]
    elif data_type == 'vendors':
        return [
            {
                'name': 'Sample Vendor 1',
                'email': 'vendor1@example.com',
                'phone': '+971501234567',
                'address': 'Dubai, UAE',
                'contact_person': 'Vendor Contact',
                'payment_terms': '30 days'
            }
        ]
    elif data_type == 'products':
        return [
            {
                'name': 'Sample Product 1',
                'code': 'PROD001',
                'description': 'Sample product description',
                'category': 'Electronics',
                'unit_price': '100.00',
                'cost_price': '80.00'
            }
        ]
    elif data_type == 'chart_of_accounts':
        return [
            {
                'account_code': '1000',
                'account_name': 'Cash',
                'account_type': 'Asset',
                'description': 'Cash on hand and in bank'
            }
        ]
    elif data_type == 'employees':
        return [
            {
                'first_name': 'John',
                'last_name': 'Doe',
                'email': 'john.doe@company.com',
                'phone': '+971501234567',
                'department': 'IT',
                'position': 'Developer',
                'hire_date': '2023-01-15'
            }
        ]
    elif data_type == 'inventory_items':
        return [
            {
                'name': 'Sample Item 1',
                'code': 'INV001',
                'description': 'Sample inventory item',
                'category': 'Raw Materials',
                'unit_cost': '50.00',
                'reorder_level': '10'
            }
        ]
    elif data_type == 'price_lists':
        return [
            {
                'name': 'Standard Price List',
                'description': 'Standard pricing for all customers',
                'currency': 'AED',
                'effective_date': '2023-01-01'
            }
        ]
    else:
        return []
